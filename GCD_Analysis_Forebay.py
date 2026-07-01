"""
================================================================================
GCD Forebay Water Temperature — Full Analysis & Figures (June–July–August)
================================================================================
Grand Coulee Dam (GCD) Forebay — Historical Trend Analysis & Climate Projections
Primary Station of Interest: FDRW (Grand Coulee Dam Forebay)
Supporting Stations: GCGW → CHJ → CHQW (provided for comparison and context)

--------------------------------------------------------------------------------
DATA SOURCES
--------------------------------------------------------------------------------
WATER TEMPERATURE DATA:
  Source  : DART (Data Access in Real Time)
            Columbia Basin Research, University of Washington
  URL     : http://www.cbr.washington.edu/dart/query/river_graph_wmq
  Stations: FDRW  — Grand Coulee Dam Forebay  ← Primary station
            GCGW  — Grand Coulee Dam Tailrace
            CHJ   — Chief Joseph Dam Forebay
            CHQW  — Chief Joseph Dam Tailrace (comparison)
  Format  : Hourly water temperature (°C and °F)
  Period  : FDRW: 2000–2025 (primary analysis)
            CHQW: 1997–2025 (comparison)
            Others (CHJ, GCGW): 2000–2025

AIR TEMPERATURE DATA:
  Source  : NOAA National Centers for Environmental Information (NCEI)
            Global Historical Climatology Network Daily (GHCND)
  Station : DOUGLAS WASHINGTON, WA US
  ID      : GHCND:USR0000WDOU
  Lat/Lon : 47.619124, -119.899292
  Period  : 1990-07-30 to 2026-02-21  (coverage: ~88%)
  Used    : 1995–2025 subset, daily TMAX/TMIN averaged to daily mean (°C)

--------------------------------------------------------------------------------
METHODS
--------------------------------------------------------------------------------
  1. Mann-Kendall trend test
     Reference: Hirsch et al. (1991); Helsel & Hirsch (2002)
     Applied to annual June/July/August mean water temperatures, 2000–2025

  2. Theil-Sen slope estimator with 95% confidence interval
     Reference: Helsel & Hirsch (2002)
     Provides median trend slope robust to outliers and non-normality

  3. Mohseni et al. (1998) nonlinear logistic air-water temperature regression
        Tw = μ + (α − μ) / (1 + exp(γ(β − Ta)))
     Reference: Mohseni et al. (1998); Mantua et al. (2010); WACCIA Ch. 6
     Fit to all weekly air-water temperature pairs (2000–2025)
     Goodness-of-fit assessed via Nash-Sutcliffe Efficiency (NSE)
     NOTE: All-weeks model used for projections. JJA-only model may show
     poorer fit due to forebay thermal inertia and reservoir operations.
     Delta method (applying modeled changes to observed baselines) reduces
     bias from absolute prediction errors.

  4. Future climate projections via CMIP5 air temperature deltas
     Source:  MACAv2-METDATA (Abatzoglou & Brown 2012), 20 CMIP5 GCMs
              accessed via Climate Toolbox (climatetoolbox.org)
     Grid cell: 47.619 N, 119.899 W (GCD area, Douglas County, WA)
     Variable: JJA mean air temperature = (daily Tmax + daily Tmin) / 2
     Baseline: 1970-1999 model historical period
     Scenarios: RCP 4.5 (moderate emissions): +2.36 C by 2040s, +3.46 C by 2080s
                RCP 8.5 (high emissions):     +2.97 C by 2040s, +6.11 C by 2080s

  5. Warm Day Scenario Analysis (10% Exceedance)
     90th percentile (10% exceedance) of daily air temperatures for June,
     July, and August. Mohseni delta method applied to observed 90th
     percentile water temps.

--------------------------------------------------------------------------------
OUTPUTS
--------------------------------------------------------------------------------
  Excel Workbook:
    GCD_Temperature_Results.xlsx           — Multi-sheet workbook

  Figures:
    Fig1_Annual_Summer_Trends.png          — annual Jun/Jul/Aug means
    Fig2_Seasonal_Climatology.png          — mean annual temperature cycle
    Fig3_FDRW_Trend_Detail.png             — FDRW detailed trend + air overlay
    Fig4_JJA_Daily_By_Year.png             — daily JJA temps, all years overlaid
    Fig5_Mohseni_Regression.png            — Mohseni air-water regression
    Fig6_Mohseni_Observed_vs_Predicted.png — observed vs predicted weekly
    Fig7_JJA_Obs_vs_Pred_TimeSeries.png    — JJA annual observed vs predicted
    Fig8_Mohseni_All_Years_Overlaid.png    — all years overlaid on Mohseni curve
    Fig9_Projected_Temps_Bar.png           — projected temps bar chart
    Fig10_Projection_Points_Mohseni.png    — projection points on Mohseni curve
    Fig11_Period_Comparison_Boxplots.png   — early vs. recent period comparison
    Fig12_Monthly_Projections_Bar.png      — monthly projections (Jun/Jul/Aug)
    Fig13_Monthly_Projection_Points.png    — monthly projections on curve
    Fig14_Observed_Weekly_MinMaxMean.png   — observed weekly temperature ranges
    Fig15_Projected_Weekly_MinMaxMean.png  — projected weekly temperature ranges
    Fig16_Average_vs_WarmDay_Comparison.png — average vs warm day comparison

--------------------------------------------------------------------------------
REFERENCES
--------------------------------------------------------------------------------
  Helsel, D.R. and Hirsch, R.M. (2002). Statistical Methods in Water Resources.
  Hirsch, R.M., Alexander, R.B., and Smith, R.A. (1991). WRR, 27(5), 803–813.
  Mantua, N., Tohver, I., and Hamlet, A. (2010). Climatic Change, 102, 187–223.
  Abatzoglou, J.T. and Brown, T.J. (2012). Int. J. Climatology, 32, 772-780.
  Mohseni, O., Stefan, H.G., and Erickson, T.R. (1998). WRR, 34(10), 2685–2692.
  O'Connor, J.E. et al. (2021). USACE, Institute for Water Resources.
  Vermeyen, T.B. (2000). USBR PAP-0854.
  Washington Climate Change Impacts Assessment (WACCIA), Chapter 6 — Salmon.
================================================================================
"""

import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
import matplotlib.cm as cm
import matplotlib.colors as mcolors
from matplotlib.lines import Line2D
from matplotlib.ticker import MultipleLocator
from scipy import stats, optimize
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_DIR = r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\results'

STATION_FILES = {
    'CHJ':  r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\CHJ\CHJ_Hourly_Data_1995_2025.xlsx',
    'GCGW': r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\GCGW\GCGW_Hourly_Data_1995_2025.xlsx',
    'FDRW': r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\FDRW\FDRW_Hourly_Data_1995_2025.xlsx',
    'CHQW': r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\CHQW\CHQW_Hourly_Data_1995_2025.xlsx',
}

AIR_TEMP_FILE = r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\DailyClimatologicalData.csv'

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── KEY CHANGE: June, July, August ──────────────────────────────────────────
SUMMER_MONTHS  = [6, 7, 8]
SUMMER_LABEL   = 'Jun-Jul-Aug'
SUMMER_LABEL_LONG = 'June–July–August'
MONTH_TUPLES   = [(6, 'June'), (7, 'July'), (8, 'August')]
MONTH_COLORS   = {'June': '#27ae60', 'July': '#e74c3c', 'August': '#2980b9'}

STATION_ORDER  = ['FDRW', 'GCGW', 'CHJ', 'CHQW']
STATION_LABELS = {
    'FDRW': 'Grand Coulee Forebay (FDRW)',
    'GCGW': 'Grand Coulee Tailrace (GCGW)',
    'CHJ':  'Chief Joseph Forebay (CHJ)',
    'CHQW': 'Chief Joseph Tailrace (CHQW)',
}

# CMIP5 Climate Projection Delta Computation
MACA_TMAX_FILE = r"C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\FutureTimeSeriesMax.csv"
MACA_TMIN_FILE = r"C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\FutureTimeSeriesMin.csv"


def compute_cmip5_deltas(tmax_path, tmin_path, header_rows=51):
    """Compute JJA mean air temperature deltas from MACAv2-METDATA CMIP5 ensemble."""
    import pandas as pd

    dfx = pd.read_csv(tmax_path, skiprows=header_rows, header=None, on_bad_lines='skip')
    dfn = pd.read_csv(tmin_path, skiprows=header_rows, header=None, on_bad_lines='skip')

    years = dfx.iloc[:, 0].astype(int).values

    tmax_45 = dfx.iloc[:, 2:22].astype(float).values
    tmax_85 = dfx.iloc[:, 22:42].astype(float).values
    tmin_45 = dfn.iloc[:, 2:22].astype(float).values
    tmin_85 = dfn.iloc[:, 22:42].astype(float).values

    tmean_45 = (tmax_45 + tmin_45) / 2
    tmean_85 = (tmax_85 + tmin_85) / 2

    ens_45 = np.nanmean(tmean_45, axis=1)
    ens_85 = np.nanmean(tmean_85, axis=1)

    hist = (years >= 1970) & (years <= 1999)
    hist_mean = ens_45[hist].mean()

    windows = {
        '2040s': (years >= 2030) & (years <= 2059),
        '2080s': (years >= 2070) & (years <= 2099),
    }

    deltas = {}
    print(f"\n{'='*72}")
    print(f"  CMIP5 JJA Air Temperature Deltas — 20-Model MACAv2-METDATA Ensemble")
    print(f"  Grid cell: 47.619 N, 119.899 W  |  Baseline: 1970-1999 = {hist_mean:.2f} C")
    print(f"{'='*72}")
    print(f"  {'Period':<22} {'RCP 4.5':>10} {'RCP 8.5':>10}")
    print(f"  {'-'*44}")
    for period, mask in windows.items():
        d45 = ens_45[mask].mean() - hist_mean
        d85 = ens_85[mask].mean() - hist_mean
        deltas[f'{period} RCP 4.5'] = round(d45, 2)
        deltas[f'{period} RCP 8.5'] = round(d85, 2)
        print(f"  {period:<22} {d45:>+10.2f} {d85:>+10.2f}")
    print(f"{'='*72}\n")

    return deltas


_maca_files_exist = os.path.isfile(MACA_TMAX_FILE) and os.path.isfile(MACA_TMIN_FILE)

if _maca_files_exist:
    _deltas = compute_cmip5_deltas(MACA_TMAX_FILE, MACA_TMIN_FILE)
else:
    _deltas = {
        '2040s RCP 4.5': 2.36,
        '2040s RCP 8.5': 2.97,
        '2080s RCP 4.5': 3.46,
        '2080s RCP 8.5': 6.11,
    }
    print("  [INFO] MACA CSV files not found — using pre-computed CMIP5 deltas.")

ANALYSIS_START_YEAR = 1995
ANALYSIS_END_YEAR   = 2025

CLIMATE_SCENARIOS = [
    ('Baseline',              0.00,                     '#2980b9'),
    ('2040s RCP 4.5',        _deltas['2040s RCP 4.5'],  '#f57c00'),
    ('2040s RCP 8.5',        _deltas['2040s RCP 8.5'],  '#d32f2f'),
    ('2080s RCP 4.5',        _deltas['2080s RCP 4.5'],  '#e65100'),
    ('2080s RCP 8.5',        _deltas['2080s RCP 8.5'],  '#b71c1c'),
]

C = {
    'FDRW':  '#c0392b',
    'GCGW':  '#2d9e6b',
    'CHJ':   '#1a6faf',
    'CHQW':  '#e07b2a',
    'air':   '#636e72',
    'trend': '#2c3e50',
}

plt.rcParams.update({
    'font.family':       'DejaVu Sans',
    'font.size':         12,
    'axes.spines.top':   False,
    'axes.spines.right': False,
    'axes.grid':         True,
    'grid.alpha':        0.2,
    'grid.linestyle':    '--',
    'axes.labelsize':    13,
    'axes.titlesize':    13,
    'axes.titleweight':  'bold',
    'legend.framealpha': 0.93,
    'legend.edgecolor':  '#cccccc',
    'legend.fontsize':   11,
    'xtick.labelsize':   11,
    'ytick.labelsize':   11,
    'figure.dpi':        150,
})

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1 — STATISTICAL HELPER FUNCTIONS
# ═════════════════════════════════════════════════════════════════════════════

def mann_kendall(y):
    n = len(y)
    if n < 4:
        return np.nan, np.nan, 0, 0.0, 'n/a'
    S = int(sum(np.sign(y[j] - y[i]) for i in range(n - 1) for j in range(i + 1, n)))
    var_S = n * (n - 1) * (2 * n + 5) / 18.0
    if   S > 0: Z = (S - 1) / np.sqrt(var_S)
    elif S < 0: Z = (S + 1) / np.sqrt(var_S)
    else:       Z = 0.0
    p_value = 2 * (1 - stats.norm.cdf(abs(Z)))
    tau     = S / (n * (n - 1) / 2.0)
    trend   = 'increasing' if S > 0 else ('decreasing' if S < 0 else 'no trend')
    return tau, p_value, S, Z, trend


def theil_sen(x, y):
    n = len(x)
    slopes = [(y[j] - y[i]) / (x[j] - x[i])
              for i in range(n) for j in range(i + 1, n) if x[j] != x[i]]
    slopes = np.array(slopes)
    slope  = np.median(slopes)
    intercept = np.median(y) - slope * np.median(x)
    M       = len(slopes)
    C_alpha = 1.96 * np.sqrt(n * (n - 1) * (2 * n + 5) / 18.0)
    m1 = max(0, int((M - C_alpha) / 2))
    m2 = min(M - 1, int((M + C_alpha) / 2 + 1))
    ss  = np.sort(slopes)
    return slope, intercept, ss[m1], ss[m2]


def mohseni(Ta, mu, alpha, gamma, beta):
    return mu + (alpha - mu) / (1 + np.exp(gamma * (beta - Ta)))


def fit_mohseni(Ta, Tw):
    Ta, Tw = np.asarray(Ta, float), np.asarray(Tw, float)
    p0  = [max(0.0, Tw.min() - 0.5), Tw.max() + 0.5, 0.2, float(np.median(Ta))]
    bds = ([0.0, p0[0] + 0.05, 0.01, -20.0],
           [Tw.max() + 2, 40.0, 2.0, 40.0])
    try:
        popt, _ = optimize.curve_fit(mohseni, Ta, Tw, p0=p0, bounds=bds, maxfev=12000)
    except Exception:
        print("  [WARNING] Mohseni fit did not converge; using linear fallback.")
        m, b, *_ = stats.linregress(Ta, Tw)
        popt = [max(0.0, b), b + m * 40, 0.15, 15.0]
    Tw_pred = mohseni(Ta, *popt)
    NSE = 1 - np.sum((Tw - Tw_pred) ** 2) / np.sum((Tw - Tw.mean()) ** 2)
    return popt, float(NSE)


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1B — QA/QC SAFEGUARDS
# ═════════════════════════════════════════════════════════════════════════════

class TemperatureQAQC:
    """Quality Assurance / Quality Control for water and air temperature data."""

    WATER_BOUNDS = {
        1: (0.0, 10),    2: (0.0, 10),    3: (0.5, 12),
        4: (2, 14),      5: (5, 16),      6: (8, 22),
        7: (10, 25),     8: (10, 25),     9: (8, 24),
        10: (4, 20),     11: (2, 14),     12: (0.5, 10),
    }
    AIR_BOUNDS = {
        1: (-20, 10),    2: (-20, 12),    3: (-10, 18),
        4: (-5, 25),     5: (0, 32),      6: (5, 38),
        7: (8, 42),      8: (8, 42),      9: (2, 35),
        10: (-5, 28),    11: (-12, 18),   12: (-20, 12),
    }

    @staticmethod
    def check_physical_bounds(df, temp_col, bounds_dict, source='water'):
        lo = df['month'].map({m: b[0] for m, b in bounds_dict.items()})
        hi = df['month'].map({m: b[1] for m, b in bounds_dict.items()})
        t = df[temp_col]
        # Rows whose month is not in bounds_dict (lo/hi NaN) are skipped.
        bad = lo.notna() & ~((lo <= t) & (t <= hi))
        if not bad.any():
            return pd.DataFrame(columns=['idx', 'issue']), 0
        idx = df.index[bad]
        flags = pd.DataFrame({
            'idx': idx,
            'issue': [f'{temp_col}={v:.2f}°C outside bounds' for v in t[bad]],
        })
        return flags, len(flags)

    @staticmethod
    def detect_outliers_iqr(series, iqr_multiplier=3.0):
        Q1, Q3 = series.quantile(0.25), series.quantile(0.75)
        IQR = Q3 - Q1
        lo = Q1 - iqr_multiplier * IQR
        hi = Q3 + iqr_multiplier * IQR
        bad = series[(series < lo) | (series > hi)].index.tolist()
        return bad, len(bad), (lo, hi)

    @staticmethod
    def detect_constant_spans(df, temp_col, min_span_hours=24, tolerance=0.05):
        # Greedy runs where each value stays within tolerance of the run's start.
        vals = df.sort_values('datetime')[temp_col].to_numpy()
        n = len(vals)
        flags = []
        i = 0
        while i < n:
            j = i
            start = vals[i]
            while j < n and abs(vals[j] - start) < tolerance:
                j += 1
            if j - i >= min_span_hours:
                flags.append({'duration_hours': j - i, 'note': 'Possible sensor failure'})
            i = j
        return pd.DataFrame(flags) if flags else pd.DataFrame(), len(flags)

    @staticmethod
    def detect_impossible_jumps(df, temp_col, max_hourly_change=3.0):
        d = df.sort_values('datetime')
        dt = d['datetime'].diff().dt.total_seconds() / 3600
        dtemp = d[temp_col].diff().abs()
        valid = d[temp_col].notna() & d[temp_col].shift().notna()
        bad = valid & (dt > 0) & (dtemp / dt > max_hourly_change)
        if not bad.any():
            return pd.DataFrame(), 0
        flags = pd.DataFrame({'idx': d.index[bad], 'dtemp': dtemp[bad], 'dt_hr': dt[bad]})
        return flags, len(flags)

    @staticmethod
    def detect_data_gaps(df, freq='D'):
        if 'datetime' in df.columns:
            time_col = 'datetime'
        elif 'date' in df.columns:
            time_col = 'date'
        else:
            raise KeyError("detect_data_gaps: expected 'datetime' or 'date' column")

        df_sorted = df.sort_values(time_col)
        expected_freq = pd.to_timedelta(f'1{freq}')
        t = df_sorted[time_col]
        gap = t.diff()
        bad = gap > expected_freq
        if not bad.any():
            return pd.DataFrame(), 0
        dur_days = gap[bad].dt.total_seconds() / 86400.0
        gaps = pd.DataFrame({
            'gap_start': t.shift()[bad],
            'gap_end': t[bad],
            'duration_days': dur_days,
            'issue': [f'Gap: {d:.1f} days' for d in dur_days],
        })
        return gaps, len(gaps)

    @staticmethod
    def qaqc_report(source, report_dict):
        print(f"  QA/QC Report: {source}")
        print("  " + "─" * 70)
        for check_name, (count, _) in report_dict.items():
            status = "[!] FLAGGED" if count > 0 else "[OK] PASS"
            print(f"    {check_name:<40} {count:>8} records  {status}")
        print("  " + "─" * 70)


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 2 — DATA LOADING
# ═════════════════════════════════════════════════════════════════════════════

def load_station(name, path):
    # Reading the hourly .xlsx via openpyxl is the slowest step. Cache the
    # cleaned/QA'd result and reuse it while the source file is unchanged.
    cache = path + '.qaqc.pkl'
    if os.path.isfile(cache) and os.path.getmtime(cache) >= os.path.getmtime(path):
        print(f"  Loading {name} (cached) …")
        df = pd.read_pickle(cache)
        print(f"    → {len(df):,} hourly records (cached)  |  {df.year.min()}–{df.year.max()}")
        return df

    print(f"  Loading {name} …")
    xl = pd.ExcelFile(path)
    df = pd.concat([xl.parse(s) for s in xl.sheet_names], ignore_index=True)
    df['datetime'] = pd.to_datetime(df['Pacific Timestamp'], errors='coerce')
    df = df.dropna(subset=['datetime', 'Temperature (C)']).copy()
    df.rename(columns={
        'Temperature (C)': 'wtc',
        'Temperature (F)': 'wtf',
        'Outflow (kcfs)':  'flow_kcfs'
    }, inplace=True)
    df['year']  = df['datetime'].dt.year
    df['month'] = df['datetime'].dt.month
    df['doy']   = df['datetime'].dt.dayofyear
    df['week']  = df['datetime'].dt.isocalendar().week.astype(int)
    df['station'] = name
    df = df.sort_values('datetime').reset_index(drop=True)

    qaqc = TemperatureQAQC()
    qaqc_report = {}

    bounds_flags, bounds_count = qaqc.check_physical_bounds(df, 'wtc', qaqc.WATER_BOUNDS, source='water')
    qaqc_report['Physical Bounds'] = (bounds_count, bounds_flags)
    if bounds_count > 0:
        print(f"    [!] {bounds_count} records flagged: outside physical temperature bounds")
        df = df.drop(bounds_flags['idx'].values).reset_index(drop=True)

    outlier_idx, outlier_count, _ = qaqc.detect_outliers_iqr(df['wtc'], iqr_multiplier=3.0)
    qaqc_report['IQR Outliers (3.0×IQR)'] = (outlier_count, None)
    if outlier_count > 0:
        print(f"    [!] {outlier_count} records flagged as statistical outliers (>3×IQR)")
        df = df.drop(outlier_idx).reset_index(drop=True)

    const_flags, const_count = qaqc.detect_constant_spans(df, 'wtc', min_span_hours=24, tolerance=0.05)
    qaqc_report['Constant Spans (24h+)'] = (const_count, const_flags)
    if const_count > 0:
        print(f"    [!] {const_count} long constant-value spans detected (possible sensor issues)")

    jump_flags, jump_count = qaqc.detect_impossible_jumps(df, 'wtc', max_hourly_change=3.0)
    qaqc_report['Impossible Jumps (>3°C/hr)'] = (jump_count, jump_flags)
    if jump_count > 0:
        print(f"    [!] {jump_count} impossible jumps detected (>3°C/hour)")
        df = df.drop(jump_flags['idx'].values).reset_index(drop=True)

    gap_flags, gap_count = qaqc.detect_data_gaps(df, freq='D')
    qaqc_report['Data Gaps'] = (gap_count, gap_flags)
    if gap_count > 0:
        print(f"    [!] {gap_count} data gap(s) detected")

    qaqc.qaqc_report(name, qaqc_report)

    df = df.sort_values('datetime').reset_index(drop=True)
    print(f"    → {len(df):,} hourly records after QA/QC  |  {df.year.min()}–{df.year.max()}")
    try:
        df.to_pickle(cache)
    except Exception as e:
        print(f"    [!] cache write skipped: {e}")
    return df


def load_air(path):
    print(f"  Loading air temperature …")
    # USBR AgriMet ABEI station (Grand Coulee Dam). Rows 1-19 are header/metadata,
    # data headers on row 20. Missing values coded as 'NO RECORD'. Values in °F.
    df = pd.read_csv(path, skiprows=19)
    df.columns = df.columns.str.strip()
    df.rename(columns={'DATE': 'date', 'ABEI MX': 'tmax_f',
                       'ABEI MN': 'tmin_f', 'ABEI MM': 'tavg_f'}, inplace=True)
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    # 'NO RECORD' (with trailing padding) → NaN; to_numeric handles both padding and text.
    for c in ['tmax_f', 'tmin_f', 'tavg_f']:
        df[c] = pd.to_numeric(df[c], errors='coerce')
    df = df.dropna(subset=['date', 'tmax_f', 'tmin_f']).reset_index(drop=True)
    df['tmax_c']  = (df['tmax_f'] - 32) * 5 / 9
    df['tmin_c']  = (df['tmin_f'] - 32) * 5 / 9
    df['tavg_f_raw'] = df['tavg_f'].copy()
    df['tavg_c']  = ((df['tmax_c'] + df['tmin_c']) / 2)
    df['year']    = df['date'].dt.year
    df['month']   = df['date'].dt.month
    df['doy']     = df['date'].dt.dayofyear
    df['week']    = df['date'].dt.isocalendar().week.astype(int)

    qaqc = TemperatureQAQC()
    qaqc_report = {}

    bounds_flags, bounds_count = qaqc.check_physical_bounds(df, 'tavg_c', qaqc.AIR_BOUNDS, source='air')
    qaqc_report['Physical Bounds'] = (bounds_count, bounds_flags)
    if bounds_count > 0:
        print(f"    [!] {bounds_count} records flagged: outside physical temperature bounds")
        df = df.drop(bounds_flags['idx'].values).reset_index(drop=True)

    outlier_idx, outlier_count, _ = qaqc.detect_outliers_iqr(df['tavg_c'], iqr_multiplier=3.0)
    qaqc_report['IQR Outliers (3.0×IQR)'] = (outlier_count, None)
    if outlier_count > 0:
        print(f"    [!] {outlier_count} records flagged as statistical outliers (>3×IQR)")
        df = df.drop(outlier_idx).reset_index(drop=True)

    gap_flags, gap_count = qaqc.detect_data_gaps(df, freq='D')
    qaqc_report['Data Gaps'] = (gap_count, gap_flags)
    if gap_count > 0:
        print(f"    [!] {gap_count} data gap(s) detected")

    qaqc.qaqc_report('Air Temperature (Douglas, WA)', qaqc_report)

    df = df.sort_values('date').reset_index(drop=True)
    print(f"    → {len(df):,} daily records after QA/QC  |  {df.year.min()}–{df.year.max()}")
    return df


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 3 — ANALYSIS
# ═════════════════════════════════════════════════════════════════════════════

def run_analysis(stations, air):
    print("\n" + "=" * 70)
    print("STATISTICAL ANALYSIS")
    print("=" * 70)

    def ann_summer(df):
        return df[df['month'].isin(SUMMER_MONTHS)].groupby('year')['wtc'].mean()

    def ann_month(df, month):
        return df[df['month'] == month].groupby('year')['wtc'].mean()

    fdrw = stations['FDRW']
    fdrw_wk = (fdrw.groupby(['year', 'week'])['wtc'].mean().reset_index()
               .assign(yw=lambda d: d['year'].astype(str) + '_' + d['week'].astype(str).str.zfill(2)))
    air_wk  = (air.groupby(['year', 'week'])['tavg_c'].mean().reset_index()
               .assign(yw=lambda d: d['year'].astype(str) + '_' + d['week'].astype(str).str.zfill(2)))
    merged  = fdrw_wk.merge(air_wk[['yw', 'tavg_c']], on='yw').dropna()

    # Mohseni fits — JJA weeks 22–35 covers Jun-Jul-Aug
    popt_all, NSE_all = fit_mohseni(merged['tavg_c'].values, merged['wtc'].values)
    jja_mask          = merged['week'].between(22, 35)
    popt_jja, NSE_jja = fit_mohseni(merged.loc[jja_mask, 'tavg_c'].values,
                                    merged.loc[jja_mask, 'wtc'].values)

    print(f"\n  Mohseni Model Parameters — FDRW (All Weeks)")
    print(f"    μ = {popt_all[0]:.3f} °C,  α = {popt_all[1]:.3f} °C,"
          f"  γ = {popt_all[2]:.4f},  β = {popt_all[3]:.3f} °C")
    print(f"    NSE = {NSE_all:.4f}")
    print(f"\n  Mohseni Model Parameters — FDRW (Jun–Aug Only)")
    print(f"    μ = {popt_jja[0]:.3f} °C,  α = {popt_jja[1]:.3f} °C,"
          f"  γ = {popt_jja[2]:.4f},  β = {popt_jja[3]:.3f} °C")
    print(f"    NSE = {NSE_jja:.4f}")

    if NSE_jja < 0.3:
        print(f"\n  ⚠ MODEL FIT QUALITY NOTE:")
        print(f"    The Jun-Aug only model shows poor fit (NSE = {NSE_jja:.3f}).")
        print(f"    PROJECTIONS USE: All-weeks model (NSE = {NSE_all:.3f}) for stability.")

    # Combined JJA trend — all stations
    print(f"\n  {'Station':<34} {'n':>4} {'Mean JJA °C':>11} {'Slope °C/yr':>13}"
          f" {'MK τ':>8} {'p':>8} {'Sig':>5}")
    print("  " + "-" * 90)

    summary_rows = []
    for stn in STATION_ORDER:
        ann  = ann_summer(stations[stn])
        x, y = ann.index.values.astype(float), ann.values
        tau, p, S, Z, direction = mann_kendall(y)
        sl, ic, ci_lo, ci_hi   = theil_sen(x, y)
        sig = '★' if p < 0.05 else ''
        print(f"  {STATION_LABELS[stn]:<34} {len(y):>4} {y.mean():>11.2f} {sl:>+13.4f}"
              f" {tau:>8.4f} {p:>8.4f} {sig:>5}")
        summary_rows.append({
            'Station': STATION_LABELS[stn], 'Code': stn,
            'n_years': len(y),
            'Mean_JJA_C': round(y.mean(), 3),
            'Mean_JJA_F': round(y.mean() * 9/5 + 32, 3),
            'TheilSen_slope_C_yr': round(sl, 5),
            'TheilSen_CI_low':  round(ci_lo, 5),
            'TheilSen_CI_high': round(ci_hi, 5),
            'MK_tau': round(tau, 4),
            'MK_S': S, 'MK_Z': round(Z, 4),
            'MK_p': round(p, 4),
            'Trend': direction,
            'Significant_p05': 'Yes' if p < 0.05 else 'No',
        })
    summary_df = pd.DataFrame(summary_rows)

    # Month-specific trends — all stations
    print(f"\n  Month-Specific Trends:")
    print(f"  {'Station':<34} {'Month':>6} {'n':>4} {'Mean °C':>9} {'Slope °C/yr':>13}"
          f" {'MK τ':>8} {'p':>8} {'Sig':>5}")
    print("  " + "-" * 95)

    monthly_summary_rows = []
    for stn in STATION_ORDER:
        for month, month_name in MONTH_TUPLES:
            ann = ann_month(stations[stn], month)
            x, y = ann.index.values.astype(float), ann.values
            if len(y) < 4:
                continue
            tau, p, S, Z, direction = mann_kendall(y)
            sl, ic, ci_lo, ci_hi = theil_sen(x, y)
            sig = '★' if p < 0.05 else ''
            print(f"  {STATION_LABELS[stn]:<34} {month_name:>6} {len(y):>4} {y.mean():>9.2f} {sl:>+13.4f}"
                  f" {tau:>8.4f} {p:>8.4f} {sig:>5}")
            monthly_summary_rows.append({
                'Station': STATION_LABELS[stn], 'Code': stn,
                'Month': month_name,
                'n_years': len(y),
                'Mean_C': round(y.mean(), 3),
                'Mean_F': round(y.mean() * 9/5 + 32, 3),
                'TheilSen_slope_C_yr': round(sl, 5),
                'TheilSen_CI_low': round(ci_lo, 5),
                'TheilSen_CI_high': round(ci_hi, 5),
                'MK_tau': round(tau, 4),
                'MK_S': S, 'MK_Z': round(Z, 4),
                'MK_p': round(p, 4),
                'Trend': direction,
                'Significant_p05': 'Yes' if p < 0.05 else 'No',
            })
    monthly_summary_df = pd.DataFrame(monthly_summary_rows)

    # FDRW-specific stats
    ann_fdrw     = ann_summer(stations['FDRW'])
    x_c, y_c     = ann_fdrw.index.values.astype(float), ann_fdrw.values
    tau_c, p_c, S_c, Z_c, dir_c = mann_kendall(y_c)
    sl_c, ic_c, ci_lo_c, ci_hi_c = theil_sen(x_c, y_c)

    # FDRW per-month water trends
    fdrw_month_trends = {}
    for month, month_name in MONTH_TUPLES:
        am = ann_month(stations['FDRW'], month)
        xm, ym = am.index.values.astype(float), am.values
        taum, pm, *_ = mann_kendall(ym)
        slm, icm, ci_lo_m, ci_hi_m = theil_sen(xm, ym)
        fdrw_month_trends[month_name] = dict(
            ann=am, x=xm, y=ym, tau=taum, p=pm,
            sl=slm, ic=icm, ci_lo=ci_lo_m, ci_hi=ci_hi_m,
        )

    # Air temperature trends
    ann_air       = air[air['month'].isin(SUMMER_MONTHS)].groupby('year')['tavg_c'].mean()
    x_a, y_a      = ann_air.index.values.astype(float), ann_air.values
    tau_a, p_a, *_ = mann_kendall(y_a)
    sl_a, ic_a, *_ = theil_sen(x_a, y_a)

    air_month_trends = {}
    for month, month_name in MONTH_TUPLES:
        am = air[air['month'] == month].groupby('year')['tavg_c'].mean()
        xm, ym = am.index.values.astype(float), am.values
        taum, pm, *_ = mann_kendall(ym)
        slm, icm, *_ = theil_sen(xm, ym)
        air_month_trends[month_name] = dict(
            x=xm, y=ym, tau=taum, p=pm, sl=slm, ic=icm,
        )

    baseline_yr_label = f"{int(fdrw.year.min())}–{int(fdrw.year.max())}"
    climate_scenarios = [
        (f'Baseline ({baseline_yr_label})', 0.00, '#2980b9'),
    ] + [(label, delta, color) for label, delta, color in CLIMATE_SCENARIOS[1:]]

    # Combined JJA projections
    baseline_air = air[air['month'].isin(SUMMER_MONTHS)]['tavg_c'].mean()
    baseline_wt  = stations['FDRW'][stations['FDRW']['month'].isin(SUMMER_MONTHS)]['wtc'].mean()

    print(f"\n" + "=" * 76)
    print(f"  CLIMATE PROJECTIONS — FDRW (Grand Coulee Dam Forebay)")
    print(f"=" * 76)
    print(f"\n  Observed baseline ({baseline_yr_label}):")
    print(f"    Jun-Aug mean air temp  : {baseline_air:.2f} °C  ({baseline_air*9/5+32:.2f} °F)")
    print(f"    Jun-Aug mean water temp: {baseline_wt:.2f} °C  ({baseline_wt*9/5+32:.2f} °F)")
    print(f"\n  Method: Delta approach using all-weeks Mohseni model (NSE={NSE_all:.3f})")
    print(f"\n  {'Scenario':<24} {'ΔTair':>8} {'Proj.Air':>10} {'Proj.Tw':>10}"
          f" {'Proj.Tw(°F)':>12} {'ΔTw':>8}")
    print("  " + "-" * 76)

    proj_rows = []
    baseline_tw_model = float(mohseni(baseline_air, *popt_all))
    for label, delta, color in climate_scenarios:
        proj_air = baseline_air + delta
        proj_tw  = mohseni(proj_air, *popt_all)
        dtw      = proj_tw - baseline_tw_model
        print(f"  {label:<24} {delta:>+8.2f} {proj_air:>10.2f} {proj_tw:>10.2f}"
              f" {proj_tw*9/5+32:>12.2f} {dtw:>+8.2f}")
        proj_rows.append({
            'Scenario': label, 'Delta_Air_C': round(delta, 2),
            'Proj_Air_C': round(proj_air, 2),
            'Proj_Tw_C':  round(proj_tw, 2),
            'Proj_Tw_F':  round(proj_tw * 9/5 + 32, 2),
            'Delta_Tw_C': round(dtw, 2),
            'Delta_Tw_F': round(dtw * 9/5, 2),
            'Color': color,
        })
    proj_df = pd.DataFrame(proj_rows)

    # Month-specific projections
    monthly_proj = {}
    for month, month_name in MONTH_TUPLES:
        m_air_baseline = air[air['month'] == month]['tavg_c'].mean()
        fdrw_month      = stations['FDRW'][stations['FDRW']['month'] == month]['wtc']
        m_wt_baseline  = fdrw_month.mean()
        m_wt_max       = fdrw_month.max()
        m_tw_model     = float(mohseni(m_air_baseline, *popt_all))

        m_rows = []
        print(f"\n  {month_name}-only projections (baseline air = {m_air_baseline:.2f} °C, "
              f"obs water = {m_wt_baseline:.2f} °C, obs max = {m_wt_max:.2f} °C, "
              f"model water = {m_tw_model:.2f} °C):")
        print(f"  {'Scenario':<24} {'ΔTair':>8} {'Proj.Air':>10} {'Proj.Tw':>10}"
              f" {'Proj.Tw(°F)':>12} {'Proj.Max':>10} {'ΔTw':>8}")
        print("  " + "-" * 88)
        for label, delta, color in climate_scenarios:
            proj_air = m_air_baseline + delta
            mohseni_dtw = float(mohseni(proj_air, *popt_all)) - m_tw_model
            proj_tw     = m_wt_baseline + mohseni_dtw
            proj_max_tw = m_wt_max      + mohseni_dtw
            print(f"  {label:<24} {delta:>+8.2f} {proj_air:>10.2f} {proj_tw:>10.2f}"
                  f" {proj_tw*9/5+32:>12.2f} {proj_max_tw:>10.2f} {mohseni_dtw:>+8.2f}")
            m_rows.append({
                'Scenario': label, 'Month': month_name,
                'Delta_Air_C': round(delta, 2),
                'Baseline_Air_C': round(m_air_baseline, 2),
                'Proj_Air_C': round(proj_air, 2),
                'Obs_Baseline_Tw_C': round(m_wt_baseline, 2),
                'Proj_Tw_C': round(proj_tw, 2),
                'Proj_Tw_F': round(proj_tw * 9/5 + 32, 2),
                'Delta_Tw_C': round(mohseni_dtw, 2),
                'Delta_Tw_F': round(mohseni_dtw * 9/5, 2),
                'Obs_Max_Tw_C': round(m_wt_max, 2),
                'Proj_Max_Tw_C': round(proj_max_tw, 2),
                'Proj_Max_Tw_F': round(proj_max_tw * 9/5 + 32, 2),
                'Color': color,
            })
        monthly_proj[month_name] = pd.DataFrame(m_rows)

    # Warm Day Scenario projections
    print(f"\n{'='*76}")
    print(f"  WARM DAY SCENARIO — 10% Exceedance (90th Percentile) Air Temperature")
    print(f"{'='*76}")

    warm_day_proj = {}
    for month, month_name in MONTH_TUPLES:
        air_month_daily = air[air['month'] == month]['tavg_c']
        air_p90 = air_month_daily.quantile(0.90)
        water_month_daily = stations['FDRW'][stations['FDRW']['month'] == month]['wtc']
        water_p90_obs = water_month_daily.quantile(0.90)
        water_p90_model = float(mohseni(air_p90, *popt_all))

        print(f"\n  {month_name} Warm Day (90th percentile):")
        print(f"    Air temp (10% exceed):   {air_p90:.2f} °C  ({air_p90*9/5+32:.2f} °F)")
        print(f"    Observed water (90th):   {water_p90_obs:.2f} °C  ({water_p90_obs*9/5+32:.2f} °F)")
        print(f"    Mohseni water (90th):    {water_p90_model:.2f} °C  ({water_p90_model*9/5+32:.2f} °F)")

        print(f"\n  {month_name} Warm Day Projections:")
        print(f"  {'Scenario':<24} {'ΔTair':>8} {'Proj.Air':>10} {'Proj.Tw':>10}"
              f" {'Proj.Tw(°F)':>12} {'ΔTw':>8}")
        print("  " + "-" * 76)

        wd_rows = []
        for label, delta, color in climate_scenarios:
            proj_air_wd = air_p90 + delta
            mohseni_dtw_wd = float(mohseni(proj_air_wd, *popt_all)) - water_p90_model
            proj_tw_wd = water_p90_obs + mohseni_dtw_wd

            print(f"  {label:<24} {delta:>+8.2f} {proj_air_wd:>10.2f} {proj_tw_wd:>10.2f}"
                  f" {proj_tw_wd*9/5+32:>12.2f} {mohseni_dtw_wd:>+8.2f}")

            wd_rows.append({
                'Scenario': label, 'Month': month_name,
                'Delta_Air_C': round(delta, 2),
                'Baseline_Air_P90_C': round(air_p90, 2),
                'Proj_Air_P90_C': round(proj_air_wd, 2),
                'Obs_Baseline_Tw_P90_C': round(water_p90_obs, 2),
                'Proj_Tw_P90_C': round(proj_tw_wd, 2),
                'Proj_Tw_P90_F': round(proj_tw_wd * 9/5 + 32, 2),
                'Delta_Tw_C': round(mohseni_dtw_wd, 2),
                'Delta_Tw_F': round(mohseni_dtw_wd * 9/5, 2),
                'Color': color,
            })
        warm_day_proj[month_name] = pd.DataFrame(wd_rows)

    print(f"{'='*76}\n")

    # Week-by-week stats — FDRW JJA
    fdrw_summer = stations['FDRW'][stations['FDRW']['month'].isin(SUMMER_MONTHS)].copy()
    weekly_by_yr = fdrw_summer.groupby(['year', 'week', 'month'])['wtc'].mean().reset_index()

    week_stats_df = weekly_by_yr.groupby('week').agg(
        mean_tw=('wtc', 'mean'),
        min_tw=('wtc', 'min'),
        max_tw=('wtc', 'max'),
        n_years=('wtc', 'count'),
    ).reset_index()

    week_month = weekly_by_yr.groupby('week')['month'].agg(lambda x: x.mode()[0]).reset_index()
    week_month.columns = ['week', 'primary_month']
    week_stats_df = week_stats_df.merge(week_month, on='week')

    week_stats_df = week_stats_df[(week_stats_df['primary_month'].isin(SUMMER_MONTHS)) &
                                  (week_stats_df['n_years'] >= 3)].sort_values('week').reset_index(drop=True)

    print(f"\n  Week-by-week water temp stats (FDRW, Jun-Aug weeks):")
    month_short = {6: 'Jun', 7: 'Jul', 8: 'Aug'}
    for _, r in week_stats_df.iterrows():
        m_name = month_short.get(r['primary_month'], '?')
        print(f"    Week {int(r['week']):>2} ({m_name}): mean={r['mean_tw']:.2f}, "
              f"min={r['min_tw']:.2f}, max={r['max_tw']:.2f}, n={int(r['n_years'])}")

    results = dict(
        summary_df=summary_df, monthly_summary_df=monthly_summary_df,
        proj_df=proj_df,
        stations=stations, air=air, merged=merged,
        popt_all=popt_all, NSE_all=NSE_all,
        popt_jja=popt_jja, NSE_jja=NSE_jja,
        ann_fdrw=ann_fdrw,
        fdrw_month_trends=fdrw_month_trends,
        air_month_trends=air_month_trends,
        x_c=x_c, y_c=y_c,
        tau_c=tau_c, p_c=p_c, sl_c=sl_c, ic_c=ic_c,
        ci_lo_c=ci_lo_c, ci_hi_c=ci_hi_c,
        x_a=x_a, y_a=y_a, sl_a=sl_a, ic_a=ic_a, tau_a=tau_a, p_a=p_a,
        baseline_air=baseline_air, baseline_wt=baseline_wt,
        jja_mask=jja_mask,
        climate_scenarios=climate_scenarios,
        monthly_proj=monthly_proj,
        warm_day_proj=warm_day_proj,
        week_stats_df=week_stats_df,
    )
    return results


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 4 — EXCEL EXPORT
# ═════════════════════════════════════════════════════════════════════════════

def export_excel(results):
    print("\n  Exporting Excel results …")
    summary_df  = results['summary_df']
    proj_df     = results['proj_df']
    popt_all    = results['popt_all']
    NSE_all     = results['NSE_all']
    popt_jja    = results['popt_jja']
    NSE_jja     = results['NSE_jja']
    ann_fdrw    = results['ann_fdrw']
    fdrw_month_trends = results['fdrw_month_trends']
    stations_data = results['stations']
    air_data     = results['air']

    fdrw_yr_range = f"{int(ann_fdrw.index.min())}–{int(ann_fdrw.index.max())}"
    water_yr_min = min(stations_data[s]['year'].min() for s in STATION_ORDER)
    water_yr_max = max(stations_data[s]['year'].max() for s in STATION_ORDER)
    water_yr_range = f"{water_yr_min}–{water_yr_max}"
    air_yr_range = f"{air_data['year'].min()}–{air_data['year'].max()}"

    HDR_FILL  = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    HDR_FONT  = Font(color='FFFFFF', bold=True)
    HDR_ALIGN = Alignment(horizontal='center')
    SEC_FILL  = PatternFill(start_color='1a4f72', end_color='1a4f72', fill_type='solid')
    SEC_FONT  = Font(color='FFFFFF', bold=True, size=10)
    FDRW_FILL = PatternFill(start_color='fde8e8', end_color='fde8e8', fill_type='solid')
    CHQW_FILL = PatternFill(start_color='fff4e6', end_color='fff4e6', fill_type='solid')
    BOLD      = Font(bold=True, size=10)
    NORMAL    = Font(size=10)
    ITALIC    = Font(italic=True, size=10)
    WRAP      = Alignment(wrap_text=True, vertical='top')

    def style_header(ws):
        for cell in ws[1]:
            cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = max(w + 2, 12)

    wb = Workbook()

    # Sheet 0 — Data Sources & Methods
    ws0 = wb.active
    ws0.title = 'Data Sources & Methods'

    def sec_header(ws, row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = SEC_FONT
        cell.fill = SEC_FILL
        cell.alignment = Alignment(vertical='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 18
        return row + 1

    def add_row(ws, row, label, value, note=None, lf=None, vf=None):
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = lf or BOLD
        c2.font = vf or NORMAL
        c2.alignment = WRAP
        if note:
            c3 = ws.cell(row=row, column=3, value=note)
            c3.font = ITALIC
            c3.alignment = WRAP
        return row + 1

    ws0.column_dimensions['A'].width = 32
    ws0.column_dimensions['B'].width = 60
    ws0.column_dimensions['C'].width = 50

    r = 1
    r = sec_header(ws0, r, 'ANALYSIS WINDOW')
    r = add_row(ws0, r, 'Summer months', SUMMER_LABEL_LONG)
    r += 1
    r = sec_header(ws0, r, 'WATER TEMPERATURE DATA')
    r = add_row(ws0, r, 'Source', 'DART — Columbia Basin Research, University of Washington')
    r = add_row(ws0, r, 'URL', 'http://www.cbr.washington.edu/dart/query/river_graph_wmq')
    r = add_row(ws0, r, 'Stations', 'FDRW (GC Forebay - PRIMARY), GCGW, CHJ, CHQW')
    r = add_row(ws0, r, 'Period', f'{water_yr_range}, hourly resolution')
    r += 1
    r = sec_header(ws0, r, 'AIR TEMPERATURE DATA')
    r = add_row(ws0, r, 'Source', 'NOAA GHCND — Station USR0000WDOU (Douglas, WA)')
    r = add_row(ws0, r, 'Location', '47.619°N, 119.899°W')
    r = add_row(ws0, r, 'Period', f'{air_yr_range}, daily TMAX/TMIN → mean')
    r += 1
    r = sec_header(ws0, r, 'STATISTICAL METHODS')
    r = add_row(ws0, r, 'Trend detection', 'Mann-Kendall test (Hirsch et al. 1991)')
    r = add_row(ws0, r, 'Trend magnitude', 'Theil-Sen median slope with 95% CI')
    r = add_row(ws0, r, 'Air-water model', 'Mohseni et al. (1998) nonlinear logistic regression')
    r = add_row(ws0, r, 'Goodness of fit', f'NSE: All-weeks={NSE_all:.3f}, Jun-Aug={NSE_jja:.3f}')
    r += 1
    r = sec_header(ws0, r, 'CLIMATE PROJECTIONS')
    r = add_row(ws0, r, 'Source', 'MACAv2-METDATA, 20 CMIP5 GCMs via Climate Toolbox')
    r = add_row(ws0, r, 'Scenarios', 'RCP 4.5 and RCP 8.5; 2040s and 2080s horizons')
    r = add_row(ws0, r, 'Method', 'Delta approach: Mohseni changes applied to observed baselines')

    # Sheet 1 — Combined JJA Trend Summary
    ws1 = wb.create_sheet(f'Trend Summary ({SUMMER_LABEL})')
    cols = list(summary_df.columns)
    ws1.append(cols)
    style_header(ws1)
    for _, row_data in summary_df.iterrows():
        ws1.append(list(row_data))
        if row_data['Code'] == 'FDRW':
            for cell in ws1[ws1.max_row]:
                cell.fill = FDRW_FILL
        elif row_data['Code'] == 'CHQW':
            for cell in ws1[ws1.max_row]:
                cell.fill = CHQW_FILL
    autofit(ws1)

    # Sheet 2 — Monthly Trend Summary
    monthly_summary_df = results['monthly_summary_df']
    ws1m = wb.create_sheet('Trend Summary (Monthly)')
    mcols = list(monthly_summary_df.columns)
    ws1m.append(mcols)
    style_header(ws1m)
    for _, row_data in monthly_summary_df.iterrows():
        ws1m.append(list(row_data))
        if row_data['Code'] == 'FDRW':
            for cell in ws1m[ws1m.max_row]:
                cell.fill = FDRW_FILL
        elif row_data['Code'] == 'CHQW':
            for cell in ws1m[ws1m.max_row]:
                cell.fill = CHQW_FILL
    autofit(ws1m)

    # Sheet 3 — Monthly Projections
    monthly_proj = results.get('monthly_proj', {})
    if monthly_proj:
        ws5 = wb.create_sheet('Monthly Projections')
        ws5.append(['Month', 'Scenario', 'Baseline_Air_C', 'Delta_Air_C',
                     'Proj_Air_C', 'Obs_Baseline_Tw_C', 'Proj_Tw_C', 'Proj_Tw_F',
                     'Delta_Tw_C', 'Delta_Tw_F',
                     'Obs_Max_Tw_C', 'Proj_Max_Tw_C', 'Proj_Max_Tw_F'])
        style_header(ws5)
        for _, month_name in MONTH_TUPLES:
            mdf = monthly_proj[month_name]
            for _, row in mdf.iterrows():
                ws5.append([month_name, row['Scenario'],
                           row['Baseline_Air_C'], row['Delta_Air_C'],
                           row['Proj_Air_C'], row['Obs_Baseline_Tw_C'],
                           row['Proj_Tw_C'], row['Proj_Tw_F'],
                           row['Delta_Tw_C'], row['Delta_Tw_F'],
                           row['Obs_Max_Tw_C'], row['Proj_Max_Tw_C'], row['Proj_Max_Tw_F']])
        autofit(ws5)

    # Sheet 3b — Warm Day Projections
    warm_day_proj = results.get('warm_day_proj', {})
    if warm_day_proj:
        ws_wd = wb.create_sheet('Warm Day Projections (P90)')
        ws_wd.append(['Month', 'Scenario', 'Baseline_Air_P90_C', 'Delta_Air_C',
                      'Proj_Air_P90_C', 'Obs_Baseline_Tw_P90_C', 'Proj_Tw_P90_C', 'Proj_Tw_P90_F',
                      'Delta_Tw_C', 'Delta_Tw_F'])
        style_header(ws_wd)
        for _, month_name in MONTH_TUPLES:
            wdf = warm_day_proj[month_name]
            for _, row in wdf.iterrows():
                ws_wd.append([month_name, row['Scenario'],
                             row['Baseline_Air_P90_C'], row['Delta_Air_C'],
                             row['Proj_Air_P90_C'], row['Obs_Baseline_Tw_P90_C'],
                             row['Proj_Tw_P90_C'], row['Proj_Tw_P90_F'],
                             row['Delta_Tw_C'], row['Delta_Tw_F']])
        autofit(ws_wd)

    # Sheet 4 — Combined JJA Projections
    ws2 = wb.create_sheet(f'Projections ({SUMMER_LABEL})')
    pcols = [c for c in proj_df.columns if c != 'Color']
    ws2.append(pcols)
    style_header(ws2)
    for _, row_data in proj_df.iterrows():
        ws2.append([row_data[c] for c in pcols])
    autofit(ws2)

    # Sheet 5 — FDRW Annual (Jun, Jul, Aug, combined)
    ws3 = wb.create_sheet('FDRW Annual')
    ws3.append(['Year',
                'Mean_Jun_Tw_C', 'Mean_Jun_Tw_F',
                'Mean_Jul_Tw_C', 'Mean_Jul_Tw_F',
                'Mean_Aug_Tw_C', 'Mean_Aug_Tw_F',
                'Mean_JJA_Tw_C', 'Mean_JJA_Tw_F'])
    style_header(ws3)
    jun_series = fdrw_month_trends['June']['ann']
    jul_series = fdrw_month_trends['July']['ann']
    aug_series = fdrw_month_trends['August']['ann']
    all_yrs = sorted(set(ann_fdrw.index) | set(jun_series.index) |
                     set(jul_series.index) | set(aug_series.index))
    for yr in all_yrs:
        jun_val = jun_series.get(yr, None)
        jul_val = jul_series.get(yr, None)
        aug_val = aug_series.get(yr, None)
        comb_val = ann_fdrw.get(yr, None)
        def cf(v):
            return round(v * 9/5 + 32, 3) if v is not None else None
        ws3.append([
            int(yr),
            round(jun_val, 3) if jun_val is not None else None, cf(jun_val),
            round(jul_val, 3) if jul_val is not None else None, cf(jul_val),
            round(aug_val, 3) if aug_val is not None else None, cf(aug_val),
            round(comb_val, 3) if comb_val is not None else None, cf(comb_val),
        ])
    autofit(ws3)

    # Sheet 4 — Mohseni Parameters
    ws4 = wb.create_sheet('Mohseni Parameters')
    ws4.append(['Fit', 'mu_C', 'alpha_C', 'gamma', 'beta_C', 'NSE', 'Notes'])
    style_header(ws4)
    ws4.append(['All weeks']    + [round(v, 5) for v in popt_all] + [round(NSE_all, 5), 'Fit to all weekly pairs'])
    ws4.append(['Jun–Aug only'] + [round(v, 5) for v in popt_jja] + [round(NSE_jja, 5), 'Fit to weeks 22–35 only'])
    autofit(ws4)

    # Sheet — Weekly Stats
    week_stats_df = results.get('week_stats_df', None)
    if week_stats_df is not None and len(week_stats_df) > 0:
        ws_wk = wb.create_sheet('Weekly Stats')
        ws_wk.append(['Scenario', 'Week', 'Month', 'Mean_Tw_C', 'Min_Tw_C', 'Max_Tw_C',
                       'Mean_Tw_F', 'Min_Tw_F', 'Max_Tw_F', 'n_years', 'Delta_Tw_C'])
        style_header(ws_wk)

        month_name_lookup = {6: 'June', 7: 'July', 8: 'August'}

        for _, r in week_stats_df.iterrows():
            m_name = month_name_lookup.get(r['primary_month'], '?')
            ws_wk.append(['Observed Baseline', int(r['week']), m_name,
                         round(r['mean_tw'], 3), round(r['min_tw'], 3), round(r['max_tw'], 3),
                         round(r['mean_tw'] * 9/5 + 32, 3), round(r['min_tw'] * 9/5 + 32, 3),
                         round(r['max_tw'] * 9/5 + 32, 3), int(r['n_years']), 0.0])

        if monthly_proj:
            for month_num, month_name in MONTH_TUPLES:
                mdf = monthly_proj[month_name]
                scen_rows = mdf[mdf['Delta_Air_C'] > 0]
                month_weeks = week_stats_df[week_stats_df['primary_month'] == month_num]
                for _, sr in scen_rows.iterrows():
                    dtw = sr['Delta_Tw_C']
                    for _, wr in month_weeks.iterrows():
                        ws_wk.append([sr['Scenario'], int(wr['week']), month_name,
                                     round(wr['mean_tw'] + dtw, 3),
                                     round(wr['min_tw'] + dtw, 3),
                                     round(wr['max_tw'] + dtw, 3),
                                     round((wr['mean_tw'] + dtw) * 9/5 + 32, 3),
                                     round((wr['min_tw'] + dtw) * 9/5 + 32, 3),
                                     round((wr['max_tw'] + dtw) * 9/5 + 32, 3),
                                     int(wr['n_years']), round(dtw, 3)])
        autofit(ws_wk)

    path = f'{OUTPUT_DIR}/GCD_Temperature_Results.xlsx'
    wb.save(path)
    print(f"    → Saved {path}")


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 5 — FIGURES
# ═════════════════════════════════════════════════════════════════════════════

def make_figures(results):
    print("\n" + "=" * 70)
    print("GENERATING FIGURES")
    print("=" * 70)

    stations    = results['stations']
    air         = results['air']
    merged      = results['merged']
    popt_all    = results['popt_all']
    NSE_all     = results['NSE_all']
    popt_jja    = results['popt_jja']
    NSE_jja     = results['NSE_jja']
    x_c, y_c    = results['x_c'], results['y_c']
    tau_c, p_c  = results['tau_c'], results['p_c']
    sl_c, ic_c  = results['sl_c'], results['ic_c']
    fdrw_month_trends = results['fdrw_month_trends']
    air_month_trends  = results['air_month_trends']
    x_a, y_a    = results['x_a'], results['y_a']
    sl_a, ic_a  = results['sl_a'], results['ic_a']
    tau_a, p_a  = results['tau_a'], results['p_a']

    baseline_air = results['baseline_air']
    baseline_wt  = results['baseline_wt']
    proj_df      = results['proj_df']
    jja_mask     = results['jja_mask']
    climate_scenarios = results['climate_scenarios']

    def ann_month(df, month):
        return df[df['month'] == month].groupby('year')['wtc'].mean()

    fdrw_yrs = stations['FDRW']['year']
    air_yrs  = air['year']

    def add_footnote(fig):
        pass

    all_years = sorted(stations['FDRW']['year'].unique())
    yr_cmap = cm.coolwarm
    yr_norm = mcolors.Normalize(vmin=min(all_years), vmax=max(all_years))

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 1 — Annual Jun/Jul/Aug means, all stations
    # ─────────────────────────────────────────────────────────────────────────
    print("\n  Figure 1 …")
    all_min_yr = min(stations[s]['year'].min() for s in STATION_ORDER)
    all_max_yr = max(stations[s]['year'].max() for s in STATION_ORDER)
    fdrw_min_yr = stations['FDRW']['year'].min()
    fdrw_max_yr = stations['FDRW']['year'].max()
    fdrw_n_years = len(stations['FDRW'][stations['FDRW']['month'].isin(SUMMER_MONTHS)]
                       .groupby('year')['wtc'].mean())

    fig, axes = plt.subplots(4, 3, figsize=(30, 21), sharex=True, sharey='row')
    fig.suptitle(
        f'Annual Mean Water Temperature by Station — June, July, and August Separate\n'
        f'Columbia River: Grand Coulee → Chief Joseph Dam  ({all_min_yr}–{all_max_yr})',
        fontsize=20, fontweight='bold', y=0.99
    )
    for i, stn in enumerate(STATION_ORDER):
        for j, (month, month_name) in enumerate(MONTH_TUPLES):
            ax = axes[i][j]
            ann = ann_month(stations[stn], month)
            x, y = ann.index.values.astype(float), ann.values
            if len(y) < 4:
                ax.set_visible(False)
                continue
            sl, ic, *_ = theil_sen(x, y)
            tau, p, S, Z, direction = mann_kendall(y)

            ax.fill_between(x, y, alpha=0.10, color=C[stn], zorder=1)
            ax.plot(x, y, color=C[stn], lw=2, marker='o', ms=4, zorder=4,
                    label=f'Annual {month_name} Mean')
            ax.plot(x, sl * x + ic, '--', color=C['trend'], lw=1.8, zorder=5,
                    label='Theil-Sen Trend')
            ax.set_xlim(all_min_yr - 1, all_max_yr + 1)
            ax.yaxis.set_minor_locator(MultipleLocator(0.5))
            sig_mark = '★' if p < 0.05 else ''
            ax.set_title(
                f'{STATION_LABELS[stn]} — {month_name}\n'
                f'TS: {sl:+.4f} °C/yr  |  MK τ={tau:.3f}, p={p:.3f} {sig_mark}',
                fontsize=15, color=C[stn], fontweight='bold', loc='left', pad=4
            )
            if j == 0:
                ax.set_ylabel('Water Temperature (°C)', fontsize=18)
            ax.legend(loc='lower left', fontsize=13, frameon=True,
                      edgecolor='#cccccc', framealpha=0.93)
    for j in range(3):
        axes[-1][j].set_xlabel('Year', fontsize=18)
    plt.tight_layout(rect=[0, 0.0, 1, 0.97])
    plt.savefig(f'{OUTPUT_DIR}/Fig1_Annual_Summer_Trends.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig1_Annual_Summer_Trends.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 2 — Seasonal climatology
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 2 …")
    fig, ax = plt.subplots(figsize=(18, 9))
    for stn in STATION_ORDER:
        clim = (stations[stn].groupby('doy')['wtc'].mean()
                .rolling(7, center=True, min_periods=1).mean())
        ax.plot(clim.index, clim.values, color=C[stn], lw=2.2, label=STATION_LABELS[stn])
    air_clim = (air.groupby('doy')['tavg_c'].mean()
                .rolling(7, center=True, min_periods=1).mean())
    ax.plot(air_clim.index, air_clim.values, color=C['air'], lw=1.5, ls=':',
            label='Air Temperature (GCD Area)')
    # Jun 1 = doy 152, Aug 31 = doy 243
    ax.axvspan(152, 243, alpha=0.09, color='gold', zorder=0, label='June–August Analysis Window')
    ax.text(197, 0.5, 'Jun–Aug\nwindow', ha='center', va='bottom',
            fontsize=16, color='#b7950b', style='italic')
    month_starts = [1, 32, 60, 91, 121, 152, 182, 213, 244, 274, 305, 335]
    month_names  = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    ax.set_xticks(month_starts)
    ax.set_xticklabels(month_names)
    ax.set_xlim(1, 365)
    ax.set_ylim(-1, 28)
    ax.set_xlabel('Month', fontsize=20)
    ax.set_ylabel('Temperature (°C)', fontsize=20)
    ax.set_title(f'Mean Annual Temperature Cycle — 7-Day Rolling Average ({all_min_yr}–{all_max_yr})',
                 fontsize=20, fontweight='bold')
    ax.legend(loc='upper left', ncol=1, fontsize=16, frameon=True, edgecolor='#cccccc')
    plt.tight_layout()
    plt.savefig(f'{OUTPUT_DIR}/Fig2_Seasonal_Climatology.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig2_Seasonal_Climatology.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 3 — FDRW detailed: Jun/Jul/Aug water + Jun/Jul/Aug Mohseni + air
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 3 …")
    fig = plt.figure(figsize=(18, 52))
    gs  = gridspec.GridSpec(7, 1, height_ratios=[3, 3, 3, 3, 3, 3, 2.5], hspace=0.30)
    fig.suptitle(
        'Grand Coulee Dam Forebay (FDRW) — June, July, August Temperature Trends\n'
        'Mann-Kendall Trend Test & Theil-Sen Slope Estimator',
        fontsize=20, fontweight='bold'
    )

    # Panels 1–3: water trends per month (top-down)
    panel_idx = 0
    for month, month_name in MONTH_TUPLES:
        ax = fig.add_subplot(gs[panel_idx])
        t = fdrw_month_trends[month_name]
        xm, ym = t['x'], t['y']
        slm, icm = t['sl'], t['ic']
        taum, pm = t['tau'], t['p']
        color = MONTH_COLORS[month_name]

        np.random.seed(42 + panel_idx)
        boot = []
        for _ in range(1500):
            idx = np.random.choice(len(xm), len(xm), replace=True)
            s_b, i_b, *_ = theil_sen(xm[idx], ym[idx])
            boot.append(s_b * xm + i_b)
        boot = np.array(boot)
        ax.fill_between(xm, np.percentile(boot, 2.5, axis=0),
                        np.percentile(boot, 97.5, axis=0),
                        alpha=0.15, color=color, label='95% Bootstrap CI')
        ax.bar(xm, ym, width=0.65, color=color, alpha=0.45, zorder=2)
        ax.plot(xm, ym, 'o-', color=color, lw=1.8, ms=5, zorder=4,
                label=f'Annual {month_name} Water Temp')
        ax.plot(xm, slm * xm + icm, '--', color=C['trend'], lw=2.2, zorder=5,
                label=f'Theil-Sen: {slm:+.4f} °C/yr  ({slm*10:+.3f} °C/decade)')
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.set_xlabel('Year', fontsize=20)
        sig_mark = '★ Significant' if pm < 0.05 else 'Not significant'
        ax.set_title(f'{month_name} Water Temperature Trend  '
                     f'(MK τ={taum:.3f}, p={pm:.3f}, {sig_mark})',
                     fontsize=20, fontweight='bold', color=color)
        ax.legend(loc='lower left', fontsize=16, frameon=True, edgecolor='#cccccc')
        panel_idx += 1

    # Panels 4–6: Mohseni predicted per month
    jja_merged = merged[merged['week'].between(22, 35)].copy()
    jja_merged['tw_pred'] = mohseni(jja_merged['tavg_c'].values, *popt_all)
    # Week ranges by month: Jun 22–26, Jul 27–30, Aug 31–35
    week_ranges = {'June': (22, 26), 'July': (27, 30), 'August': (31, 35)}
    pred_colors = {'June': '#2980b9', 'July': '#2980b9', 'August': '#2980b9'}

    for month, month_name in MONTH_TUPLES:
        ax = fig.add_subplot(gs[panel_idx])
        wlo, whi = week_ranges[month_name]
        sub = jja_merged[jja_merged['week'].between(wlo, whi)]
        if len(sub) == 0:
            ax.set_visible(False)
            panel_idx += 1
            continue
        ann_obs = sub.groupby('year')['wtc'].mean()
        ann_pred = sub.groupby('year')['tw_pred'].mean()
        xm = ann_obs.index.values.astype(float)
        ax.bar(xm - 0.18, ann_obs.values, width=0.36, color='#2980b9', alpha=0.65,
               zorder=2, label=f'Observed {month_name} Mean')
        ax.bar(xm + 0.18, ann_pred.values, width=0.36, color='#e74c3c', alpha=0.55,
               zorder=2, label=f'Mohseni Predicted {month_name} Mean')
        mae = np.mean(np.abs(ann_obs.values - ann_pred.values))
        ax.text(0.98, 0.03, f'MAE: {mae:.3f}°C', transform=ax.transAxes,
                fontsize=16, va='bottom', ha='right', family='monospace',
                bbox=dict(boxstyle='round,pad=0.2', facecolor='#fdfefe',
                          edgecolor='#bdc3c7', alpha=0.9))
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.set_xlabel('Year', fontsize=20)
        ax.set_title(f'{month_name} — Observed vs. Mohseni Predicted',
                     fontsize=20, fontweight='bold')
        ax.legend(loc='lower left', fontsize=16, frameon=True, edgecolor='#cccccc')
        panel_idx += 1

    # Panel 7: Air temperature trends Jun/Jul/Aug
    ax_air = fig.add_subplot(gs[panel_idx])
    offsets = {'June': -0.27, 'July': 0.0, 'August': 0.27}
    air_colors = {'June': '#27ae60', 'July': '#636e72', 'August': '#b2bec3'}
    for month, month_name in MONTH_TUPLES:
        t = air_month_trends[month_name]
        off = offsets[month_name]
        col = air_colors[month_name]
        sig = '  ★' if t['p'] < 0.05 else ''
        ax_air.bar(t['x'] + off, t['y'], width=0.27, color=col, alpha=0.5, zorder=2,
                   label=f'{month_name} Air  (TS: {t["sl"]:+.4f} °C/yr, p={t["p"]:.3f}{sig})')
        ax_air.plot(t['x'], t['sl'] * t['x'] + t['ic'], '--', color=col, lw=1.5, zorder=5)
    ax_air.set_ylabel('Air Temperature (°C)', fontsize=20)
    ax_air.set_xlabel('Year', fontsize=20)
    ax_air.set_title('Air Temperature Trend — June, July, August', fontsize=20, fontweight='bold')
    ax_air.legend(loc='lower left', fontsize=14, frameon=True, edgecolor='#cccccc')

    add_footnote(fig)
    plt.savefig(f'{OUTPUT_DIR}/Fig3_FDRW_Trend_Detail.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig3_FDRW_Trend_Detail.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 4 — Daily JJA water temp, all years overlaid
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 4 …")
    fdrw = stations['FDRW']
    summer = fdrw[fdrw['month'].isin(SUMMER_MONTHS)].copy()
    summer['jun1'] = pd.to_datetime(summer['year'].astype(str) + '-06-01')
    summer['day_of_summer'] = (summer['datetime'] - summer['jun1']).dt.total_seconds() / 86400.0

    fig, ax = plt.subplots(figsize=(21, 10.5))
    for yr in all_years:
        yr_data = summer[summer['year'] == yr]
        daily = yr_data.groupby(yr_data['day_of_summer'].astype(int))['wtc'].mean()
        ax.plot(daily.index, daily.values, color=yr_cmap(yr_norm(yr)), lw=1.0, alpha=0.55, zorder=2)

    overall_daily = summer.groupby(summer['day_of_summer'].astype(int))['wtc'].mean()
    ax.plot(overall_daily.index, overall_daily.values,
            color='black', lw=3, alpha=0.9, zorder=5, label=f'{fdrw_n_years}-Year Average Daily Cycle')

    ax.set_xlabel('Day of Summer', fontsize=20)
    ax.set_ylabel('Water Temperature (°C)', fontsize=20)
    ax.set_title('Grand Coulee Dam Forebay (FDRW) — Daily Mean Water Temperature\n'
                 f'June–August, Every Year {fdrw_min_yr}–{fdrw_max_yr}',
                 fontsize=20, fontweight='bold')
    # June (30 days) + July (31) + August (31) = 92 days
    ax.set_xlim(-1, 93)
    ax.set_xticks([0, 10, 20, 30, 40, 50, 61, 71, 81, 91])
    ax.set_xticklabels(['Jun 1', 'Jun 11', 'Jun 21', 'Jul 1', 'Jul 11', 'Jul 21',
                        'Aug 1', 'Aug 11', 'Aug 21', 'Aug 31'])

    sm = cm.ScalarMappable(cmap=yr_cmap, norm=yr_norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax, pad=0.01, fraction=0.03, aspect=30)
    cbar.set_label('Year', fontsize=20)
    cbar.set_ticks(list(range(all_min_yr, all_max_yr+1, 5)))
    year_line = Line2D([0], [0], color='gray', lw=1.0, alpha=0.55,
                       label='Individual Year Daily Mean (colored by year →)')
    handles, labels = ax.get_legend_handles_labels()
    handles.insert(0, year_line)
    ax.legend(handles=handles, loc='lower right', fontsize=16, frameon=True, edgecolor='#cccccc')
    add_footnote(fig)
    plt.tight_layout(rect=[0, 0.03, 1, 1])
    plt.savefig(f'{OUTPUT_DIR}/Fig4_JJA_Daily_By_Year.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig4_JJA_Daily_By_Year.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 5 — Mohseni regression scatter
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 5 …")
    SEASONS = {
        'Winter (Dec–Feb)': (merged['week'].isin(range(49, 53)) | (merged['week'] <= 8),  '#3498db'),
        'Spring (Mar–May)': (merged['week'].between(9, 21),   '#27ae60'),
        'Summer (Jun–Aug)': (merged['week'].between(22, 35),  '#e74c3c'),
        'Fall (Sep–Nov)':   (merged['week'].between(36, 48),  '#e67e22'),
    }
    fig, ax = plt.subplots(figsize=(15, 10.5))
    for label, (mask, color) in SEASONS.items():
        sub = merged[mask]
        ax.scatter(sub['tavg_c'], sub['wtc'], s=8, alpha=0.28, color=color, label=label, zorder=2)
    Ta_r = np.linspace(merged['tavg_c'].min() - 2, merged['tavg_c'].max() + 2, 400)
    ax.plot(Ta_r, mohseni(Ta_r, *popt_all), 'k-', lw=2.5, zorder=6,
            label=f'Mohseni fit — all weeks  (NSE = {NSE_all:.3f})')
    ax.plot(Ta_r, mohseni(Ta_r, *popt_jja), '--', color='darkred', lw=2, zorder=7,
            label=f'Mohseni fit — Jun–Aug  (NSE = {NSE_jja:.3f})')
    ax.axhline(popt_all[1], color='#555', lw=0.9, ls=':', zorder=1,
               label=f'Thermal Ceiling (α = {popt_all[1]:.1f}°C)')
    ax.axhline(popt_all[0], color='#555', lw=0.9, ls=':', zorder=1,
               label=f'Estimated Minimum (μ = {popt_all[0]:.1f}°C)')
    ax.set_xlabel('Weekly Mean Air Temperature (°C)', fontsize=20, labelpad=12)
    ax.set_ylabel('Weekly Mean Water Temperature (°C)', fontsize=20)
    ax.set_title('Mohseni Air–Water Temperature Regression — FDRW (GCD Forebay)\n'
                 f'Weekly Averages, {fdrw_min_yr}–{fdrw_max_yr}', fontsize=20, fontweight='bold')
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.13), ncol=4,
              fontsize=16, frameon=True, edgecolor='#cccccc')
    plt.tight_layout(rect=[0, 0.07, 1, 1.0])
    plt.savefig(f'{OUTPUT_DIR}/Fig5_Mohseni_Regression.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig5_Mohseni_Regression.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 6 — Observed vs. predicted scatter
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 6 …")
    merged_pred = merged.copy()
    merged_pred['tw_pred'] = mohseni(merged_pred['tavg_c'].values, *popt_all)
    ss_res = np.sum((merged_pred['wtc'] - merged_pred['tw_pred'])**2)
    ss_tot = np.sum((merged_pred['wtc'] - merged_pred['wtc'].mean())**2)
    r2 = 1 - ss_res / ss_tot

    fig, ax = plt.subplots(figsize=(15, 13.5))
    sc = ax.scatter(merged_pred['tw_pred'], merged_pred['wtc'],
                    c=merged_pred['year'], cmap=yr_cmap, norm=yr_norm,
                    s=14, alpha=0.4, edgecolors='none', zorder=3,
                    label=f'Weekly Data Points (n = {len(merged_pred):,}, colored by year →)')
    lo = min(merged_pred['wtc'].min(), merged_pred['tw_pred'].min()) - 1
    hi = max(merged_pred['wtc'].max(), merged_pred['tw_pred'].max()) + 1
    ax.plot([lo, hi], [lo, hi], 'k--', lw=1.5, alpha=0.7, zorder=5, label='1:1 Line (perfect prediction)')
    ax.axhline(popt_all[1], color='#c0392b', lw=1, ls=':', alpha=0.6, zorder=2,
               label=f'Thermal Ceiling — Observed Axis (α = {popt_all[1]:.1f}°C)')
    ax.axvline(popt_all[1], color='#c0392b', lw=1, ls=':', alpha=0.6, zorder=2,
               label=f'Thermal Ceiling — Predicted Axis (α = {popt_all[1]:.1f}°C)')
    ax.set_xlim(lo, hi); ax.set_ylim(lo, hi)
    ax.set_xlabel('Mohseni Predicted Water Temperature (°C)', fontsize=20)
    ax.set_ylabel('Observed Water Temperature (°C)', fontsize=20)
    ax.set_title('Mohseni Model — Observed vs. Predicted Weekly Water Temperature\n'
                 f'FDRW (GCD Forebay), All Weeks {fdrw_min_yr}–{fdrw_max_yr}',
                 fontsize=20, fontweight='bold')
    stats_text = f'NSE = {NSE_all:.3f}    R² = {r2:.3f}'
    ax.text(0.03, 0.97, stats_text, transform=ax.transAxes, fontsize=16, va='top', ha='left',
            family='monospace', bbox=dict(boxstyle='round,pad=0.4', facecolor='#fdfefe',
                                          edgecolor='#bdc3c7', alpha=0.9))
    ax.legend(loc='lower right', fontsize=16, frameon=True, edgecolor='#cccccc')
    cbar = plt.colorbar(sc, ax=ax, pad=0.01, fraction=0.03, aspect=25)
    cbar.set_label('Year', fontsize=20)
    add_footnote(fig)
    plt.tight_layout(rect=[0, 0.03, 1, 1])
    plt.savefig(f'{OUTPUT_DIR}/Fig6_Mohseni_Observed_vs_Predicted.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig6_Mohseni_Observed_vs_Predicted.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 7 — Jun, Jul, Aug obs vs predicted (3 panels)
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 7 …")
    jja = merged_pred[merged_pred['week'].between(22, 35)].copy()

    fig, axes = plt.subplots(1, 3, figsize=(30, 10.5), sharey=True)
    fig.suptitle(
        'Monthly Mean Water Temperature — Observed vs. Mohseni Predicted\n'
        f'FDRW (GCD Forebay), {fdrw_min_yr}–{fdrw_max_yr}',
        fontsize=20, fontweight='bold'
    )

    for idx, (month, month_name) in enumerate(MONTH_TUPLES):
        ax = axes[idx]
        wlo, whi = week_ranges[month_name]
        sub = jja[jja['week'].between(wlo, whi)]
        if len(sub) == 0:
            ax.set_visible(False)
            continue
        ann_obs = sub.groupby('year')['wtc'].mean()
        ann_pred = sub.groupby('year')['tw_pred'].mean()
        xm = ann_obs.index.values.astype(float)
        ax.bar(xm - 0.18, ann_obs.values, width=0.36, color='#2980b9', alpha=0.65,
               zorder=2, label=f'Observed {month_name} Mean')
        ax.bar(xm + 0.18, ann_pred.values, width=0.36, color='#e74c3c', alpha=0.55,
               zorder=2, label='Mohseni Predicted')
        mae = np.mean(np.abs(ann_obs.values - ann_pred.values))
        ax.text(0.98, 0.03, f'MAE: {mae:.3f}°C', transform=ax.transAxes,
                fontsize=16, va='bottom', ha='right', family='monospace',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='#fdfefe',
                          edgecolor='#bdc3c7', alpha=0.9))
        ax.set_xlabel('Year', fontsize=20)
        if idx == 0:
            ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.set_title(month_name, fontsize=20, fontweight='bold')
        ax.set_xlim(fdrw_min_yr - 1.5, fdrw_max_yr + 1.5)
        ax.legend(loc='lower left', fontsize=15, frameon=True, edgecolor='#cccccc')

    add_footnote(fig)
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.savefig(f'{OUTPUT_DIR}/Fig7_JJA_Obs_vs_Pred_TimeSeries.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig7_JJA_Obs_vs_Pred_TimeSeries.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 8 — All years overlaid on Mohseni curve
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 8 …")
    fig, ax = plt.subplots(figsize=(21, 13.5))
    for yr in all_years:
        yr_data = merged[merged['year'] == yr].sort_values('tavg_c')
        ax.plot(yr_data['tavg_c'], yr_data['wtc'],
                'o-', color=yr_cmap(yr_norm(yr)), ms=3.5, lw=0.8, alpha=0.45, zorder=2)
    ta_range = np.linspace(merged['tavg_c'].min() - 2, merged['tavg_c'].max() + 2, 300)
    ax.plot(ta_range, mohseni(ta_range, *popt_all), 'k-', lw=3, alpha=0.9, zorder=5,
            label=f'Mohseni Fitted Curve (NSE = {NSE_all:.3f})')
    ax.axhline(popt_all[1], color='#c0392b', lw=1.5, ls=':', alpha=0.7, zorder=4,
               label=f'Thermal Ceiling (α = {popt_all[1]:.1f}°C)')
    ax.axhline(popt_all[0], color='#2980b9', lw=1.5, ls=':', alpha=0.7, zorder=4,
               label=f'Estimated Minimum (μ = {popt_all[0]:.1f}°C)')
    ax.set_xlabel('Weekly Mean Air Temperature (°C)', fontsize=20)
    ax.set_ylabel('Weekly Mean Water Temperature (°C)', fontsize=20)
    ax.set_title(f'Mohseni Air–Water Temperature Relationship — All {fdrw_n_years} Years Overlaid\n'
                 f'FDRW (GCD Forebay), Weekly Averages {fdrw_min_yr}–{fdrw_max_yr}',
                 fontsize=20, fontweight='bold')
    sm = cm.ScalarMappable(cmap=yr_cmap, norm=yr_norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax, pad=0.01, fraction=0.03, aspect=30)
    cbar.set_label('Year', fontsize=20)
    cbar.set_ticks(list(range(all_min_yr, all_max_yr+1, 5)))
    year_line = Line2D([0], [0], color='gray', lw=0.8, marker='o', ms=3.5, alpha=0.45,
                       label='Individual Year Weekly Data (colored by year →)')
    handles, labels = ax.get_legend_handles_labels()
    handles.insert(0, year_line)
    ax.legend(handles=handles, loc='lower right', fontsize=16, frameon=True, edgecolor='#cccccc')
    add_footnote(fig)
    plt.tight_layout(rect=[0, 0.03, 1, 1])
    plt.savefig(f'{OUTPUT_DIR}/Fig8_Mohseni_All_Years_Overlaid.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig8_Mohseni_All_Years_Overlaid.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 9 — Projected temps bar chart (3 months)
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 9 …")
    monthly_proj = results.get('monthly_proj', {})
    if monthly_proj:
        jun_df_9 = monthly_proj['June']
        jul_df_9 = monthly_proj['July']
        aug_df_9 = monthly_proj['August']

        scenario_labels_9 = jun_df_9['Scenario'].values
        n9 = len(scenario_labels_9)
        x9 = np.arange(n9)
        bar_w9 = 0.27

        fig, ax = plt.subplots(figsize=(20, 10.5))
        bars_n9 = ax.bar(x9 - bar_w9, jun_df_9['Proj_Tw_C'].values, bar_w9,
                         label='June', color=MONTH_COLORS['June'], alpha=0.85,
                         edgecolor='white', linewidth=0.5)
        bars_j9 = ax.bar(x9,           jul_df_9['Proj_Tw_C'].values, bar_w9,
                         label='July', color=MONTH_COLORS['July'], alpha=0.85,
                         edgecolor='white', linewidth=0.5)
        bars_a9 = ax.bar(x9 + bar_w9,  aug_df_9['Proj_Tw_C'].values, bar_w9,
                         label='August', color=MONTH_COLORS['August'], alpha=0.85,
                         edgecolor='white', linewidth=0.5)

        for bars in [bars_n9, bars_j9, bars_a9]:
            for bar in bars:
                h = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2, h + 0.02,
                        f'{h:.2f}°C\n({h*9/5+32:.1f}°F)',
                        ha='center', va='bottom', fontsize=13, fontweight='bold')

        ax.set_xticks(x9)
        ax.set_xticklabels(scenario_labels_9, fontsize=14)
        ax.set_xlabel('Climate Scenario', fontsize=20)
        ax.set_ylabel('Projected Water Temperature (°C)', fontsize=20)
        ax.set_title('Projected June, July, and August Water Temperature at GCD Forebay (FDRW)\n'
                     'Mohseni (1998) Delta Method + CMIP5 Climate Scenarios (RCP 4.5 / 8.5)',
                     fontsize=20, fontweight='bold')
        all_vals = (list(jun_df_9['Proj_Tw_C'].values) + list(jul_df_9['Proj_Tw_C'].values)
                    + list(aug_df_9['Proj_Tw_C'].values))
        ax.set_ylim(min(all_vals) - 1.0, max(all_vals) + 1.5)
        ax.legend(fontsize=18, loc='upper left')
        ax.grid(axis='y', alpha=0.3)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.03, 1, 1])
        plt.savefig(f'{OUTPUT_DIR}/Fig9_Projected_Temps_Bar.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig9_Projected_Temps_Bar.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 10 — Projection points on Mohseni curve (3 months)
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 10 …")
    if monthly_proj:
        fig, ax = plt.subplots(figsize=(15, 10.5))
        ta_range = np.linspace(-10, 40, 500)
        ax.plot(ta_range, mohseni(ta_range, *popt_all), 'k-', lw=2.5, alpha=0.9, zorder=3,
                label='Mohseni Fitted Curve')
        ax.axhline(popt_all[1], color='#c0392b', lw=1.5, ls=':', alpha=0.7, zorder=2,
                   label=f'Thermal Ceiling (α = {popt_all[1]:.1f}°C)')
        ax.axhline(popt_all[0], color='#2980b9', lw=1.5, ls=':', alpha=0.7, zorder=2,
                   label=f'Estimated Minimum (μ = {popt_all[0]:.1f}°C)')

        scenario_colors = {'Baseline': '#2980b9', '2040s RCP 4.5': '#f57c00',
                          '2040s RCP 8.5': '#d32f2f', '2080s RCP 4.5': '#e65100',
                          '2080s RCP 8.5': '#b71c1c'}
        markers = {'June': '^', 'July': 'o', 'August': 's'}

        for month, month_name in MONTH_TUPLES:
            mdf = monthly_proj[month_name]
            marker = markers[month_name]
            for _, row in mdf.iterrows():
                tw_on_curve = float(mohseni(row['Proj_Air_C'], *popt_all))
                color = scenario_colors.get(row['Scenario'].split('(')[0].strip(),
                                            row.get('Color', '#333'))
                label_str = f"{month_name}: {row['Scenario']} ({row['Proj_Tw_C']:.1f}°C)"
                ax.plot(row['Proj_Air_C'], tw_on_curve, marker, color=color,
                        ms=12, zorder=5, markeredgecolor='white', markeredgewidth=1.5,
                        label=label_str)

        ax.set_xlabel('Air Temperature (°C)', fontsize=20)
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.set_title('Climate Projection Points on the Mohseni Air–Water Curve\n'
                     'FDRW (GCD Forebay) — June (▲), July (○), August (□) — CMIP5 RCP 4.5 / 8.5',
                     fontsize=20, fontweight='bold')
        ax.set_xlim(10, 35); ax.set_ylim(10, 22)
        ax.legend(loc='lower right', fontsize=12, frameon=True, edgecolor='#cccccc', ncol=3)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.03, 1, 1])
        plt.savefig(f'{OUTPUT_DIR}/Fig10_Projection_Points_Mohseni.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig10_Projection_Points_Mohseni.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 11 — Period comparison boxplots
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 11 …")
    mid_year = (all_min_yr + all_max_yr) // 2
    early_label = f'Early Record ({all_min_yr}–{mid_year})'
    recent_label = f'Recent Record ({mid_year+1}–{all_max_yr})'
    fig, axes = plt.subplots(2, 2, figsize=(19.5, 13.5))
    fig.suptitle(
        f'Water Temperature Distribution by Month — Early ({all_min_yr}–{mid_year}) vs. Recent ({mid_year+1}–{all_max_yr})\n'
        'All Stations  |  Wilcoxon p-values shown for June, July & August',
        fontsize=20, fontweight='bold'
    )
    for idx, stn in enumerate(STATION_ORDER):
        ax  = axes[idx // 2][idx % 2]
        df  = stations[stn][stations[stn]['month'].between(4, 10)].copy()
        early  = df[df['year'] <= mid_year]
        recent = df[df['year'] >= mid_year + 1]
        months = list(range(4, 11))
        mlabels = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct']

        def bp_kw(positions, color):
            return dict(positions=positions, widths=0.38, patch_artist=True,
                        medianprops=dict(color='white', lw=2),
                        whiskerprops=dict(color=color, lw=1.2),
                        capprops=dict(color=color),
                        boxprops=dict(facecolor=color, alpha=0.65),
                        flierprops=dict(marker='.', ms=2, alpha=0.25, color=color))

        ax.boxplot([early[early['month']  == m]['wtc'].dropna() for m in months],
                   **bp_kw([m - 0.22 for m in months], '#2980b9'))
        ax.boxplot([recent[recent['month'] == m]['wtc'].dropna() for m in months],
                   **bp_kw([m + 0.22 for m in months], '#c0392b'))

        ax.set_xticks(months)
        ax.set_xticklabels(mlabels)
        ax.set_title(STATION_LABELS[stn], fontsize=20, fontweight='bold',
                     color=C[stn], pad=6)
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        # Highlight Jun-Aug window
        ax.axvspan(5.6, 8.4, alpha=0.07, color='gold', zorder=0)
        ymax = ax.get_ylim()[1]
        for m in SUMMER_MONTHS:
            e = early[early['month']  == m]['wtc'].dropna().values
            r = recent[recent['month'] == m]['wtc'].dropna().values
            if len(e) > 5 and len(r) > 5:
                _, wp = stats.mannwhitneyu(e, r, alternative='two-sided')
                ax.text(m, ymax * 0.98, f'{"★" if wp < 0.05 else ""}p={wp:.3f}',
                        ha='center', va='top', fontsize=14,
                        color='darkred' if wp < 0.05 else '#888', fontweight='bold')

    axes[1][0].set_xlabel('Month', fontsize=20)
    axes[1][1].set_xlabel('Month', fontsize=20)
    fig.legend(
        handles=[mpatches.Patch(facecolor='#2980b9', alpha=0.65, label=early_label),
                 mpatches.Patch(facecolor='#c0392b', alpha=0.65, label=recent_label),
                 Line2D([0], [0], color='white', lw=2, label='Median (white line in box)'),
                 Line2D([0], [0], color='gray', lw=1.2, label='Whiskers (1.5× IQR)'),
                 Line2D([0], [0], color='gray', marker='.', ms=4, lw=0, alpha=0.4, label='Outliers'),
                 mpatches.Patch(facecolor='gold', alpha=0.15, label='June–August Analysis Window'),
                 Line2D([0], [0], color='white', marker='', label='p = Mann-Whitney U test'),
                 Line2D([0], [0], color='white', marker='', label='★ = significant (p < 0.05)')],
        loc='lower center', ncol=4, fontsize=16,
        bbox_to_anchor=(0.5, -0.01), frameon=True, edgecolor='#cccccc'
    )
    plt.tight_layout(rect=[0, 0.04, 1, 0.97])
    plt.savefig(f'{OUTPUT_DIR}/Fig11_Period_Comparison_Boxplots.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig11_Period_Comparison_Boxplots.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 12 — Monthly Projections: ΔTw by Month (3 bars per scenario)
    # ─────────────────────────────────────────────────────────────────────────
    if monthly_proj:
        print("  Figure 12 …")
        fig, ax = plt.subplots(figsize=(20, 10.5))

        jun_df = monthly_proj['June']
        jul_df = monthly_proj['July']
        aug_df = monthly_proj['August']

        jun_scen = jun_df[jun_df['Delta_Air_C'] > 0].reset_index(drop=True)
        jul_scen = jul_df[jul_df['Delta_Air_C'] > 0].reset_index(drop=True)
        aug_scen = aug_df[aug_df['Delta_Air_C'] > 0].reset_index(drop=True)

        scenario_labels = jun_scen['Scenario'].values
        n = len(scenario_labels)
        x = np.arange(n)
        bar_w = 0.27

        jun_dtw = jun_scen['Delta_Tw_C'].values
        jul_dtw = jul_scen['Delta_Tw_C'].values
        aug_dtw = aug_scen['Delta_Tw_C'].values

        bars_jun = ax.bar(x - bar_w, jun_dtw, bar_w, label='June',
                          color=MONTH_COLORS['June'], alpha=0.85,
                          edgecolor='white', linewidth=0.5)
        bars_jul = ax.bar(x,         jul_dtw, bar_w, label='July',
                          color=MONTH_COLORS['July'], alpha=0.85,
                          edgecolor='white', linewidth=0.5)
        bars_aug = ax.bar(x + bar_w, aug_dtw, bar_w, label='August',
                          color=MONTH_COLORS['August'], alpha=0.85,
                          edgecolor='white', linewidth=0.5)

        for bars in [bars_jun, bars_jul, bars_aug]:
            for bar in bars:
                h = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2, h + 0.01,
                        f'+{h:.2f}°C\n(+{h*9/5:.2f}°F)',
                        ha='center', va='bottom', fontsize=13, fontweight='bold')

        ax.set_xticks(x)
        ax.set_xticklabels(scenario_labels, fontsize=14)
        ax.set_xlabel('Climate Scenario', fontsize=20)
        ax.set_ylabel('Projected Change in Water Temperature (°C)', fontsize=20)
        ax.set_ylim(0, max(max(jun_dtw), max(jul_dtw), max(aug_dtw)) * 1.45)

        jun_base = jun_df.loc[jun_df['Delta_Air_C'] == 0, 'Obs_Baseline_Tw_C'].values[0]
        jul_base = jul_df.loc[jul_df['Delta_Air_C'] == 0, 'Obs_Baseline_Tw_C'].values[0]
        aug_base = aug_df.loc[aug_df['Delta_Air_C'] == 0, 'Obs_Baseline_Tw_C'].values[0]
        baseline_label_fig = f"{fdrw_min_yr}–{fdrw_max_yr}"
        ax.text(0.02, 0.97,
                f'Observed Baselines ({baseline_label_fig}):\n'
                f'  June:    {jun_base:.2f}°C  ({jun_base*9/5+32:.1f}°F)\n'
                f'  July:    {jul_base:.2f}°C  ({jul_base*9/5+32:.1f}°F)\n'
                f'  August:  {aug_base:.2f}°C  ({aug_base*9/5+32:.1f}°F)',
                transform=ax.transAxes, fontsize=14, va='top', family='monospace',
                bbox=dict(boxstyle='round,pad=0.4', facecolor='lightyellow', alpha=0.9))

        ax.set_title('Projected Change in June, July, and August Water Temperature at GCD Forebay (FDRW)\n'
                     'Mohseni (1998) Delta Method + CMIP5 Climate Scenarios (RCP 4.5 / 8.5)',
                     fontsize=20, fontweight='bold')
        ax.legend(fontsize=18, loc='upper right')
        ax.grid(axis='y', alpha=0.3)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.04, 1, 0.97])
        plt.savefig(f'{OUTPUT_DIR}/Fig12_Monthly_Projections_Bar.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig12_Monthly_Projections_Bar.png")

        # ── Fig 13 — Mohseni sensitivity arrows (3 months) ──────────────
        print("  Figure 13 …")
        fig, ax = plt.subplots(figsize=(15, 10.5))
        Ta_r = np.linspace(10, 35, 500)
        ax.plot(Ta_r, mohseni(Ta_r, *popt_all), 'k-', lw=3, label='Mohseni Fitted Curve')
        ax.axhline(y=float(popt_all[1]), color='red', ls=':', alpha=0.6, lw=1,
                   label=f'Thermal Ceiling (α = {popt_all[1]:.1f}°C)')
        ax.axhline(y=float(popt_all[0]), color='blue', ls=':', alpha=0.6, lw=1,
                   label=f'Estimated Minimum (μ = {popt_all[0]:.1f}°C)')

        markers_13 = {'June': '^', 'July': 'o', 'August': 's'}
        for month, month_name in MONTH_TUPLES:
            mdf = monthly_proj[month_name]
            marker_style = markers_13[month_name]
            base_row = mdf[mdf['Delta_Air_C'] == 0].iloc[0]
            base_air = base_row['Baseline_Air_C']
            base_tw_mohseni = float(mohseni(base_air, *popt_all))

            ax.scatter(base_air, base_tw_mohseni, s=150, marker=marker_style,
                      c='#2980b9', edgecolors='black', linewidth=1, zorder=10,
                      label=f'{month_name} Baseline Air ({base_air:.1f}°C)')

            extreme = mdf[mdf['Delta_Air_C'] > 0].iloc[-1]
            ext_air = extreme['Proj_Air_C']
            ext_tw_mohseni = float(mohseni(ext_air, *popt_all))
            dtw = ext_tw_mohseni - base_tw_mohseni

            arrow_color = MONTH_COLORS[month_name]
            ax.annotate('', xy=(ext_air, ext_tw_mohseni),
                       xytext=(base_air, base_tw_mohseni),
                       arrowprops=dict(arrowstyle='->', color=arrow_color,
                                      lw=2.5, connectionstyle='arc3,rad=0.1'))
            mid_air = (base_air + ext_air) / 2
            mid_tw = (base_tw_mohseni + ext_tw_mohseni) / 2
            offset_y = {'June': 0.8, 'July': 0.4, 'August': -0.8}[month_name]
            ax.text(mid_air, mid_tw + offset_y,
                    f'{month_name}: ΔTw = +{dtw:.2f}°C\n(2080s RCP 8.5)',
                    fontsize=13, ha='center', fontweight='bold',
                    color=arrow_color,
                    bbox=dict(boxstyle='round,pad=0.2', facecolor='white', alpha=0.85))

            ax.scatter(ext_air, ext_tw_mohseni, s=150, marker=marker_style,
                      c='#8b0000', edgecolors='black', linewidth=1, zorder=10)

        ax.text(0.02, 0.98,
                'Arrows show Mohseni-derived ΔTw\n'
                'used in the delta method.\n'
                'Observed baselines differ from\n'
                'curve due to reservoir thermal lag.',
                transform=ax.transAxes, fontsize=14, va='top',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow', alpha=0.9))

        ax.set_xlabel('Air Temperature (°C)', fontsize=20)
        ax.set_ylabel('Mohseni-Predicted Water Temperature (°C)', fontsize=20)
        ax.set_title('Mohseni Model Sensitivity Used for Monthly Climate Projections\n'
                     'FDRW (GCD Forebay) — Delta Method (RCP 4.5 / 8.5)',
                     fontsize=20, fontweight='bold')
        ax.legend(loc='lower right', fontsize=12)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.04, 1, 0.97])
        plt.savefig(f'{OUTPUT_DIR}/Fig13_Monthly_Projection_Points.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig13_Monthly_Projection_Points.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 14 — Observed weekly min/max/mean, Jun-Aug
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 14 …")
    week_stats_df = results.get('week_stats_df', None)

    if week_stats_df is not None and len(week_stats_df) > 0:
        wdf = week_stats_df.sort_values('week').reset_index(drop=True)
        n_weeks = len(wdf)
        x = np.arange(n_weeks)
        colors = [MONTH_COLORS['June'] if m == 6 else
                  (MONTH_COLORS['July'] if m == 7 else MONTH_COLORS['August'])
                  for m in wdf['primary_month']]

        fig, ax = plt.subplots(figsize=(24, 10.5))
        bars = ax.bar(x, wdf['mean_tw'].values, width=0.7, color=colors, alpha=0.85,
                      edgecolor='white', linewidth=0.5, zorder=3)
        err_lo = (wdf['mean_tw'] - wdf['min_tw']).values
        err_hi = (wdf['max_tw'] - wdf['mean_tw']).values
        ax.errorbar(x, wdf['mean_tw'].values, yerr=[err_lo, err_hi], fmt='none',
                    ecolor='#2c3e50', elinewidth=1.8, capsize=6, capthick=1.8, zorder=5)

        for i, (_, r) in enumerate(wdf.iterrows()):
            ax.text(x[i], r['max_tw'] + 0.08, f'{r["max_tw"]:.1f}',
                    ha='center', va='bottom', fontsize=13, fontweight='bold', color='#555')
            ax.text(x[i], r['min_tw'] - 0.08, f'{r["min_tw"]:.1f}',
                    ha='center', va='top', fontsize=13, fontweight='bold', color='#555')

        month_short = {6: 'Jun', 7: 'Jul', 8: 'Aug'}
        week_labels = []
        for _, r in wdf.iterrows():
            m_name = month_short.get(r['primary_month'], '?')
            week_labels.append(f'Wk {int(r["week"])}\n({m_name})')
        ax.set_xticks(x)
        ax.set_xticklabels(week_labels, fontsize=13)
        ax.set_xlabel('Week', fontsize=20)
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.set_title(f'Observed Weekly Water Temperature at FDRW — June through August ({fdrw_min_yr}–{fdrw_max_yr})',
                     fontsize=20, fontweight='bold')

        jun_patch = mpatches.Patch(color=MONTH_COLORS['June'], alpha=0.85, label='June weeks')
        jul_patch = mpatches.Patch(color=MONTH_COLORS['July'], alpha=0.85, label='July weeks')
        aug_patch = mpatches.Patch(color=MONTH_COLORS['August'], alpha=0.85, label='August weeks')
        err_line = Line2D([0], [0], color='#2c3e50', lw=1.8, marker='_', ms=10,
                          label='Min/Max range across years')
        ax.legend(handles=[jun_patch, jul_patch, aug_patch, err_line], loc='lower right',
                  fontsize=14, frameon=True, edgecolor='#cccccc')

        ax.grid(axis='y', alpha=0.3)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.04, 1, 1])
        plt.savefig(f'{OUTPUT_DIR}/Fig14_Observed_Weekly_MinMaxMean.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig14_Observed_Weekly_MinMaxMean.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 15 — Projected weekly under 2080s RCP 8.5
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 15 …")
    monthly_proj_fig = results.get('monthly_proj', {})
    if week_stats_df is not None and len(week_stats_df) > 0 and monthly_proj_fig:
        wdf = week_stats_df.sort_values('week').reset_index(drop=True)
        n_weeks = len(wdf)
        x = np.arange(n_weeks)
        bar_w = 0.38

        jun_proj_df = monthly_proj_fig['June']
        jul_proj_df = monthly_proj_fig['July']
        aug_proj_df = monthly_proj_fig['August']
        jun_scen = jun_proj_df[jun_proj_df['Delta_Air_C'] > 0].reset_index(drop=True)
        jul_scen = jul_proj_df[jul_proj_df['Delta_Air_C'] > 0].reset_index(drop=True)
        aug_scen = aug_proj_df[aug_proj_df['Delta_Air_C'] > 0].reset_index(drop=True)
        scenarios_to_plot = []
        for i in range(len(jun_scen)):
            scenarios_to_plot.append({
                'label': jun_scen.iloc[i]['Scenario'],
                6: jun_scen.iloc[i]['Delta_Tw_C'],
                7: jul_scen.iloc[i]['Delta_Tw_C'],
                8: aug_scen.iloc[i]['Delta_Tw_C'],
            })

        fig, ax = plt.subplots(figsize=(24, 10.5))

        colors_obs = [MONTH_COLORS['June'] if m == 6 else
                      (MONTH_COLORS['July'] if m == 7 else MONTH_COLORS['August'])
                      for m in wdf['primary_month']]
        bars_obs = ax.bar(x - bar_w/2, wdf['mean_tw'].values, bar_w, color=colors_obs, alpha=0.5,
                          edgecolor='white', linewidth=0.5, zorder=3)
        err_lo_obs = (wdf['mean_tw'] - wdf['min_tw']).values
        err_hi_obs = (wdf['max_tw'] - wdf['mean_tw']).values
        ax.errorbar(x - bar_w/2, wdf['mean_tw'].values, yerr=[err_lo_obs, err_hi_obs], fmt='none',
                    ecolor='#888', elinewidth=1.2, capsize=4, capthick=1.2, zorder=5)

        extreme = scenarios_to_plot[-1]
        proj_means = []
        proj_mins = []
        proj_maxs = []
        for _, r in wdf.iterrows():
            dtw = extreme[r['primary_month']]
            proj_means.append(r['mean_tw'] + dtw)
            proj_mins.append(r['min_tw'] + dtw)
            proj_maxs.append(r['max_tw'] + dtw)
        proj_means = np.array(proj_means)
        proj_mins = np.array(proj_mins)
        proj_maxs = np.array(proj_maxs)
        dark_colors = {'June': '#1e8449', 'July': '#c0392b', 'August': '#1a5276'}
        colors_proj = [dark_colors['June'] if m == 6 else
                       (dark_colors['July'] if m == 7 else dark_colors['August'])
                       for m in wdf['primary_month']]

        bars_proj = ax.bar(x + bar_w/2, proj_means, bar_w, color=colors_proj, alpha=0.85,
                           edgecolor='white', linewidth=0.5, zorder=3)
        err_lo_proj = proj_means - proj_mins
        err_hi_proj = proj_maxs - proj_means
        ax.errorbar(x + bar_w/2, proj_means, yerr=[err_lo_proj, err_hi_proj], fmt='none',
                    ecolor='#2c3e50', elinewidth=1.8, capsize=4, capthick=1.8, zorder=5)

        for i in range(n_weeks):
            ax.text(x[i] + bar_w/2, proj_maxs[i] + 0.08, f'{proj_maxs[i]:.1f}',
                    ha='center', va='bottom', fontsize=12, fontweight='bold', color='#555')
            ax.text(x[i] + bar_w/2, proj_mins[i] - 0.08, f'{proj_mins[i]:.1f}',
                    ha='center', va='top', fontsize=12, fontweight='bold', color='#555')

        for i, (_, r) in enumerate(wdf.iterrows()):
            ax.text(x[i] - bar_w/2, r['max_tw'] + 0.08, f'{r["max_tw"]:.1f}',
                    ha='center', va='bottom', fontsize=12, fontweight='bold', color='#999')
            ax.text(x[i] - bar_w/2, r['min_tw'] - 0.08, f'{r["min_tw"]:.1f}',
                    ha='center', va='top', fontsize=12, fontweight='bold', color='#999')

        month_short = {6: 'Jun', 7: 'Jul', 8: 'Aug'}
        week_labels = []
        for _, r in wdf.iterrows():
            m_name = month_short.get(r['primary_month'], '?')
            week_labels.append(f'Wk {int(r["week"])}\n({m_name})')
        ax.set_xticks(x)
        ax.set_xticklabels(week_labels, fontsize=13)
        ax.set_xlabel('Week', fontsize=20)
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.set_title('Projected Weekly Water Temperature at FDRW — June through August\n'
                     f'Observed vs. {extreme["label"]}',
                     fontsize=20, fontweight='bold')

        obs_jun = mpatches.Patch(color=MONTH_COLORS['June'], alpha=0.5, label='Observed June')
        obs_jul = mpatches.Patch(color=MONTH_COLORS['July'], alpha=0.5, label='Observed July')
        obs_aug = mpatches.Patch(color=MONTH_COLORS['August'], alpha=0.5, label='Observed August')
        proj_jun = mpatches.Patch(color=dark_colors['June'], alpha=0.85, label=f'Projected June')
        proj_jul = mpatches.Patch(color=dark_colors['July'], alpha=0.85, label=f'Projected July')
        proj_aug = mpatches.Patch(color=dark_colors['August'], alpha=0.85, label=f'Projected August')
        err_line = Line2D([0], [0], color='#2c3e50', lw=1.8, marker='_', ms=10,
                          label='Min/Max range')
        ax.legend(handles=[obs_jun, obs_jul, obs_aug, proj_jun, proj_jul, proj_aug, err_line],
                  loc='lower right', fontsize=12, frameon=True, edgecolor='#cccccc', ncol=2)

        ax.grid(axis='y', alpha=0.3)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.04, 1, 1])
        plt.savefig(f'{OUTPUT_DIR}/Fig15_Projected_Weekly_MinMaxMean.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig15_Projected_Weekly_MinMaxMean.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 16 — Average vs Warm Day Comparison (3 panels)
    # ─────────────────────────────────────────────────────────────────────────
    warm_day_proj = results.get('warm_day_proj', {})
    if warm_day_proj and monthly_proj:
        print("  Figure 16 …")
        fig, axes = plt.subplots(1, 3, figsize=(30, 10.5))

        for idx, (month, month_name) in enumerate(MONTH_TUPLES):
            ax = axes[idx]

            avg_df = monthly_proj[month_name]
            warm_df = warm_day_proj[month_name]

            avg_df_proj = avg_df[avg_df['Delta_Air_C'] > 0]
            warm_df_proj = warm_df[warm_df['Delta_Air_C'] > 0]

            if len(avg_df_proj) == 0:
                continue

            n_scenarios = len(avg_df_proj)
            x = np.arange(n_scenarios)
            width = 0.35

            bars1 = ax.bar(x - width/2, avg_df_proj['Proj_Tw_C'].values, width,
                          label='Average Day', color='#3498db', alpha=0.85, edgecolor='white', linewidth=1)
            bars2 = ax.bar(x + width/2, warm_df_proj['Proj_Tw_P90_C'].values, width,
                          label='Warm Day (90th %)', color='#e74c3c', alpha=0.85, edgecolor='white', linewidth=1)

            for bars in [bars1, bars2]:
                for bar in bars:
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'{height:.1f}°C\n({height*9/5+32:.1f}°F)',
                           ha='center', va='bottom', fontsize=10, fontweight='bold')

            baseline_avg = avg_df[avg_df['Delta_Air_C'] == 0]['Proj_Tw_C'].values[0]
            baseline_warm = warm_df[warm_df['Delta_Air_C'] == 0]['Proj_Tw_P90_C'].values[0]

            ax.axhline(baseline_avg, color='#3498db', linestyle='--', linewidth=1.5, alpha=0.5,
                      label=f'Baseline Avg: {baseline_avg:.1f}°C')
            ax.axhline(baseline_warm, color='#e74c3c', linestyle='--', linewidth=1.5, alpha=0.5,
                      label=f'Baseline Warm: {baseline_warm:.1f}°C')

            ax.set_xlabel('Climate Scenario', fontsize=16, fontweight='bold')
            if idx == 0:
                ax.set_ylabel('Water Temperature (°C)', fontsize=16, fontweight='bold')
            ax.set_title(f'{month_name} — Average vs Warm Day',
                        fontsize=18, fontweight='bold', pad=15)
            ax.set_xticks(x)
            ax.set_xticklabels([s.replace('Baseline ', '').replace(' RCP ', '\n')
                               for s in avg_df_proj['Scenario'].values],
                              fontsize=11, fontweight='bold')
            ax.legend(loc='lower left', fontsize=12, frameon=True, edgecolor='#cccccc', framealpha=0.95)
            ax.grid(axis='y', alpha=0.3)
            ax.tick_params(axis='both', labelsize=11)

        fig.suptitle('Projected Water Temperature at FDRW (GCD Forebay)\n'
                     'Average Day vs Warm Day (10% Exceedance) Scenarios',
                     fontsize=22, fontweight='bold', y=0.98)

        plt.tight_layout(rect=[0, 0.0, 1, 0.96])
        plt.savefig(f'{OUTPUT_DIR}/Fig16_Average_vs_WarmDay_Comparison.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig16_Average_vs_WarmDay_Comparison.png")

    print(f"\n  All figures saved to {OUTPUT_DIR}")


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 6 — MAIN ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print("=" * 70)
    print("GCD Forebay Water Temperature — Full Analysis (Jun-Jul-Aug)")
    print("=" * 70)

    print("\n[1] Loading data …")
    stations = {s: load_station(s, STATION_FILES[s])
                for s in STATION_ORDER}
    air = load_air(AIR_TEMP_FILE)

    MIN_YEAR = 1991
    MAX_YEAR = 2026
    for s in stations:
        before = len(stations[s])
        stations[s] = stations[s][(stations[s]['year'] >= MIN_YEAR) &
                                  (stations[s]['year'] <= MAX_YEAR)].reset_index(drop=True)
        dropped = before - len(stations[s])
        if dropped > 0:
            print(f"    [{s}] Dropped {dropped:,} records outside {MIN_YEAR}–{MAX_YEAR}")
    before_air = len(air)
    air = air[(air['year'] >= MIN_YEAR) & (air['year'] <= MAX_YEAR)].reset_index(drop=True)
    dropped_air = before_air - len(air)
    if dropped_air > 0:
        print(f"    [Air] Dropped {dropped_air:,} records outside {MIN_YEAR}–{MAX_YEAR}")

    print("\n[2] Running statistical analysis …")
    results = run_analysis(stations, air)

    print("\n[3] Exporting results …")
    export_excel(results)

    print("\n[4] Generating figures …")
    make_figures(results)

    print("\n" + "=" * 70)
    print("COMPLETE — all outputs saved to:", OUTPUT_DIR)
    print("=" * 70)