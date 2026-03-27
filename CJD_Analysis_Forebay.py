"""
================================================================================
CJD Forebay Water Temperature — Full Analysis & Figures
================================================================================
Chief Joseph Dam (CJD) Forebay — Historical Trend Analysis & Climate Projections
Primary Station of Interest: CHJ (Chief Joseph Dam Forebay)
Supporting Stations: FDRW → GCGW → CHQW (provided for comparison and context)

--------------------------------------------------------------------------------
DATA SOURCES
--------------------------------------------------------------------------------
WATER TEMPERATURE DATA:
  Source  : DART (Data Access in Real Time)
            Columbia Basin Research, University of Washington
  URL     : http://www.cbr.washington.edu/dart/query/river_graph_wmq
  Stations: FDRW  — Grand Coulee Dam Forebay
            GCGW  — Grand Coulee Dam Tailrace
            CHJ   — Chief Joseph Dam Forebay  ← Primary station
            CHQW  — Chief Joseph Dam Tailrace (comparison)
  Format  : Hourly water temperature (°C and °F)
  Period  : CHJ: 2000–2025 (primary analysis)
            CHQW: 1997–2025 (comparison)
            Others (FDRW, GCGW): 2000–2025

AIR TEMPERATURE DATA:
  Source  : NOAA National Centers for Environmental Information (NCEI)
            Global Historical Climatology Network Daily (GHCND)
  Station : DOUGLAS WASHINGTON, WA US
  ID      : GHCND:USR0000WDOU
  Lat/Lon : 47.619124, -119.899292
  Period  : 1990-07-30 to 2026-02-21  (coverage: ~88%)
  Used    : 1995–2025 subset, daily TMAX/TMIN averaged to daily mean (°C)

--------------------------------------------------------------------------------
METHODS
--------------------------------------------------------------------------------
  1. Mann-Kendall trend test
     Reference: Hirsch et al. (1991); Helsel & Hirsch (2002)
     Applied to annual July/August mean water temperatures, 2000–2025

  2. Theil-Sen slope estimator with 95% confidence interval
     Reference: Helsel & Hirsch (2002)
     Provides median trend slope robust to outliers and non-normality

  3. Mohseni et al. (1998) nonlinear logistic air-water temperature regression
        Tw = μ + (α − μ) / (1 + exp(γ(β − Ta)))
     Reference: Mohseni et al. (1998); Mantua et al. (2010); WACCIA Ch. 6
     Fit to all weekly air-water temperature pairs (2000–2025)
     Goodness-of-fit assessed via Nash-Sutcliffe Efficiency (NSE)

  4. Future climate projections via CMIP5 air temperature deltas
     Source:  MACAv2-METDATA (Abatzoglou & Brown 2012), 20 CMIP5 GCMs
              accessed via Climate Toolbox (climatetoolbox.org)
     Grid cell: 47.619 N, 119.899 W (CJD area, Douglas County, WA)
     Variable: JJA mean air temperature = (daily Tmax + daily Tmin) / 2
     Baseline: 1970-1999 model historical period
     Scenarios: RCP 4.5 (moderate emissions): +2.36 C by 2040s, +3.46 C by 2080s
                RCP 8.5 (high emissions):     +2.97 C by 2040s, +6.11 C by 2080s
     NOTE: Replaces CMIP3-era A1B/B1 scenarios from Mantua et al. (2010).
     Deltas computed as difference between future-period ensemble means and
     the 1970-1999 historical baseline, applied to observed 2000-2025 air
     temperature at the Douglas County station.

--------------------------------------------------------------------------------
OUTPUTS
--------------------------------------------------------------------------------
  CJD_Temperature_Results.xlsx             — trend summary, projections, Mohseni params
  Fig1_Annual_Summer_Trends.png            — annual Jul/Aug means, all stations
  Fig2_Seasonal_Climatology.png            — mean annual temperature cycle
  Fig3_CHJ_Trend_Detail.png                — CHJ detailed trend + air temp overlay
  Fig4_JulAug_Daily_By_Year.png            — daily Jul/Aug temps, all years overlaid
  Fig5_Mohseni_Regression.png              — Mohseni air-water regression by season
  Fig6_Mohseni_Observed_vs_Predicted.png   — observed vs predicted weekly scatter
  Fig7_JulAug_Obs_vs_Pred_TimeSeries.png   — Jul/Aug annual observed vs predicted
  Fig8_Mohseni_All_Years_Overlaid.png      — all years overlaid on Mohseni curve
  Fig9_Projected_Temps_Bar.png             — projected temps bar chart by scenario
  Fig10_Projection_Points_Mohseni.png      — projection points on Mohseni curve
  Fig11_Period_Comparison_Boxplots.png     — early vs. recent period comparison

--------------------------------------------------------------------------------
REFERENCES
--------------------------------------------------------------------------------
  Helsel, D.R. and Hirsch, R.M. (2002). Statistical Methods in Water Resources.
    USGS Techniques of Water-Resources Investigations, Book 4, Ch. A3.

  Hirsch, R.M., Alexander, R.B., and Smith, R.A. (1991). Selection of methods
    for the detection and estimation of trends in water quality. Water Resources
    Research, 27(5), 803–813.

  Mantua, N., Tohver, I., and Hamlet, A. (2010). Climate change impacts on
    streamflow extremes and summertime stream temperature and their possible
    consequences for freshwater salmon habitat in Washington State. Climatic
    Change, 102, 187–223.

  Abatzoglou, J.T. and Brown, T.J. (2012). A comparison of statistical
    downscaling methods suited for wildfire applications. International Journal
    of Climatology, 32, 772-780.  [MACAv2-METDATA downscaling method]

  Mohseni, O., Stefan, H.G., and Erickson, T.R. (1998). A nonlinear regression
    model for weekly stream temperatures. Water Resources Research, 34(10),
    2685–2692.

  O'Connor, J.E. et al. (2021). Water temperature trends in the Columbia River
    Basin. USACE, Institute for Water Resources.

  Vermeyen, T.B. (2000). Using selective withdrawal to control reservoir
    releases and downstream temperatures. USBR PAP-0854.

  Washington Climate Change Impacts Assessment (WACCIA), Chapter 6 — Salmon.
    Mantua et al. (contributing authors).
================================================================================
"""

import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
import matplotlib.cm as cm
import matplotlib.colors as mcolors
from matplotlib.lines import Line2D
from matplotlib.ticker import MultipleLocator
from scipy import stats, optimize
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# FILE PATHS  —  update these if data is moved
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT_DIR = r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\results'

STATION_FILES = {
    'FDRW': r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\FDRW\FDRW_Hourly_Data_1995_2025.xlsx',
    'GCGW': r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\GCGW\GCGW_Hourly_Data_1995_2025.xlsx',
    'CHJ':  r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\CHJ\CHJ_Hourly_Data_1995_2025.xlsx',
    # CHQW: included for comparison purposes; primary analysis focus is on CHJ
    'CHQW': r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\CHQW\CHQW_Hourly_Data_1995_2025.xlsx',
}

AIR_TEMP_FILE = r'C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\Douglas_Temp_1995_2025.csv'

os.makedirs(OUTPUT_DIR, exist_ok=True)

SUMMER_MONTHS  = [7, 8]
STATION_ORDER  = ['FDRW', 'GCGW', 'CHJ', 'CHQW']
STATION_LABELS = {
    'FDRW': 'Grand Coulee Forebay (FDRW)',
    'GCGW': 'Grand Coulee Tailrace (GCGW)',
    'CHJ':  'Chief Joseph Forebay (CHJ)',
    'CHQW': 'Chief Joseph Tailrace (CHQW)',
}

# ── CMIP5 Climate Projection Delta Computation ────────────────────────────────
# Source: MACAv2-METDATA (Abatzoglou & Brown 2012), 20 CMIP5 GCMs
#         Climate Toolbox (climatetoolbox.org)
# Grid cell: 47.619 N, 119.899 W (Douglas County, WA — nearest to CJD)
# Variable: JJA mean air temperature = (daily Tmax + daily Tmin) / 2
# Scenarios: RCP 4.5 (moderate emissions), RCP 8.5 (high emissions)
# Baseline:  1970-1999 model historical period (20-model ensemble mean)
#
# Input files (download from Climate Toolbox > Future Time Series > Download
#   all models, selecting JJA Max and Min Temperature in Celsius):
MACA_TMAX_FILE = r"C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\FutureTimeSeriesMax.csv"
MACA_TMIN_FILE = r"C:\Users\Ethan.Muhlestein\OneDrive - Kleinschmidt Associates\Documents\CJD_Temp_Monitoring\Data\FutureTimeSeriesMin.csv"

def compute_cmip5_deltas(tmax_path, tmin_path, header_rows=51):
    """
    Compute JJA mean air temperature deltas from MACAv2-METDATA CMIP5 ensemble.

    Reads Climate Toolbox CSV exports of JJA Max and Min Temperature (Celsius),
    computes mean = (Tmax + Tmin) / 2, then calculates deltas for the 2040s
    (2030-2059) and 2080s (2070-2099) relative to the 1970-1999 historical
    baseline for both RCP 4.5 and RCP 8.5.

    Returns dict with keys like '2040s RCP 4.5' -> delta value (float, deg C).
    Also prints a summary table to stdout.
    """
    import pandas as pd

    dfx = pd.read_csv(tmax_path, skiprows=header_rows, header=None, on_bad_lines='skip')
    dfn = pd.read_csv(tmin_path, skiprows=header_rows, header=None, on_bad_lines='skip')

    years = dfx.iloc[:, 0].astype(int).values

    # Each file has 40 model columns: cols 2-21 = RCP 4.5, cols 22-41 = RCP 8.5
    tmax_45 = dfx.iloc[:, 2:22].astype(float).values
    tmax_85 = dfx.iloc[:, 22:42].astype(float).values
    tmin_45 = dfn.iloc[:, 2:22].astype(float).values
    tmin_85 = dfn.iloc[:, 22:42].astype(float).values

    tmean_45 = (tmax_45 + tmin_45) / 2
    tmean_85 = (tmax_85 + tmin_85) / 2

    ens_45 = np.nanmean(tmean_45, axis=1)
    ens_85 = np.nanmean(tmean_85, axis=1)

    hist = (years >= 1970) & (years <= 1999)
    hist_mean = ens_45[hist].mean()

    windows = {
        '2040s': (years >= 2030) & (years <= 2059),
        '2080s': (years >= 2070) & (years <= 2099),
    }

    deltas = {}
    print(f"\n{'='*72}")
    print(f"  CMIP5 JJA Air Temperature Deltas — 20-Model MACAv2-METDATA Ensemble")
    print(f"  Grid cell: 47.619 N, 119.899 W  |  Baseline: 1970-1999 = {hist_mean:.2f} C")
    print(f"{'='*72}")
    print(f"  {'Period':<22} {'RCP 4.5':>10} {'RCP 8.5':>10}")
    print(f"  {'-'*44}")
    for period, mask in windows.items():
        d45 = ens_45[mask].mean() - hist_mean
        d85 = ens_85[mask].mean() - hist_mean
        deltas[f'{period} RCP 4.5'] = round(d45, 2)
        deltas[f'{period} RCP 8.5'] = round(d85, 2)
        print(f"  {period:<22} {d45:>+10.2f} {d85:>+10.2f}")
    print(f"{'='*72}\n")

    return deltas

# --- Compute deltas if MACA files exist, otherwise use pre-computed values ---
_maca_files_exist = os.path.isfile(MACA_TMAX_FILE) and os.path.isfile(MACA_TMIN_FILE)

if _maca_files_exist:
    _deltas = compute_cmip5_deltas(MACA_TMAX_FILE, MACA_TMIN_FILE)
else:
    # Pre-computed deltas (from MACAv2-METDATA, computed 2026-03-02)
    # These are the 20-model ensemble means; recompute by placing the
    # Climate Toolbox CSV files at the paths defined above.
    _deltas = {
        '2040s RCP 4.5': 2.36,
        '2040s RCP 8.5': 2.97,
        '2080s RCP 4.5': 3.46,
        '2080s RCP 8.5': 6.11,
    }
    print("  [INFO] MACA CSV files not found — using pre-computed CMIP5 deltas.")

# ── Analysis period — determined from loaded data (updated in main) ────────
# These are placeholder defaults; main() overwrites them after loading data.
ANALYSIS_START_YEAR = 1995   # earliest year in any station file
ANALYSIS_END_YEAR   = 2025

CLIMATE_SCENARIOS = [
    ('Baseline',              0.00,                     '#2980b9'),
    ('2040s RCP 4.5',        _deltas['2040s RCP 4.5'],  '#f57c00'),
    ('2040s RCP 8.5',        _deltas['2040s RCP 8.5'],  '#d32f2f'),
    ('2080s RCP 4.5',        _deltas['2080s RCP 4.5'],  '#e65100'),
    ('2080s RCP 8.5',        _deltas['2080s RCP 8.5'],  '#b71c1c'),
]

# Color palette
C = {
    'FDRW':  '#1a6faf',
    'GCGW':  '#2d9e6b',
    'CHJ':   '#c0392b',      # Primary analysis station (red)
    'CHQW':  '#e07b2a',      # Included for comparison (orange)
    'air':   '#636e72',
    'trend': '#2c3e50',
}

# ─────────────────────────────────────────────────────────────────────────────
# MATPLOTLIB STYLE
# ─────────────────────────────────────────────────────────────────────────────

plt.rcParams.update({
    'font.family':       'DejaVu Sans',
    'font.size':         12,
    'axes.spines.top':   False,
    'axes.spines.right': False,
    'axes.grid':         True,
    'grid.alpha':        0.2,
    'grid.linestyle':    '--',
    'axes.labelsize':    13,
    'axes.titlesize':    13,
    'axes.titleweight':  'bold',
    'legend.framealpha': 0.93,
    'legend.edgecolor':  '#cccccc',
    'legend.fontsize':   11,
    'xtick.labelsize':   11,
    'ytick.labelsize':   11,
    'figure.dpi':        150,
})

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1 — STATISTICAL HELPER FUNCTIONS
# ═════════════════════════════════════════════════════════════════════════════

def mann_kendall(y):
    """
    Mann-Kendall trend test.
    Reference: Hirsch et al. (1991); Helsel & Hirsch (2002)
    Returns: tau, p_value, S statistic, Z score, trend direction string
    """
    n   = len(y)
    if n < 4:
        return np.nan, np.nan, 0, 0.0, 'n/a'
    S = int(sum(np.sign(y[j] - y[i]) for i in range(n - 1) for j in range(i + 1, n)))
    var_S = n * (n - 1) * (2 * n + 5) / 18.0
    if   S > 0: Z = (S - 1) / np.sqrt(var_S)
    elif S < 0: Z = (S + 1) / np.sqrt(var_S)
    else:       Z = 0.0
    p_value = 2 * (1 - stats.norm.cdf(abs(Z)))
    tau     = S / (n * (n - 1) / 2.0)
    trend   = 'increasing' if S > 0 else ('decreasing' if S < 0 else 'no trend')
    return tau, p_value, S, Z, trend


def theil_sen(x, y):
    """
    Theil-Sen median slope estimator with 95% CI.
    Reference: Helsel & Hirsch (2002)
    Returns: slope, intercept, ci_low, ci_high
    """
    n      = len(x)
    slopes = [(y[j] - y[i]) / (x[j] - x[i])
              for i in range(n) for j in range(i + 1, n) if x[j] != x[i]]
    slopes = np.array(slopes)
    slope  = np.median(slopes)
    intercept = np.median(y) - slope * np.median(x)
    # 95% CI — normal approximation
    M       = len(slopes)
    C_alpha = 1.96 * np.sqrt(n * (n - 1) * (2 * n + 5) / 18.0)
    m1 = max(0, int((M - C_alpha) / 2))
    m2 = min(M - 1, int((M + C_alpha) / 2 + 1))
    ss  = np.sort(slopes)
    return slope, intercept, ss[m1], ss[m2]


def mohseni(Ta, mu, alpha, gamma, beta):
    """
    Mohseni et al. (1998) nonlinear logistic air-to-water temperature model.
        Tw = μ + (α − μ) / (1 + exp(γ(β − Ta)))
    Parameters:
        mu    — minimum stream temperature (°C)
        alpha — maximum stream temperature (°C)
        gamma — steepness of the S-curve
        beta  — air temperature at inflection point (°C)
        Ta    — weekly mean air temperature (°C)
    """
    return mu + (alpha - mu) / (1 + np.exp(gamma * (beta - Ta)))


def fit_mohseni(Ta, Tw):
    """
    Fit Mohseni model via least-squares with Nash-Sutcliffe Efficiency.
    Reference: Mohseni et al. (1998); Mantua et al. (2010)
    Returns: popt [mu, alpha, gamma, beta], NSE
    """
    Ta, Tw = np.asarray(Ta, float), np.asarray(Tw, float)
    p0  = [max(0.0, Tw.min() - 0.5), Tw.max() + 0.5, 0.2, float(np.median(Ta))]
    bds = ([0.0, p0[0] + 0.05, 0.01, -20.0],
           [Tw.max() + 2, 40.0, 2.0, 40.0])
    try:
        popt, _ = optimize.curve_fit(mohseni, Ta, Tw, p0=p0, bounds=bds, maxfev=12000)
    except Exception:
        print("  [WARNING] Mohseni fit did not converge; using linear fallback.")
        m, b, *_ = stats.linregress(Ta, Tw)
        popt = [max(0.0, b), b + m * 40, 0.15, 15.0]
    Tw_pred = mohseni(Ta, *popt)
    NSE = 1 - np.sum((Tw - Tw_pred) ** 2) / np.sum((Tw - Tw.mean()) ** 2)
    return popt, float(NSE)


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1B — QA/QC SAFEGUARDS FOR TEMPERATURE DATA
# ═════════════════════════════════════════════════════════════════════════════

class TemperatureQAQC:
    """
    Quality Assurance / Quality Control for water and air temperature data.
    Flags and removes problematic records to prevent analysis bias.
    
    QA/QC standards based on published literature:
    - USBR (2018): Thermal Regime of Columbia River at Lake Roosevelt
    - EPA (2002): Columbia River Temperature Assessment, Lake Roosevelt
    - O'Connor (2021): Water Temperature Trends in Columbia River Basin
    - Helsel & Hirsch (2002): Statistical Methods in Water Resources
    - EPA/Tetra Tech (2021): RBM10 Model Temperature TMDL
    
    Physical bounds are set conservatively to allow for natural extremes
    while flagging sensor errors and physically impossible values.
    """

    # Physical bounds for Columbia River water temp (°C), by month
    # Based on USBR (2018) observed range (3-19°C) with buffer for climate extremes
    # Reference: Median temps at Grand Coulee/Chief Joseph 2000-2015
    WATER_BOUNDS = {
        1: (0.0, 10),    2: (0.0, 10),    3: (0.5, 12),
        4: (2, 14),      5: (5, 16),      6: (8, 22),
        7: (10, 25),     8: (10, 25),     9: (8, 24),
        10: (4, 20),     11: (2, 14),     12: (0.5, 10),
    }
    # Air temperature bounds (°C), by month; for Douglas, WA area (Eastern WA)
    # Conservative bounds for continental climate with cold winters, hot summers
    AIR_BOUNDS = {
        1: (-20, 10),    2: (-20, 12),    3: (-10, 18),
        4: (-5, 25),     5: (0, 32),      6: (5, 38),
        7: (8, 42),      8: (8, 42),      9: (2, 35),
        10: (-5, 28),    11: (-12, 18),   12: (-20, 12),
    }

    @staticmethod
    def check_physical_bounds(df, temp_col, bounds_dict, source='water'):
        """Flag records falling outside monthly physical bounds."""
        flags = []
        for idx, row in df.iterrows():
            month = row['month']
            if month not in bounds_dict:
                continue
            lo, hi = bounds_dict[month]
            if not (lo <= row[temp_col] <= hi):
                flags.append({'idx': idx, 'issue': f'{temp_col}={row[temp_col]:.2f}°C outside bounds'})
        return pd.DataFrame(flags) if flags else pd.DataFrame(columns=['idx', 'issue']), len(flags)

    @staticmethod
    def detect_outliers_iqr(series, iqr_multiplier=3.0):
        """Flag outliers using IQR method. Returns (bad_indices, count, (lo_bound, hi_bound))."""
        Q1, Q3 = series.quantile(0.25), series.quantile(0.75)
        IQR = Q3 - Q1
        lo = Q1 - iqr_multiplier * IQR
        hi = Q3 + iqr_multiplier * IQR
        bad = series[(series < lo) | (series > hi)].index.tolist()
        return bad, len(bad), (lo, hi)

    @staticmethod
    def detect_constant_spans(df, temp_col, min_span_hours=24, tolerance=0.05):
        """Flag long constant-temperature spans (sensor failure indicator)."""
        flags, count = [], 0
        df_sorted = df.sort_values('datetime')
        i = 0
        sorted_idx = df_sorted.index.tolist()
        while i < len(df_sorted):
            j = i
            while j < len(df_sorted) and abs(df_sorted.iloc[j][temp_col] - df_sorted.iloc[i][temp_col]) < tolerance:
                j += 1
            if j - i >= min_span_hours:
                flags.append({'duration_hours': j - i, 'note': 'Possible sensor failure'})
                count = len(flags)
            i = j
        return pd.DataFrame(flags) if flags else pd.DataFrame(), count

    @staticmethod
    def detect_impossible_jumps(df, temp_col, max_hourly_change=3.0):
        """
        Flag impossible temperature jumps.
        EPA (2002) standard: >3°C flagged as suspect in large reservoirs.
        Reference: EPA Columbia River Temperature Assessment, Lake Roosevelt thermal regime study.
        """
        df_sorted = df.sort_values('datetime')
        flags, idx_list = [], []
        for i in range(1, len(df_sorted)):
            curr_idx = df_sorted.index[i]
            prev_idx = df_sorted.index[i-1]
            if pd.notna(df_sorted.loc[curr_idx, temp_col]) and pd.notna(df_sorted.loc[prev_idx, temp_col]):
                dt = (df_sorted.loc[curr_idx, 'datetime'] - df_sorted.loc[prev_idx, 'datetime']).total_seconds() / 3600
                dtemp = abs(df_sorted.loc[curr_idx, temp_col] - df_sorted.loc[prev_idx, temp_col])
                if dt > 0 and dtemp / dt > max_hourly_change:
                    flags.append({'idx': curr_idx, 'dtemp': dtemp, 'dt_hr': dt})
                    idx_list.append(curr_idx)
        return pd.DataFrame(flags) if flags else pd.DataFrame(), len(flags)

    @staticmethod
    def detect_data_gaps(df, freq='D'):
        """Identify data gaps exceeding specified frequency.

        Works for both hourly water data ("datetime" column)
        and daily air temperature data ("date" column).
        """
        # Determine the time column to use
        if 'datetime' in df.columns:
            time_col = 'datetime'
        elif 'date' in df.columns:
            time_col = 'date'
        else:
            raise KeyError("detect_data_gaps: expected 'datetime' or 'date' column")

        df_sorted = df.sort_values(time_col)
        expected_freq = pd.to_timedelta(f'1{freq}')
        gaps = []
        for i in range(1, len(df_sorted)):
            t_prev = df_sorted.iloc[i-1][time_col]
            t_curr = df_sorted.iloc[i][time_col]
            actual_gap = t_curr - t_prev
            if actual_gap > expected_freq:
                duration_days = actual_gap.total_seconds() / 86400.0
                gaps.append({
                    'gap_start': t_prev,
                    'gap_end': t_curr,
                    'duration_days': duration_days,
                    'issue': f'Gap: {duration_days:.1f} days'
                })
        return pd.DataFrame(gaps) if gaps else pd.DataFrame(), len(gaps)

    @staticmethod
    def qaqc_report(source, report_dict):
        """Print QA/QC summary."""
        print(f"  QA/QC Report: {source}")
        print("  " + "─" * 70)
        for check_name, (count, _) in report_dict.items():
            status = "[!] FLAGGED" if count > 0 else "[OK] PASS"
            print(f"    {check_name:<40} {count:>8} records  {status}")
        print("  " + "─" * 70)


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 2 — DATA LOADING
# ═════════════════════════════════════════════════════════════════════════════

def load_station(name, path):
    """
    Load multi-sheet hourly water temperature XLSX (one sheet per year)
    and return a tidy daily-resolution DataFrame.
    Includes QA/QC checks to flag/remove problematic records.
    """
    print(f"  Loading {name} …")
    xl = pd.ExcelFile(path)
    df = pd.concat([xl.parse(s) for s in xl.sheet_names], ignore_index=True)
    df['datetime'] = pd.to_datetime(df['Pacific Timestamp'], errors='coerce')
    df = df.dropna(subset=['datetime', 'Temperature (C)']).copy()
    df.rename(columns={
        'Temperature (C)': 'wtc',
        'Temperature (F)': 'wtf',
        'Outflow (kcfs)':  'flow_kcfs'
    }, inplace=True)
    df['year']  = df['datetime'].dt.year
    df['month'] = df['datetime'].dt.month
    df['doy']   = df['datetime'].dt.dayofyear
    df['week']  = df['datetime'].dt.isocalendar().week.astype(int)
    df['station'] = name
    df = df.sort_values('datetime').reset_index(drop=True)
    
    # ──── QA/QC CHECKS ────────────────────────────────────────────────────────
    qaqc = TemperatureQAQC()
    qaqc_report = {}
    
    # Check 1: Physical bounds
    bounds_flags, bounds_count = qaqc.check_physical_bounds(df, 'wtc', qaqc.WATER_BOUNDS, source='water')
    qaqc_report['Physical Bounds'] = (bounds_count, bounds_flags)
    if bounds_count > 0:
        print(f"    [!] {bounds_count} records flagged: outside physical temperature bounds")
        df = df.drop(bounds_flags['idx'].values)
        df = df.reset_index(drop=True)
    
    # Check 2: Outliers (IQR method)
    outlier_idx, outlier_count, (lo_bound, hi_bound) = qaqc.detect_outliers_iqr(df['wtc'], iqr_multiplier=3.0)
    qaqc_report['IQR Outliers (3.0×IQR)'] = (outlier_count, None)
    if outlier_count > 0:
        print(f"    [!] {outlier_count} records flagged as statistical outliers (>3×IQR)")
        df = df.drop(outlier_idx)
        df = df.reset_index(drop=True)
    
    # Check 3: Constant spans (sensor failure indicator)
    const_flags, const_count = qaqc.detect_constant_spans(df, 'wtc', min_span_hours=24, tolerance=0.05)
    qaqc_report['Constant Spans (24h+)'] = (const_count, const_flags)
    if const_count > 0:
        print(f"    [!] {const_count} long constant-value spans detected (possible sensor issues)")
    
    # Check 4: Impossible jumps (EPA 2002 standard: >3°C flagged as suspect)
    jump_flags, jump_count = qaqc.detect_impossible_jumps(df, 'wtc', max_hourly_change=3.0)
    qaqc_report['Impossible Jumps (>3°C/hr)'] = (jump_count, jump_flags)
    if jump_count > 0:
        print(f"    [!] {jump_count} impossible jumps detected (>3°C/hour)")
        df = df.drop(jump_flags['idx'].values)
        df = df.reset_index(drop=True)
    
    # Check 5: Data gaps
    gap_flags, gap_count = qaqc.detect_data_gaps(df, freq='D')
    qaqc_report['Data Gaps'] = (gap_count, gap_flags)
    if gap_count > 0:
        print(f"    [!] {gap_count} data gap(s) detected")
    
    # Print summary
    qaqc.qaqc_report(name, qaqc_report)
    
    # Final count
    df = df.sort_values('datetime').reset_index(drop=True)
    print(f"    → {len(df):,} hourly records after QA/QC  |  {df.year.min()}–{df.year.max()}")
    return df


def load_air(path):
    """
    Load NOAA daily air temperature CSV and convert °F → °C.
    Includes QA/QC checks to flag/remove problematic records.
    """
    print(f"  Loading air temperature …")
    df = pd.read_csv(path, parse_dates=['DATE'])
    df.rename(columns={'DATE': 'date', 'TMAX': 'tmax_f',
                       'TMIN': 'tmin_f', 'TAVG': 'tavg_f'}, inplace=True)
    df['tmax_c']  = (df['tmax_f'] - 32) * 5 / 9
    df['tmin_c']  = (df['tmin_f'] - 32) * 5 / 9
    df['tavg_f_raw'] = df['tavg_f'].copy()
    df['tavg_c']  = ((df['tmax_c'] + df['tmin_c']) / 2)   # always use mean of max/min
    df['year']    = df['date'].dt.year
    df['month']   = df['date'].dt.month
    df['doy']     = df['date'].dt.dayofyear
    df['week']    = df['date'].dt.isocalendar().week.astype(int)
    
    # ──── QA/QC CHECKS ────────────────────────────────────────────────────────
    qaqc = TemperatureQAQC()
    qaqc_report = {}
    
    # Check 1: Physical bounds
    bounds_flags, bounds_count = qaqc.check_physical_bounds(df, 'tavg_c', qaqc.AIR_BOUNDS, source='air')
    qaqc_report['Physical Bounds'] = (bounds_count, bounds_flags)
    if bounds_count > 0:
        print(f"    [!] {bounds_count} records flagged: outside physical temperature bounds")
        df = df.drop(bounds_flags['idx'].values)
        df = df.reset_index(drop=True)
    
    # Check 2: Outliers (IQR method)
    outlier_idx, outlier_count, (lo_bound, hi_bound) = qaqc.detect_outliers_iqr(df['tavg_c'], iqr_multiplier=3.0)
    qaqc_report['IQR Outliers (3.0×IQR)'] = (outlier_count, None)
    if outlier_count > 0:
        print(f"    [!] {outlier_count} records flagged as statistical outliers (>3×IQR)")
        df = df.drop(outlier_idx)
        df = df.reset_index(drop=True)
    
    # Check 3: Data gaps
    gap_flags, gap_count = qaqc.detect_data_gaps(df, freq='D')
    qaqc_report['Data Gaps'] = (gap_count, gap_flags)
    if gap_count > 0:
        print(f"    [!] {gap_count} data gap(s) detected")
        if len(gap_flags) > 0:
            for col_name in ['gap_start', 'gap_end', 'duration_days']:
                pass  # Just print the first few
            print(f"      {gap_flags.iloc[0]['issue']}")
    
    # Print summary
    qaqc.qaqc_report('Air Temperature (Douglas, WA)', qaqc_report)
    
    # Final count
    df = df.sort_values('date').reset_index(drop=True)
    print(f"    → {len(df):,} daily records after QA/QC  |  {df.year.min()}–{df.year.max()}")
    return df


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 3 — ANALYSIS
# ═════════════════════════════════════════════════════════════════════════════

def run_analysis(stations, air):
    """
    Execute all statistical analyses and return result DataFrames.
    """
    print("\n" + "=" * 70)
    print("STATISTICAL ANALYSIS")
    print("=" * 70)

    # ── Annual means: combined Jul-Aug AND per-month ────────────────────────
    def ann_summer(df):
        return df[df['month'].isin(SUMMER_MONTHS)].groupby('year')['wtc'].mean()

    def ann_month(df, month):
        return df[df['month'] == month].groupby('year')['wtc'].mean()

    # ── Build weekly air-water pairs for Mohseni model (CHJ) ─────────────
    chj = stations['CHJ']
    chj_wk = (chj.groupby(['year', 'week'])['wtc'].mean().reset_index()
              .assign(yw=lambda d: d['year'].astype(str) + '_' + d['week'].astype(str).str.zfill(2)))
    air_wk  = (air.groupby(['year', 'week'])['tavg_c'].mean().reset_index()
               .assign(yw=lambda d: d['year'].astype(str) + '_' + d['week'].astype(str).str.zfill(2)))
    merged  = chj_wk.merge(air_wk[['yw', 'tavg_c']], on='yw').dropna()

    # ── Mohseni model fits ────────────────────────────────────────────────────
    popt_all, NSE_all = fit_mohseni(merged['tavg_c'].values, merged['wtc'].values)
    jja_mask          = merged['week'].between(26, 35)
    popt_jja, NSE_jja = fit_mohseni(merged.loc[jja_mask, 'tavg_c'].values,
                                    merged.loc[jja_mask, 'wtc'].values)

    print(f"\n  Mohseni Model Parameters — CHJ (All Weeks)")
    print(f"    μ = {popt_all[0]:.3f} °C,  α = {popt_all[1]:.3f} °C,"
          f"  γ = {popt_all[2]:.4f},  β = {popt_all[3]:.3f} °C")
    print(f"    NSE = {NSE_all:.4f}  {'(Good ≥0.7)' if NSE_all >= 0.7 else '(Moderate)'}")
    print(f"\n  Mohseni Model Parameters — CHJ (Jul–Aug Only)")
    print(f"    μ = {popt_jja[0]:.3f} °C,  α = {popt_jja[1]:.3f} °C,"
          f"  γ = {popt_jja[2]:.4f},  β = {popt_jja[3]:.3f} °C")
    print(f"    NSE = {NSE_jja:.4f}")

    # ── Trend analysis — all stations ────────────────────────────────────────
    print(f"\n  {'Station':<34} {'n':>4} {'Mean JA °C':>11} {'Slope °C/yr':>13}"
          f" {'MK τ':>8} {'p':>8} {'Sig':>5}")
    print("  " + "-" * 90)

    summary_rows = []
    for stn in STATION_ORDER:
        ann  = ann_summer(stations[stn])
        x, y = ann.index.values.astype(float), ann.values
        tau, p, S, Z, direction = mann_kendall(y)
        sl, ic, ci_lo, ci_hi   = theil_sen(x, y)
        sig = '★' if p < 0.05 else ''
        print(f"  {STATION_LABELS[stn]:<34} {len(y):>4} {y.mean():>11.2f} {sl:>+13.4f}"
              f" {tau:>8.4f} {p:>8.4f} {sig:>5}")
        summary_rows.append({
            'Station': STATION_LABELS[stn], 'Code': stn,
            'n_years': len(y),
            'Mean_JulAug_C': round(y.mean(), 3),
            'Mean_JulAug_F': round(y.mean() * 9/5 + 32, 3),
            'TheilSen_slope_C_yr': round(sl, 5),
            'TheilSen_CI_low':  round(ci_lo, 5),
            'TheilSen_CI_high': round(ci_hi, 5),
            'MK_tau': round(tau, 4),
            'MK_S': S, 'MK_Z': round(Z, 4),
            'MK_p': round(p, 4),
            'Trend': direction,
            'Significant_p05': 'Yes' if p < 0.05 else 'No',
        })
    summary_df = pd.DataFrame(summary_rows)

    # ── Month-specific trend analysis — all stations ──────────────────────
    print(f"\n  Month-Specific Trends:")
    print(f"  {'Station':<34} {'Month':>5} {'n':>4} {'Mean °C':>9} {'Slope °C/yr':>13}"
          f" {'MK τ':>8} {'p':>8} {'Sig':>5}")
    print("  " + "-" * 95)

    monthly_summary_rows = []
    for stn in STATION_ORDER:
        for month, month_name in [(7, 'July'), (8, 'August')]:
            ann = ann_month(stations[stn], month)
            x, y = ann.index.values.astype(float), ann.values
            if len(y) < 4:
                continue
            tau, p, S, Z, direction = mann_kendall(y)
            sl, ic, ci_lo, ci_hi = theil_sen(x, y)
            sig = '★' if p < 0.05 else ''
            print(f"  {STATION_LABELS[stn]:<34} {month_name:>5} {len(y):>4} {y.mean():>9.2f} {sl:>+13.4f}"
                  f" {tau:>8.4f} {p:>8.4f} {sig:>5}")
            monthly_summary_rows.append({
                'Station': STATION_LABELS[stn], 'Code': stn,
                'Month': month_name,
                'n_years': len(y),
                'Mean_C': round(y.mean(), 3),
                'Mean_F': round(y.mean() * 9/5 + 32, 3),
                'TheilSen_slope_C_yr': round(sl, 5),
                'TheilSen_CI_low': round(ci_lo, 5),
                'TheilSen_CI_high': round(ci_hi, 5),
                'MK_tau': round(tau, 4),
                'MK_S': S, 'MK_Z': round(Z, 4),
                'MK_p': round(p, 4),
                'Trend': direction,
                'Significant_p05': 'Yes' if p < 0.05 else 'No',
            })
    monthly_summary_df = pd.DataFrame(monthly_summary_rows)

    # ── CHJ-specific stats for figures ──────────────────────────────────────
    ann_chj     = ann_summer(stations['CHJ'])
    x_c, y_c     = ann_chj.index.values.astype(float), ann_chj.values
    tau_c, p_c, S_c, Z_c, dir_c = mann_kendall(y_c)
    sl_c, ic_c, ci_lo_c, ci_hi_c = theil_sen(x_c, y_c)

    # CHJ month-specific annual means and trends
    ann_chj_jul = ann_month(stations['CHJ'], 7)
    ann_chj_aug = ann_month(stations['CHJ'], 8)
    x_cj, y_cj = ann_chj_jul.index.values.astype(float), ann_chj_jul.values
    x_ca, y_ca = ann_chj_aug.index.values.astype(float), ann_chj_aug.values
    tau_cj, p_cj, *_ = mann_kendall(y_cj)
    sl_cj, ic_cj, ci_lo_cj, ci_hi_cj = theil_sen(x_cj, y_cj)
    tau_ca, p_ca, *_ = mann_kendall(y_ca)
    sl_ca, ic_ca, ci_lo_ca, ci_hi_ca = theil_sen(x_ca, y_ca)

    # Air temperature trend
    ann_air       = air[air['month'].isin(SUMMER_MONTHS)].groupby('year')['tavg_c'].mean()
    x_a, y_a      = ann_air.index.values.astype(float), ann_air.values
    tau_a, p_a, *_ = mann_kendall(y_a)
    sl_a, ic_a, *_ = theil_sen(x_a, y_a)

    # Air temperature month-specific trends
    ann_air_jul = air[air['month'] == 7].groupby('year')['tavg_c'].mean()
    ann_air_aug = air[air['month'] == 8].groupby('year')['tavg_c'].mean()
    x_aj, y_aj = ann_air_jul.index.values.astype(float), ann_air_jul.values
    x_aa, y_aa = ann_air_aug.index.values.astype(float), ann_air_aug.values
    tau_aj, p_aj, *_ = mann_kendall(y_aj)
    sl_aj, ic_aj, *_ = theil_sen(x_aj, y_aj)
    tau_aa, p_aa, *_ = mann_kendall(y_aa)
    sl_aa, ic_aa, *_ = theil_sen(x_aa, y_aa)

    # ── Build climate scenarios with actual baseline year range ─────────────
    baseline_yr_label = f"{int(chj.year.min())}–{int(chj.year.max())}"
    climate_scenarios = [
        (f'Baseline ({baseline_yr_label})', 0.00, '#2980b9'),
    ] + [(label, delta, color) for label, delta, color in CLIMATE_SCENARIOS[1:]]

    # ── Future projections ────────────────────────────────────────────────────
    baseline_air = air[air['month'].isin(SUMMER_MONTHS)]['tavg_c'].mean()
    baseline_wt  = stations['CHJ'][stations['CHJ']['month'].isin(SUMMER_MONTHS)]['wtc'].mean()

    baseline_yr_label = f"{int(chj.year.min())}–{int(chj.year.max())}"
    print(f"\n  Observed baseline ({baseline_yr_label}):")
    print(f"    Jul/Aug mean air temp  : {baseline_air:.2f} °C  ({baseline_air*9/5+32:.2f} °F)")
    print(f"    Jul/Aug mean water temp: {baseline_wt:.2f} °C  ({baseline_wt*9/5+32:.2f} °F)")
    print(f"\n  {'Scenario':<24} {'ΔTair':>8} {'Proj.Air':>10} {'Proj.Tw':>10}"
          f" {'Proj.Tw(°F)':>12} {'ΔTw':>8}")
    print("  " + "-" * 76)

    proj_rows = []
    # Compute model-predicted baseline Tw for consistent deltas
    baseline_tw_model = float(mohseni(baseline_air, *popt_all))
    for label, delta, color in climate_scenarios:
        proj_air = baseline_air + delta
        proj_tw  = mohseni(proj_air, *popt_all)
        dtw      = proj_tw - baseline_tw_model
        print(f"  {label:<24} {delta:>+8.2f} {proj_air:>10.2f} {proj_tw:>10.2f}"
              f" {proj_tw*9/5+32:>12.2f} {dtw:>+8.2f}")
        proj_rows.append({
            'Scenario': label, 'Delta_Air_C': round(delta, 2),
            'Proj_Air_C': round(proj_air, 2),
            'Proj_Tw_C':  round(proj_tw, 2),
            'Proj_Tw_F':  round(proj_tw * 9/5 + 32, 2),
            'Delta_Tw_C': round(dtw, 2),
            'Delta_Tw_F': round(dtw * 9/5, 2),
            'Color': color,
        })
    proj_df = pd.DataFrame(proj_rows)

    # ── Month-specific projections (July and August separately) ──────────────
    # Uses observed monthly water temperature as the baseline and applies only
    # the Mohseni-derived *change* (delta) for each climate scenario.  This
    # preserves the empirical fact that August is warmer than July at CHJ
    # (due to cumulative seasonal heat loading in Rufus Woods Lake) while
    # still using the Mohseni curve to estimate sensitivity to air warming.
    monthly_proj = {}
    for month, month_name in [(7, 'July'), (8, 'August')]:
        m_air_baseline = air[air['month'] == month]['tavg_c'].mean()
        m_wt_baseline  = stations['CHJ'][(stations['CHJ']['month'] == month)]['wtc'].mean()
        m_tw_model     = float(mohseni(m_air_baseline, *popt_all))

        m_rows = []
        print(f"\n  {month_name}-only projections (baseline air = {m_air_baseline:.2f} °C, "
              f"obs water = {m_wt_baseline:.2f} °C, model water = {m_tw_model:.2f} °C):")
        print(f"  {'Scenario':<24} {'ΔTair':>8} {'Proj.Air':>10} {'Proj.Tw':>10}"
              f" {'Proj.Tw(°F)':>12} {'ΔTw':>8}")
        print("  " + "-" * 76)
        for label, delta, color in climate_scenarios:
            proj_air = m_air_baseline + delta
            # Mohseni-derived delta: change in modeled Tw for the given air temp shift
            mohseni_dtw = float(mohseni(proj_air, *popt_all)) - m_tw_model
            # Apply that delta to the OBSERVED baseline water temp
            proj_tw = m_wt_baseline + mohseni_dtw
            print(f"  {label:<24} {delta:>+8.2f} {proj_air:>10.2f} {proj_tw:>10.2f}"
                  f" {proj_tw*9/5+32:>12.2f} {mohseni_dtw:>+8.2f}")
            m_rows.append({
                'Scenario': label, 'Month': month_name,
                'Delta_Air_C': round(delta, 2),
                'Baseline_Air_C': round(m_air_baseline, 2),
                'Proj_Air_C': round(proj_air, 2),
                'Obs_Baseline_Tw_C': round(m_wt_baseline, 2),
                'Proj_Tw_C': round(proj_tw, 2),
                'Proj_Tw_F': round(proj_tw * 9/5 + 32, 2),
                'Delta_Tw_C': round(mohseni_dtw, 2),
                'Delta_Tw_F': round(mohseni_dtw * 9/5, 2),
                'Color': color,
            })
        monthly_proj[month_name] = pd.DataFrame(m_rows)

    # ── Week-by-week water temperature stats for July and August (CHJ) ────
    chj_summer = stations['CHJ'][stations['CHJ']['month'].isin([7, 8])].copy()
    weekly_by_yr = chj_summer.groupby(['year', 'week', 'month'])['wtc'].mean().reset_index()

    # Stats per ISO week: mean/min/max across all years for each week
    week_stats_df = weekly_by_yr.groupby('week').agg(
        mean_tw=('wtc', 'mean'),
        min_tw=('wtc', 'min'),
        max_tw=('wtc', 'max'),
        n_years=('wtc', 'count'),
    ).reset_index()

    # Determine which month each week primarily falls in
    week_month = weekly_by_yr.groupby('week')['month'].agg(lambda x: x.mode()[0]).reset_index()
    week_month.columns = ['week', 'primary_month']
    week_stats_df = week_stats_df.merge(week_month, on='week')

    # Filter to Jul-Aug weeks only (roughly weeks 27–35)
    week_stats_df = week_stats_df[(week_stats_df['primary_month'].isin([7, 8])) &
                                  (week_stats_df['n_years'] >= 3)].sort_values('week').reset_index(drop=True)

    print(f"\n  Week-by-week water temp stats (CHJ, Jul-Aug weeks):")
    for _, r in week_stats_df.iterrows():
        m_name = 'Jul' if r['primary_month'] == 7 else 'Aug'
        print(f"    Week {int(r['week']):>2} ({m_name}): mean={r['mean_tw']:.2f}, "
              f"min={r['min_tw']:.2f}, max={r['max_tw']:.2f}, n={int(r['n_years'])}")

    # Package everything for figures and export
    results = dict(
        summary_df=summary_df, monthly_summary_df=monthly_summary_df,
        proj_df=proj_df,
        stations=stations, air=air, merged=merged,
        popt_all=popt_all, NSE_all=NSE_all,
        popt_jja=popt_jja, NSE_jja=NSE_jja,
        ann_chj=ann_chj,
        ann_chj_jul=ann_chj_jul, ann_chj_aug=ann_chj_aug,
        x_c=x_c, y_c=y_c,
        tau_c=tau_c, p_c=p_c, sl_c=sl_c, ic_c=ic_c,
        ci_lo_c=ci_lo_c, ci_hi_c=ci_hi_c,
        x_cj=x_cj, y_cj=y_cj, tau_cj=tau_cj, p_cj=p_cj,
        sl_cj=sl_cj, ic_cj=ic_cj, ci_lo_cj=ci_lo_cj, ci_hi_cj=ci_hi_cj,
        x_ca=x_ca, y_ca=y_ca, tau_ca=tau_ca, p_ca=p_ca,
        sl_ca=sl_ca, ic_ca=ic_ca, ci_lo_ca=ci_lo_ca, ci_hi_ca=ci_hi_ca,
        x_a=x_a, y_a=y_a, sl_a=sl_a, ic_a=ic_a, tau_a=tau_a, p_a=p_a,
        x_aj=x_aj, y_aj=y_aj, sl_aj=sl_aj, ic_aj=ic_aj, tau_aj=tau_aj, p_aj=p_aj,
        x_aa=x_aa, y_aa=y_aa, sl_aa=sl_aa, ic_aa=ic_aa, tau_aa=tau_aa, p_aa=p_aa,
        baseline_air=baseline_air, baseline_wt=baseline_wt,
        jja_mask=jja_mask,
        climate_scenarios=climate_scenarios,
        monthly_proj=monthly_proj,
        week_stats_df=week_stats_df,
    )
    return results


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 4 — EXCEL EXPORT
# ═════════════════════════════════════════════════════════════════════════════

def export_excel(results):
    """Export all results to a formatted multi-sheet Excel workbook."""
    print("\n  Exporting Excel results …")
    summary_df  = results['summary_df']
    proj_df     = results['proj_df']
    popt_all    = results['popt_all']
    NSE_all     = results['NSE_all']
    popt_jja    = results['popt_jja']
    NSE_jja     = results['NSE_jja']
    ann_chj     = results['ann_chj']
    baseline_air = results['baseline_air']
    baseline_wt  = results['baseline_wt']
    stations_data = results['stations']
    air_data     = results['air']

    # Determine actual data periods from loaded data
    chj_yr_range = f"{int(ann_chj.index.min())}–{int(ann_chj.index.max())}"
    water_yr_min = min(stations_data[s]['year'].min() for s in STATION_ORDER)
    water_yr_max = max(stations_data[s]['year'].max() for s in STATION_ORDER)
    water_yr_range = f"{water_yr_min}–{water_yr_max}"
    air_yr_range = f"{air_data['year'].min()}–{air_data['year'].max()}"

    # ── Shared styles ─────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    HDR_FONT  = Font(color='FFFFFF', bold=True)
    HDR_ALIGN = Alignment(horizontal='center')
    SEC_FILL  = PatternFill(start_color='1a4f72', end_color='1a4f72', fill_type='solid')
    SEC_FONT  = Font(color='FFFFFF', bold=True, size=10)
    CHJ_FILL  = PatternFill(start_color='fde8e8', end_color='fde8e8', fill_type='solid')  # Primary station highlight
    CHQW_FILL = PatternFill(start_color='fff4e6', end_color='fff4e6', fill_type='solid')  # Comparison station highlight
    BOLD      = Font(bold=True, size=10)
    NORMAL    = Font(size=10)
    ITALIC    = Font(italic=True, size=10)
    WRAP      = Alignment(wrap_text=True, vertical='top')
    CENTER_V  = Alignment(horizontal='center', vertical='top')

    def style_header(ws):
        for cell in ws[1]:
            cell.font, cell.fill, cell.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN

    def autofit(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = max(w + 2, 12)

    wb = Workbook()

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 0 — Data Sources & Methods
    # ═════════════════════════════════════════════════════════════════════════
    ws0 = wb.active
    ws0.title = 'Data Sources & Methods'

    def sec_header(ws, row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = SEC_FONT
        cell.fill = SEC_FILL
        cell.alignment = Alignment(vertical='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 18
        return row + 1

    def add_row(ws, row, label, value, note=None, lf=None, vf=None):
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        c1.font = lf or BOLD
        c2.font = vf or NORMAL
        c2.alignment = WRAP
        if note:
            c3 = ws.cell(row=row, column=3, value=note)
            c3.font = ITALIC
            c3.alignment = WRAP
        return row + 1

    ws0.column_dimensions['A'].width = 32
    ws0.column_dimensions['B'].width = 60
    ws0.column_dimensions['C'].width = 50

    r = 1
    r = sec_header(ws0, r, 'WATER TEMPERATURE DATA')
    r = add_row(ws0, r, 'Source', 'DART — Columbia Basin Research, University of Washington')
    r = add_row(ws0, r, 'URL', 'http://www.cbr.washington.edu/dart/query/river_graph_wmq')
    r = add_row(ws0, r, 'Stations', 'FDRW (GC Forebay), GCGW (GC Tailrace), CHJ (CJ Forebay - PRIMARY), CHQW (CJ Tailrace - comparison)')
    r = add_row(ws0, r, 'Period', f'{water_yr_range}, hourly resolution')
    r += 1
    r = sec_header(ws0, r, 'AIR TEMPERATURE DATA')
    r = add_row(ws0, r, 'Source', 'NOAA GHCND — Station USR0000WDOU (Douglas, WA)')
    r = add_row(ws0, r, 'Location', '47.619°N, 119.899°W  (~10 mi from CJD)')
    r = add_row(ws0, r, 'Period', f'{air_yr_range}, daily TMAX/TMIN → mean')
    r += 1
    r = sec_header(ws0, r, 'STATISTICAL METHODS')
    r = add_row(ws0, r, 'Trend detection', 'Mann-Kendall test (Hirsch et al. 1991; Helsel & Hirsch 2002)')
    r = add_row(ws0, r, 'Trend magnitude', 'Theil-Sen median slope with 95% CI (Helsel & Hirsch 2002)')
    r = add_row(ws0, r, 'Air-water model', 'Mohseni et al. (1998) nonlinear logistic regression')
    r = add_row(ws0, r, 'Goodness of fit', 'Nash-Sutcliffe Efficiency (Nash & Sutcliffe 1970)')
    r += 1
    r = sec_header(ws0, r, 'CLIMATE PROJECTIONS')
    r = add_row(ws0, r, 'Source', 'MACAv2-METDATA (Abatzoglou & Brown 2012), 20 CMIP5 GCMs via Climate Toolbox')
    r = add_row(ws0, r, 'Scenarios', 'RCP 4.5 (moderate) and RCP 8.5 (high emissions); 2040s and 2080s horizons')
    r = add_row(ws0, r, 'Note', f'Deltas referenced to 1970–1999 baseline; applied to {water_yr_range} observed baseline (conservative)')

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Trend Summary (combined Jul-Aug, retained for reference)
    # ═════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet('Trend Summary (Jul-Aug)')
    cols = list(summary_df.columns)
    ws1.append(cols)
    style_header(ws1)
    for _, row_data in summary_df.iterrows():
        ws1.append(list(row_data))
        if row_data['Code'] == 'CHJ':
            for cell in ws1[ws1.max_row]:
                cell.fill = CHJ_FILL
        elif row_data['Code'] == 'CHQW':
            for cell in ws1[ws1.max_row]:
                cell.fill = CHQW_FILL
    autofit(ws1)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Monthly Trend Summary (July and August separate)
    # ═════════════════════════════════════════════════════════════════════════
    monthly_summary_df = results['monthly_summary_df']
    ws1m = wb.create_sheet('Trend Summary (Monthly)')
    mcols = list(monthly_summary_df.columns)
    ws1m.append(mcols)
    style_header(ws1m)
    for _, row_data in monthly_summary_df.iterrows():
        ws1m.append(list(row_data))
        if row_data['Code'] == 'CHJ':
            for cell in ws1m[ws1m.max_row]:
                cell.fill = CHJ_FILL
        elif row_data['Code'] == 'CHQW':
            for cell in ws1m[ws1m.max_row]:
                cell.fill = CHQW_FILL
    autofit(ws1m)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Monthly Projections (July and August separate) — PRIMARY
    # ═════════════════════════════════════════════════════════════════════════
    monthly_proj = results.get('monthly_proj', {})
    if monthly_proj:
        ws5 = wb.create_sheet('Monthly Projections')
        ws5.append(['Month', 'Scenario', 'Baseline_Air_C', 'Delta_Air_C',
                     'Proj_Air_C', 'Obs_Baseline_Tw_C', 'Proj_Tw_C', 'Proj_Tw_F',
                     'Delta_Tw_C', 'Delta_Tw_F'])
        style_header(ws5)
        for month_name in ['July', 'August']:
            mdf = monthly_proj[month_name]
            for _, row in mdf.iterrows():
                ws5.append([month_name, row['Scenario'],
                           row['Baseline_Air_C'], row['Delta_Air_C'],
                           row['Proj_Air_C'], row['Obs_Baseline_Tw_C'],
                           row['Proj_Tw_C'], row['Proj_Tw_F'],
                           row['Delta_Tw_C'], row['Delta_Tw_F']])
        autofit(ws5)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Projections (combined Jul-Aug, retained for reference)
    # ═════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Projections (Jul-Aug)')
    pcols = [c for c in proj_df.columns if c != 'Color']
    ws2.append(pcols)
    style_header(ws2)
    for _, row_data in proj_df.iterrows():
        ws2.append([row_data[c] for c in pcols])
    autofit(ws2)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 5 — CHJ Annual (July, August, and Jul-Aug combined)
    # ═════════════════════════════════════════════════════════════════════════
    ann_chj_jul = results['ann_chj_jul']
    ann_chj_aug = results['ann_chj_aug']
    ws3 = wb.create_sheet('CHJ Annual')
    ws3.append(['Year', 'Mean_Jul_Tw_C', 'Mean_Jul_Tw_F',
                'Mean_Aug_Tw_C', 'Mean_Aug_Tw_F',
                'Mean_JulAug_Tw_C', 'Mean_JulAug_Tw_F'])
    style_header(ws3)
    all_yrs = sorted(set(ann_chj.index) | set(ann_chj_jul.index) | set(ann_chj_aug.index))
    for yr in all_yrs:
        jul_val = ann_chj_jul.get(yr, None)
        aug_val = ann_chj_aug.get(yr, None)
        comb_val = ann_chj.get(yr, None)
        ws3.append([
            int(yr),
            round(jul_val, 3) if jul_val is not None else None,
            round(jul_val * 9/5 + 32, 3) if jul_val is not None else None,
            round(aug_val, 3) if aug_val is not None else None,
            round(aug_val * 9/5 + 32, 3) if aug_val is not None else None,
            round(comb_val, 3) if comb_val is not None else None,
            round(comb_val * 9/5 + 32, 3) if comb_val is not None else None,
        ])
    autofit(ws3)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Mohseni Parameters
    # ═════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Mohseni Parameters')
    ws4.append(['Fit', 'mu_C', 'alpha_C', 'gamma', 'beta_C', 'NSE', 'Notes'])
    style_header(ws4)
    ws4.append(['All weeks']    + [round(v, 5) for v in popt_all] + [round(NSE_all, 5), 'Fit to all weekly pairs'])
    ws4.append(['Jul–Aug only'] + [round(v, 5) for v in popt_jja] + [round(NSE_jja, 5), 'Fit to weeks 26–35 only'])
    autofit(ws4)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET — Weekly Temperature Stats (Observed & Projected, week by week)
    # ═════════════════════════════════════════════════════════════════════════
    week_stats_df = results.get('week_stats_df', None)
    monthly_proj = results.get('monthly_proj', {})
    if week_stats_df is not None and len(week_stats_df) > 0:
        ws_wk = wb.create_sheet('Weekly Stats')
        ws_wk.append(['Scenario', 'Week', 'Month', 'Mean_Tw_C', 'Min_Tw_C', 'Max_Tw_C',
                       'Mean_Tw_F', 'Min_Tw_F', 'Max_Tw_F', 'n_years', 'Delta_Tw_C'])
        style_header(ws_wk)

        # Observed baseline rows
        for _, r in week_stats_df.iterrows():
            m_name = 'July' if r['primary_month'] == 7 else 'August'
            ws_wk.append(['Observed Baseline', int(r['week']), m_name,
                         round(r['mean_tw'], 3), round(r['min_tw'], 3), round(r['max_tw'], 3),
                         round(r['mean_tw'] * 9/5 + 32, 3), round(r['min_tw'] * 9/5 + 32, 3),
                         round(r['max_tw'] * 9/5 + 32, 3), int(r['n_years']), 0.0])

        # Projected rows — apply monthly deltas to each week based on its primary month
        if monthly_proj:
            for month_name in ['July', 'August']:
                mdf = monthly_proj[month_name]
                scen_rows = mdf[mdf['Delta_Air_C'] > 0]
                month_num = 7 if month_name == 'July' else 8
                month_weeks = week_stats_df[week_stats_df['primary_month'] == month_num]
                for _, sr in scen_rows.iterrows():
                    dtw = sr['Delta_Tw_C']
                    for _, wr in month_weeks.iterrows():
                        ws_wk.append([sr['Scenario'], int(wr['week']), month_name,
                                     round(wr['mean_tw'] + dtw, 3),
                                     round(wr['min_tw'] + dtw, 3),
                                     round(wr['max_tw'] + dtw, 3),
                                     round((wr['mean_tw'] + dtw) * 9/5 + 32, 3),
                                     round((wr['min_tw'] + dtw) * 9/5 + 32, 3),
                                     round((wr['max_tw'] + dtw) * 9/5 + 32, 3),
                                     int(wr['n_years']), round(dtw, 3)])
        autofit(ws_wk)

    path = f'{OUTPUT_DIR}/CJD_Temperature_Results.xlsx'
    wb.save(path)
    print(f"    → Saved {path}")


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 5 — FIGURES
# ═════════════════════════════════════════════════════════════════════════════

def make_figures(results):
    """Generate all 11 publication-quality figures."""
    print("\n" + "=" * 70)
    print("GENERATING FIGURES")
    print("=" * 70)

    stations    = results['stations']
    air         = results['air']
    merged      = results['merged']
    popt_all    = results['popt_all']
    NSE_all     = results['NSE_all']
    popt_jja    = results['popt_jja']
    NSE_jja     = results['NSE_jja']
    x_c, y_c    = results['x_c'], results['y_c']
    tau_c, p_c  = results['tau_c'], results['p_c']
    sl_c, ic_c  = results['sl_c'], results['ic_c']
    ci_lo_c, ci_hi_c = results['ci_lo_c'], results['ci_hi_c']
    # CHJ monthly
    x_cj, y_cj  = results['x_cj'], results['y_cj']
    tau_cj, p_cj = results['tau_cj'], results['p_cj']
    sl_cj, ic_cj = results['sl_cj'], results['ic_cj']
    ci_lo_cj, ci_hi_cj = results['ci_lo_cj'], results['ci_hi_cj']
    x_ca, y_ca  = results['x_ca'], results['y_ca']
    tau_ca, p_ca = results['tau_ca'], results['p_ca']
    sl_ca, ic_ca = results['sl_ca'], results['ic_ca']
    ci_lo_ca, ci_hi_ca = results['ci_lo_ca'], results['ci_hi_ca']
    # Air combined and monthly
    x_a, y_a    = results['x_a'], results['y_a']
    sl_a, ic_a  = results['sl_a'], results['ic_a']
    tau_a, p_a  = results['tau_a'], results['p_a']
    x_aj, y_aj  = results['x_aj'], results['y_aj']
    sl_aj, ic_aj = results['sl_aj'], results['ic_aj']
    tau_aj, p_aj = results['tau_aj'], results['p_aj']
    x_aa, y_aa  = results['x_aa'], results['y_aa']
    sl_aa, ic_aa = results['sl_aa'], results['ic_aa']
    tau_aa, p_aa = results['tau_aa'], results['p_aa']

    baseline_air = results['baseline_air']
    baseline_wt  = results['baseline_wt']
    proj_df      = results['proj_df']
    jja_mask     = results['jja_mask']
    climate_scenarios = results['climate_scenarios']

    def ann_summer(df):
        return df[df['month'].isin(SUMMER_MONTHS)].groupby('year')['wtc'].mean()

    def ann_month(df, month):
        return df[df['month'] == month].groupby('year')['wtc'].mean()

    # Shared data source footnote — built from actual loaded data ranges
    chj_yrs = stations['CHJ']['year']
    air_yrs  = air['year']
    DATA_FOOTNOTE = (
        f'Water temperature: DART (Columbia Basin Research, Univ. of Washington) — '
        f'CHJ station, hourly, {chj_yrs.min()}–{chj_yrs.max()}.  '
        f'Air temperature: NOAA GHCND:USR0000WDOU (Douglas, WA), daily, {air_yrs.min()}–{air_yrs.max()}.  '
        f'Climate deltas: MACAv2-METDATA 20-model CMIP5 ensemble (RCP 4.5 / 8.5).'
    )

    def add_footnote(fig):
        pass  # metadata footnote removed

    # Year colormap used by new figures
    all_years = sorted(stations['CHJ']['year'].unique())
    yr_cmap = cm.coolwarm
    yr_norm = mcolors.Normalize(vmin=min(all_years), vmax=max(all_years))

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 1 — Annual July and August means, all stations (split by month)
    # ─────────────────────────────────────────────────────────────────────────
    print("\n  Figure 1 …")
    # Determine the common year range across all stations
    all_min_yr = min(stations[s]['year'].min() for s in STATION_ORDER)
    all_max_yr = max(stations[s]['year'].max() for s in STATION_ORDER)
    chj_min_yr = stations['CHJ']['year'].min()
    chj_max_yr = stations['CHJ']['year'].max()
    chj_n_years = len(ann_summer(stations['CHJ']))
    air_min_yr  = air['year'].min()
    air_max_yr  = air['year'].max()

    fig, axes = plt.subplots(4, 2, figsize=(24, 21), sharex=True, sharey='row')
    fig.suptitle(
        'Annual Mean Water Temperature by Station — July and August Separate\n'
        f'Columbia River: Grand Coulee → Chief Joseph Dam  ({all_min_yr}–{all_max_yr})',
        fontsize=20, fontweight='bold', y=0.99
    )
    for i, stn in enumerate(STATION_ORDER):
        for j, (month, month_name) in enumerate([(7, 'July'), (8, 'August')]):
            ax = axes[i][j]
            ann = ann_month(stations[stn], month)
            x, y = ann.index.values.astype(float), ann.values
            if len(y) < 4:
                ax.set_visible(False)
                continue
            sl, ic, *_ = theil_sen(x, y)
            tau, p, S, Z, direction = mann_kendall(y)

            ax.fill_between(x, y, alpha=0.10, color=C[stn], zorder=1)
            ax.plot(x, y, color=C[stn], lw=2, marker='o', ms=4, zorder=4,
                    label=f'Annual {month_name} Mean')
            ax.plot(x, sl * x + ic, '--', color=C['trend'], lw=1.8, zorder=5,
                    label='Theil-Sen Trend')
            ax.set_xlim(all_min_yr - 1, all_max_yr + 1)
            ax.yaxis.set_minor_locator(MultipleLocator(0.5))
            sig_mark = '★' if p < 0.05 else ''
            ax.set_title(
                f'{STATION_LABELS[stn]} — {month_name}\n'
                f'TS: {sl:+.4f} °C/yr  |  MK τ={tau:.3f}, p={p:.3f} {sig_mark}',
                fontsize=16, color=C[stn], fontweight='bold', loc='left', pad=4
            )
            if j == 0:
                ax.set_ylabel('Water Temperature (°C)', fontsize=20)
            ax.legend(loc='lower left', fontsize=16, frameon=True,
                      edgecolor='#cccccc', framealpha=0.93)
    axes[-1][0].set_xlabel('Year', fontsize=20)
    axes[-1][1].set_xlabel('Year', fontsize=20)
    plt.tight_layout(rect=[0, 0.0, 1, 0.97])
    plt.savefig(f'{OUTPUT_DIR}/Fig1_Annual_Summer_Trends.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig1_Annual_Summer_Trends.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 2 — Mean annual temperature cycle (seasonal climatology)
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 2 …")
    fig, ax = plt.subplots(figsize=(18, 9))
    for stn in STATION_ORDER:
        clim = (stations[stn].groupby('doy')['wtc'].mean()
                .rolling(7, center=True, min_periods=1).mean())
        ax.plot(clim.index, clim.values, color=C[stn], lw=2.2, label=STATION_LABELS[stn])
    air_clim = (air.groupby('doy')['tavg_c'].mean()
                .rolling(7, center=True, min_periods=1).mean())
    ax.plot(air_clim.index, air_clim.values, color=C['air'], lw=1.5, ls=':',
            label='Air Temperature (CJD Area)')
    ax.axvspan(182, 243, alpha=0.09, color='gold', zorder=0, label='July–August Analysis Window')
    ax.text(212, 0.5, 'Jul–Aug\nwindow', ha='center', va='bottom',
            fontsize=16, color='#b7950b', style='italic')
    month_starts = [1, 32, 60, 91, 121, 152, 182, 213, 244, 274, 305, 335]
    month_names  = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    ax.set_xticks(month_starts)
    ax.set_xticklabels(month_names)
    ax.set_xlim(1, 365)
    ax.set_ylim(-1, 28)
    ax.set_xlabel('Month', fontsize=20)
    ax.set_ylabel('Temperature (°C)', fontsize=20)
    ax.set_title(f'Mean Annual Temperature Cycle — 7-Day Rolling Average ({all_min_yr}–{all_max_yr})',
                 fontsize=20, fontweight='bold')
    ax.legend(loc='upper left', ncol=1,
              fontsize=16, frameon=True, edgecolor='#cccccc')
    plt.tight_layout()
    plt.savefig(f'{OUTPUT_DIR}/Fig2_Seasonal_Climatology.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig2_Seasonal_Climatology.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 3 — CHJ detailed trend: July and August separate + air temp
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 3 …")
    fig = plt.figure(figsize=(18, 39))
    gs  = gridspec.GridSpec(5, 1, height_ratios=[3, 3, 3, 3, 2.5], hspace=0.30)
    fig.suptitle(
        'Chief Joseph Dam Forebay (CHJ) — July and August Temperature Trends\n'
        'Mann-Kendall Trend Test & Theil-Sen Slope Estimator',
        fontsize=20, fontweight='bold'
    )

    # Panel 1 — July water temp trend
    ax_jw = fig.add_subplot(gs[0])
    np.random.seed(42)
    boot_jw = []
    for _ in range(1500):
        idx = np.random.choice(len(x_cj), len(x_cj), replace=True)
        s_b, i_b, *_ = theil_sen(x_cj[idx], y_cj[idx])
        boot_jw.append(s_b * x_cj + i_b)
    boot_jw = np.array(boot_jw)
    ax_jw.fill_between(x_cj, np.percentile(boot_jw, 2.5, axis=0),
                       np.percentile(boot_jw, 97.5, axis=0),
                       alpha=0.15, color=C['CHJ'], label='95% Bootstrap CI')
    ax_jw.bar(x_cj, y_cj, width=0.65, color=C['CHJ'], alpha=0.45, zorder=2)
    ax_jw.plot(x_cj, y_cj, 'o-', color=C['CHJ'], lw=1.8, ms=5, zorder=4,
               label='Annual July Water Temp')
    ax_jw.plot(x_cj, sl_cj * x_cj + ic_cj, '--', color=C['trend'], lw=2.2, zorder=5,
               label=f'Theil-Sen: {sl_cj:+.4f} °C/yr  ({sl_cj*10:+.3f} °C/decade)')
    ax_jw.set_ylabel('Water Temperature (°C)', fontsize=20)
    ax_jw.set_xlabel('Year', fontsize=20)
    ax_jw.set_title('July Water Temperature Trend', fontsize=20, fontweight='bold', color=C['CHJ'])
    mk_text_j = (f"MK: τ={tau_cj:.3f}, p={p_cj:.3f}\n"
                 f"{'★ Significant' if p_cj < 0.05 else 'Not significant'}")
    ax_jw.legend(loc='lower left', fontsize=16, frameon=True, edgecolor='#cccccc')

    # Panel 2 — August water temp trend
    ax_aw = fig.add_subplot(gs[1])
    np.random.seed(43)
    boot_aw = []
    for _ in range(1500):
        idx = np.random.choice(len(x_ca), len(x_ca), replace=True)
        s_b, i_b, *_ = theil_sen(x_ca[idx], y_ca[idx])
        boot_aw.append(s_b * x_ca + i_b)
    boot_aw = np.array(boot_aw)
    ax_aw.fill_between(x_ca, np.percentile(boot_aw, 2.5, axis=0),
                       np.percentile(boot_aw, 97.5, axis=0),
                       alpha=0.15, color='#d35400', label='95% Bootstrap CI')
    ax_aw.bar(x_ca, y_ca, width=0.65, color='#d35400', alpha=0.45, zorder=2)
    ax_aw.plot(x_ca, y_ca, 'o-', color='#d35400', lw=1.8, ms=5, zorder=4,
               label='Annual August Water Temp')
    ax_aw.plot(x_ca, sl_ca * x_ca + ic_ca, '--', color=C['trend'], lw=2.2, zorder=5,
               label=f'Theil-Sen: {sl_ca:+.4f} °C/yr  ({sl_ca*10:+.3f} °C/decade)')
    ax_aw.set_ylabel('Water Temperature (°C)', fontsize=20)
    ax_aw.set_xlabel('Year', fontsize=20)
    ax_aw.set_title('August Water Temperature Trend', fontsize=20, fontweight='bold', color='#d35400')
    mk_text_a = (f"MK: τ={tau_ca:.3f}, p={p_ca:.3f}\n"
                 f"{'★ Significant' if p_ca < 0.05 else 'Not significant'}")
    ax_aw.legend(loc='lower left', fontsize=16, frameon=True, edgecolor='#cccccc')

    # Panel 3 — July Mohseni predicted
    ax_jp = fig.add_subplot(gs[2])
    jja_merged = merged[merged['week'].between(26, 35)].copy()
    jja_merged['tw_pred'] = mohseni(jja_merged['tavg_c'].values, *popt_all)
    jul_merged = jja_merged[jja_merged['week'].between(26, 30)]
    ann_obs_j = jul_merged.groupby('year')['wtc'].mean()
    ann_pred_j = jul_merged.groupby('year')['tw_pred'].mean()
    xj = ann_obs_j.index.values.astype(float)
    ax_jp.bar(xj - 0.18, ann_obs_j.values, width=0.36, color='#2980b9', alpha=0.65,
              zorder=2, label='Observed July Mean')
    ax_jp.bar(xj + 0.18, ann_pred_j.values, width=0.36, color='#e74c3c', alpha=0.55,
              zorder=2, label='Mohseni Predicted July Mean')
    mae_j = np.mean(np.abs(ann_obs_j.values - ann_pred_j.values))
    ax_jp.text(0.98, 0.03, f'MAE: {mae_j:.3f}°C', transform=ax_jp.transAxes,
               fontsize=16, va='bottom', ha='right', family='monospace',
               bbox=dict(boxstyle='round,pad=0.2', facecolor='#fdfefe', edgecolor='#bdc3c7', alpha=0.9))
    ax_jp.set_ylabel('Water Temperature (°C)', fontsize=20)
    ax_jp.set_xlabel('Year', fontsize=20)
    ax_jp.set_title('July — Observed vs. Mohseni Predicted', fontsize=20, fontweight='bold')
    ax_jp.legend(loc='lower left', fontsize=16, frameon=True, edgecolor='#cccccc')

    # Panel 4 — August Mohseni predicted
    ax_ap = fig.add_subplot(gs[3])
    aug_merged = jja_merged[jja_merged['week'].between(31, 35)]
    ann_obs_a = aug_merged.groupby('year')['wtc'].mean()
    ann_pred_a = aug_merged.groupby('year')['tw_pred'].mean()
    xa = ann_obs_a.index.values.astype(float)
    ax_ap.bar(xa - 0.18, ann_obs_a.values, width=0.36, color='#2980b9', alpha=0.65,
              zorder=2, label='Observed August Mean')
    ax_ap.bar(xa + 0.18, ann_pred_a.values, width=0.36, color='#e74c3c', alpha=0.55,
              zorder=2, label='Mohseni Predicted August Mean')
    mae_a = np.mean(np.abs(ann_obs_a.values - ann_pred_a.values))
    ax_ap.text(0.98, 0.03, f'MAE: {mae_a:.3f}°C', transform=ax_ap.transAxes,
               fontsize=16, va='bottom', ha='right', family='monospace',
               bbox=dict(boxstyle='round,pad=0.2', facecolor='#fdfefe', edgecolor='#bdc3c7', alpha=0.9))
    ax_ap.set_ylabel('Water Temperature (°C)', fontsize=20)
    ax_ap.set_xlabel('Year', fontsize=20)
    ax_ap.set_title('August — Observed vs. Mohseni Predicted', fontsize=20, fontweight='bold')
    ax_ap.legend(loc='lower left', fontsize=16, frameon=True, edgecolor='#cccccc')

    # Panel 5 — Air temperature — July and August
    ax_air = fig.add_subplot(gs[4])
    ax_air.bar(x_aj - 0.18, y_aj, width=0.36, color='#636e72', alpha=0.4, zorder=2,
               label=f'July Air  (TS: {sl_aj:+.4f} °C/yr, p={p_aj:.3f}{"  ★" if p_aj < 0.05 else ""})')
    ax_air.bar(x_aa + 0.18, y_aa, width=0.36, color='#b2bec3', alpha=0.5, zorder=2,
               label=f'August Air  (TS: {sl_aa:+.4f} °C/yr, p={p_aa:.3f}{"  ★" if p_aa < 0.05 else ""})')
    ax_air.plot(x_aj, sl_aj * x_aj + ic_aj, '--', color='#2c3e50', lw=1.5, zorder=5)
    ax_air.plot(x_aa, sl_aa * x_aa + ic_aa, '-.', color='#636e72', lw=1.5, zorder=5)
    ax_air.set_ylabel('Air Temperature (°C)', fontsize=20)
    ax_air.set_xlabel('Year', fontsize=20)
    ax_air.set_title('Air Temperature Trend — July and August', fontsize=20, fontweight='bold')
    ax_air.set_xlim(min(x_cj.min(), x_aj.min()) - 1, max(x_ca.max(), x_aa.max()) + 1)
    ax_air.legend(loc='lower left', fontsize=16, frameon=True, edgecolor='#cccccc')

    add_footnote(fig)
    plt.savefig(f'{OUTPUT_DIR}/Fig3_CHJ_Trend_Detail.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig3_CHJ_Trend_Detail.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 4 — Daily Jul–Aug water temperature, all years overlaid
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 4 …")
    chj = stations['CHJ']
    summer = chj[chj['month'].isin([7, 8])].copy()
    summer['jul1'] = pd.to_datetime(summer['year'].astype(str) + '-07-01')
    summer['day_of_summer'] = (summer['datetime'] - summer['jul1']).dt.total_seconds() / 86400.0

    fig, ax = plt.subplots(figsize=(21, 10.5))
    for yr in all_years:
        yr_data = summer[summer['year'] == yr]
        daily = yr_data.groupby(yr_data['day_of_summer'].astype(int))['wtc'].mean()
        ax.plot(daily.index, daily.values, color=yr_cmap(yr_norm(yr)), lw=1.0, alpha=0.55, zorder=2)

    overall_daily = summer.groupby(summer['day_of_summer'].astype(int))['wtc'].mean()
    ax.plot(overall_daily.index, overall_daily.values,
            color='black', lw=3, alpha=0.9, zorder=5, label=f'{chj_n_years}-Year Average Daily Cycle')

    overall_mean = summer['wtc'].mean()

    ax.set_xlabel('Day of Summer', fontsize=20)
    ax.set_ylabel('Water Temperature (°C)', fontsize=20)
    ax.set_title('Chief Joseph Dam Forebay (CHJ) — Daily Mean Water Temperature\n'
                 f'July–August, Every Year {chj_min_yr}–{chj_max_yr}', fontsize=20, fontweight='bold')
    ax.set_xlim(-1, 62)
    ax.set_xticks([0, 10, 20, 31, 41, 51, 61])
    ax.set_xticklabels(['Jul 1', 'Jul 11', 'Jul 21', 'Aug 1', 'Aug 11', 'Aug 21', 'Aug 31'])

    sm = cm.ScalarMappable(cmap=yr_cmap, norm=yr_norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax, pad=0.01, fraction=0.03, aspect=30)
    cbar.set_label('Year', fontsize=20)
    cbar.set_ticks(list(range(all_min_yr, all_max_yr+1, 5)))
    # Add a legend entry for individual year lines
    year_line = Line2D([0], [0], color='gray', lw=1.0, alpha=0.55,
                       label='Individual Year Daily Mean (colored by year →)')
    handles, labels = ax.get_legend_handles_labels()
    handles.insert(0, year_line)
    ax.legend(handles=handles, loc='lower right', fontsize=16, frameon=True, edgecolor='#cccccc')
    add_footnote(fig)
    plt.tight_layout(rect=[0, 0.03, 1, 1])
    plt.savefig(f'{OUTPUT_DIR}/Fig4_JulAug_Daily_By_Year.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig4_JulAug_Daily_By_Year.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 5 — Mohseni air-water regression scatter (by season)
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 5 …")
    SEASONS = {
        'Winter (Dec–Feb)': (merged['week'].isin(range(49, 53)) | merged['week'] <= 8,  '#3498db'),
        'Spring (Mar–May)': (merged['week'].between(9, 21),   '#27ae60'),
        'Summer (Jun–Aug)': (merged['week'].between(22, 35),  '#e74c3c'),
        'Fall (Sep–Nov)':   (merged['week'].between(36, 48),  '#e67e22'),
    }
    fig, ax = plt.subplots(figsize=(15, 10.5))
    for label, (mask, color) in SEASONS.items():
        sub = merged[mask]
        ax.scatter(sub['tavg_c'], sub['wtc'], s=8, alpha=0.28, color=color, label=label, zorder=2)
    Ta_r = np.linspace(merged['tavg_c'].min() - 2, merged['tavg_c'].max() + 2, 400)
    ax.plot(Ta_r, mohseni(Ta_r, *popt_all), 'k-', lw=2.5, zorder=6,
            label=f'Mohseni fit — all weeks  (NSE = {NSE_all:.3f})')
    ax.plot(Ta_r, mohseni(Ta_r, *popt_jja), '--', color='darkred', lw=2, zorder=7,
            label=f'Mohseni fit — Jul–Aug  (NSE = {NSE_jja:.3f})')
    ax.axhline(popt_all[1], color='#555', lw=0.9, ls=':', zorder=1,
               label=f'Thermal Ceiling (α = {popt_all[1]:.1f}°C)')
    ax.axhline(popt_all[0], color='#555', lw=0.9, ls=':', zorder=1,
               label=f'Estimated Minimum (μ = {popt_all[0]:.1f}°C)')
    ax.set_xlabel('Weekly Mean Air Temperature (°C)', fontsize=20, labelpad=12)
    ax.set_ylabel('Weekly Mean Water Temperature (°C)', fontsize=20)
    ax.set_title('Mohseni Air–Water Temperature Regression — CHJ (CJD Forebay)\n'
                 f'Weekly Averages, {chj_min_yr}–{chj_max_yr}', fontsize=20, fontweight='bold')
    ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.13), ncol=4,
              fontsize=16, frameon=True, edgecolor='#cccccc')
    plt.tight_layout(rect=[0, 0.07, 1, 1.0])
    plt.savefig(f'{OUTPUT_DIR}/Fig5_Mohseni_Regression.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig5_Mohseni_Regression.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 6 — Observed vs. Mohseni-predicted scatter
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 6 …")
    merged_pred = merged.copy()
    merged_pred['tw_pred'] = mohseni(merged_pred['tavg_c'].values, *popt_all)
    ss_res = np.sum((merged_pred['wtc'] - merged_pred['tw_pred'])**2)
    ss_tot = np.sum((merged_pred['wtc'] - merged_pred['wtc'].mean())**2)
    r2 = 1 - ss_res / ss_tot

    fig, ax = plt.subplots(figsize=(15, 13.5))
    sc = ax.scatter(merged_pred['tw_pred'], merged_pred['wtc'],
                    c=merged_pred['year'], cmap=yr_cmap, norm=yr_norm,
                    s=14, alpha=0.4, edgecolors='none', zorder=3,
                    label=f'Weekly Data Points (n = {len(merged_pred):,}, colored by year →)')
    lo = min(merged_pred['wtc'].min(), merged_pred['tw_pred'].min()) - 1
    hi = max(merged_pred['wtc'].max(), merged_pred['tw_pred'].max()) + 1
    ax.plot([lo, hi], [lo, hi], 'k--', lw=1.5, alpha=0.7, zorder=5, label='1:1 Line (perfect prediction)')
    ax.axhline(popt_all[1], color='#c0392b', lw=1, ls=':', alpha=0.6, zorder=2,
               label=f'Thermal Ceiling — Observed Axis (α = {popt_all[1]:.1f}°C)')
    ax.axvline(popt_all[1], color='#c0392b', lw=1, ls=':', alpha=0.6, zorder=2,
               label=f'Thermal Ceiling — Predicted Axis (α = {popt_all[1]:.1f}°C)')
    ax.set_xlim(lo, hi); ax.set_ylim(lo, hi)
    ax.set_xlabel('Mohseni Predicted Water Temperature (°C)', fontsize=20)
    ax.set_ylabel('Observed Water Temperature (°C)', fontsize=20)
    ax.set_title('Mohseni Model — Observed vs. Predicted Weekly Water Temperature\n'
                 f'CHJ (CJD Forebay), All Weeks {chj_min_yr}–{chj_max_yr}', fontsize=20, fontweight='bold')
    stats_text = f'NSE = {NSE_all:.3f}    R² = {r2:.3f}'
    ax.text(0.03, 0.97, stats_text, transform=ax.transAxes, fontsize=16, va='top', ha='left',
            family='monospace', bbox=dict(boxstyle='round,pad=0.4', facecolor='#fdfefe',
                                          edgecolor='#bdc3c7', alpha=0.9))
    ax.legend(loc='lower right', fontsize=16, frameon=True, edgecolor='#cccccc')
    cbar = plt.colorbar(sc, ax=ax, pad=0.01, fraction=0.03, aspect=25)
    cbar.set_label('Year', fontsize=20)
    add_footnote(fig)
    plt.tight_layout(rect=[0, 0.03, 1, 1])
    plt.savefig(f'{OUTPUT_DIR}/Fig6_Mohseni_Observed_vs_Predicted.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig6_Mohseni_Observed_vs_Predicted.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 7 — July and August annual observed vs predicted (side by side)
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 7 …")
    jja = merged_pred[merged_pred['week'].between(26, 35)].copy()

    fig, (ax_j, ax_a) = plt.subplots(1, 2, figsize=(24, 10.5), sharey=True)
    fig.suptitle(
        'Monthly Mean Water Temperature — Observed vs. Mohseni Predicted\n'
        f'CHJ (CJD Forebay), {chj_min_yr}–{chj_max_yr}',
        fontsize=20, fontweight='bold'
    )

    # July (weeks ~26-30)
    jul_data = jja[jja['week'].between(26, 30)]
    ann_obs_j7 = jul_data.groupby('year')['wtc'].mean()
    ann_pred_j7 = jul_data.groupby('year')['tw_pred'].mean()
    xj7 = ann_obs_j7.index.values.astype(float)
    ax_j.bar(xj7 - 0.18, ann_obs_j7.values, width=0.36, color='#2980b9', alpha=0.65,
             zorder=2, label='Observed July Mean')
    ax_j.bar(xj7 + 0.18, ann_pred_j7.values, width=0.36, color='#e74c3c', alpha=0.55,
             zorder=2, label='Mohseni Predicted')
    mae_j7 = np.mean(np.abs(ann_obs_j7.values - ann_pred_j7.values))
    ax_j.text(0.98, 0.03, f'MAE: {mae_j7:.3f}°C', transform=ax_j.transAxes,
              fontsize=16, va='bottom', ha='right', family='monospace',
              bbox=dict(boxstyle='round,pad=0.3', facecolor='#fdfefe', edgecolor='#bdc3c7', alpha=0.9))
    ax_j.set_xlabel('Year', fontsize=20)
    ax_j.set_ylabel('Water Temperature (°C)', fontsize=20)
    ax_j.set_title('July', fontsize=20, fontweight='bold')
    ax_j.set_xlim(chj_min_yr - 1.5, chj_max_yr + 1.5)
    ax_j.legend(loc='lower left', fontsize=16, frameon=True, edgecolor='#cccccc')

    # August (weeks ~31-35)
    aug_data = jja[jja['week'].between(31, 35)]
    ann_obs_a7 = aug_data.groupby('year')['wtc'].mean()
    ann_pred_a7 = aug_data.groupby('year')['tw_pred'].mean()
    xa7 = ann_obs_a7.index.values.astype(float)
    ax_a.bar(xa7 - 0.18, ann_obs_a7.values, width=0.36, color='#2980b9', alpha=0.65,
             zorder=2, label='Observed August Mean')
    ax_a.bar(xa7 + 0.18, ann_pred_a7.values, width=0.36, color='#e74c3c', alpha=0.55,
             zorder=2, label='Mohseni Predicted')
    mae_a7 = np.mean(np.abs(ann_obs_a7.values - ann_pred_a7.values))
    ax_a.text(0.98, 0.03, f'MAE: {mae_a7:.3f}°C', transform=ax_a.transAxes,
              fontsize=16, va='bottom', ha='right', family='monospace',
              bbox=dict(boxstyle='round,pad=0.3', facecolor='#fdfefe', edgecolor='#bdc3c7', alpha=0.9))
    ax_a.set_xlabel('Year', fontsize=20)
    ax_a.set_title('August', fontsize=20, fontweight='bold')
    ax_a.set_xlim(chj_min_yr - 1.5, chj_max_yr + 1.5)
    ax_a.legend(loc='lower left', fontsize=16, frameon=True, edgecolor='#cccccc')

    add_footnote(fig)
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.savefig(f'{OUTPUT_DIR}/Fig7_JulAug_Obs_vs_Pred_TimeSeries.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig7_JulAug_Obs_vs_Pred_TimeSeries.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 8 — All years overlaid on Mohseni S-curve
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 8 …")
    fig, ax = plt.subplots(figsize=(21, 13.5))
    for yr in all_years:
        yr_data = merged[merged['year'] == yr].sort_values('tavg_c')
        ax.plot(yr_data['tavg_c'], yr_data['wtc'],
                'o-', color=yr_cmap(yr_norm(yr)), ms=3.5, lw=0.8, alpha=0.45, zorder=2)
    ta_range = np.linspace(merged['tavg_c'].min() - 2, merged['tavg_c'].max() + 2, 300)
    ax.plot(ta_range, mohseni(ta_range, *popt_all), 'k-', lw=3, alpha=0.9, zorder=5,
            label=f'Mohseni Fitted Curve (NSE = {NSE_all:.3f})')
    ax.axhline(popt_all[1], color='#c0392b', lw=1.5, ls=':', alpha=0.7, zorder=4,
               label=f'Thermal Ceiling (α = {popt_all[1]:.1f}°C)')
    ax.axhline(popt_all[0], color='#2980b9', lw=1.5, ls=':', alpha=0.7, zorder=4,
               label=f'Estimated Minimum (μ = {popt_all[0]:.1f}°C)')
    ax.set_xlabel('Weekly Mean Air Temperature (°C)', fontsize=20)
    ax.set_ylabel('Weekly Mean Water Temperature (°C)', fontsize=20)
    ax.set_title(f'Mohseni Air–Water Temperature Relationship — All {chj_n_years} Years Overlaid\n'
                 f'CHJ (CJD Forebay), Weekly Averages {chj_min_yr}–{chj_max_yr}', fontsize=20, fontweight='bold')
    sm = cm.ScalarMappable(cmap=yr_cmap, norm=yr_norm)
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax, pad=0.01, fraction=0.03, aspect=30)
    cbar.set_label('Year', fontsize=20)
    cbar.set_ticks(list(range(all_min_yr, all_max_yr+1, 5)))
    # Add legend entry for individual year lines
    year_line = Line2D([0], [0], color='gray', lw=0.8, marker='o', ms=3.5, alpha=0.45,
                       label='Individual Year Weekly Data (colored by year →)')
    handles, labels = ax.get_legend_handles_labels()
    handles.insert(0, year_line)
    ax.legend(handles=handles, loc='lower right', fontsize=16, frameon=True, edgecolor='#cccccc')
    add_footnote(fig)
    plt.tight_layout(rect=[0, 0.03, 1, 1])
    plt.savefig(f'{OUTPUT_DIR}/Fig8_Mohseni_All_Years_Overlaid.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig8_Mohseni_All_Years_Overlaid.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 9 — Projected temperatures bar chart (July and August separate)
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 9 …")
    monthly_proj = results.get('monthly_proj', {})
    if monthly_proj:
        jul_df_9 = monthly_proj['July']
        aug_df_9 = monthly_proj['August']

        # All rows including baseline
        scenario_labels_9 = jul_df_9['Scenario'].values
        n9 = len(scenario_labels_9)
        x9 = np.arange(n9)
        bar_w9 = 0.35

        fig, ax = plt.subplots(figsize=(18, 10.5))
        bars_j9 = ax.bar(x9 - bar_w9/2, jul_df_9['Proj_Tw_C'].values, bar_w9,
                         label='July', color='#e74c3c', alpha=0.85, edgecolor='white', linewidth=0.5)
        bars_a9 = ax.bar(x9 + bar_w9/2, aug_df_9['Proj_Tw_C'].values, bar_w9,
                         label='August', color='#2980b9', alpha=0.85, edgecolor='white', linewidth=0.5)

        for bars in [bars_j9, bars_a9]:
            for bar in bars:
                h = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2, h + 0.02,
                        f'{h:.2f}°C\n({h*9/5+32:.1f}°F)',
                        ha='center', va='bottom', fontsize=16, fontweight='bold')

        ax.set_xticks(x9)
        ax.set_xticklabels(scenario_labels_9, fontsize=16)
        ax.set_xlabel('Climate Scenario', fontsize=20)
        ax.set_ylabel('Projected Water Temperature (°C)', fontsize=20)
        ax.set_title('Projected July and August Water Temperature at CJD Forebay (CHJ)\n'
                     'Mohseni (1998) Delta Method + CMIP5 Climate Scenarios (RCP 4.5 / 8.5)',
                     fontsize=20, fontweight='bold')
        all_vals = list(jul_df_9['Proj_Tw_C'].values) + list(aug_df_9['Proj_Tw_C'].values)
        ax.set_ylim(min(all_vals) - 1.0, max(all_vals) + 1.5)
        ax.legend(fontsize=20, loc='upper left')
        ax.grid(axis='y', alpha=0.3)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.03, 1, 1])
        plt.savefig(f'{OUTPUT_DIR}/Fig9_Projected_Temps_Bar.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig9_Projected_Temps_Bar.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 10 — Projection points on Mohseni curve (July and August separate)
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 10 …")
    if monthly_proj:
        jul_df_10 = monthly_proj['July']
        aug_df_10 = monthly_proj['August']

        fig, ax = plt.subplots(figsize=(15, 10.5))
        ta_range = np.linspace(-10, 40, 500)
        ax.plot(ta_range, mohseni(ta_range, *popt_all), 'k-', lw=2.5, alpha=0.9, zorder=3,
                label='Mohseni Fitted Curve')
        ax.axhline(popt_all[1], color='#c0392b', lw=1.5, ls=':', alpha=0.7, zorder=2,
                   label=f'Thermal Ceiling (α = {popt_all[1]:.1f}°C)')
        ax.axhline(popt_all[0], color='#2980b9', lw=1.5, ls=':', alpha=0.7, zorder=2,
                   label=f'Estimated Minimum (μ = {popt_all[0]:.1f}°C)')

        scenario_colors = {'Baseline': '#2980b9', '2040s RCP 4.5': '#f57c00',
                          '2040s RCP 8.5': '#d32f2f', '2080s RCP 4.5': '#e65100',
                          '2080s RCP 8.5': '#b71c1c'}

        for month_name, mdf, marker in [('July', jul_df_10, 'o'), ('August', aug_df_10, 's')]:
            for _, row in mdf.iterrows():
                # Plot on Mohseni curve (model-predicted, not observed baseline)
                tw_on_curve = float(mohseni(row['Proj_Air_C'], *popt_all))
                color = scenario_colors.get(row['Scenario'].split('(')[0].strip(),
                                            row.get('Color', '#333'))
                label_str = f"{month_name}: {row['Scenario']} ({row['Proj_Tw_C']:.1f}°C)"
                ax.plot(row['Proj_Air_C'], tw_on_curve, marker, color=color,
                        ms=12, zorder=5, markeredgecolor='white', markeredgewidth=1.5,
                        label=label_str)

        ax.set_xlabel('Air Temperature (°C)', fontsize=20)
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.set_title('Climate Projection Points on the Mohseni Air–Water Curve\n'
                     'CHJ (CJD Forebay) — July (○) and August (□) — CMIP5 RCP 4.5 / 8.5',
                     fontsize=20, fontweight='bold')
        ax.set_xlim(15, 35); ax.set_ylim(15, 20)
        ax.legend(loc='lower right', fontsize=16, frameon=True, edgecolor='#cccccc', ncol=2)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.03, 1, 1])
        plt.savefig(f'{OUTPUT_DIR}/Fig10_Projection_Points_Mohseni.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig10_Projection_Points_Mohseni.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 11 — Period comparison boxplots: early vs recent
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 11 …")
    # Dynamic period split at midpoint of the common record
    mid_year = (all_min_yr + all_max_yr) // 2
    early_label = f'Early Record ({all_min_yr}–{mid_year})'
    recent_label = f'Recent Record ({mid_year+1}–{all_max_yr})'
    fig, axes = plt.subplots(2, 2, figsize=(19.5, 13.5))
    fig.suptitle(
        f'Water Temperature Distribution by Month — Early ({all_min_yr}–{mid_year}) vs. Recent ({mid_year+1}–{all_max_yr})\n'
        'All Stations  |  Wilcoxon p-values shown for July & August',
        fontsize=20, fontweight='bold'
    )
    for idx, stn in enumerate(STATION_ORDER):
        ax  = axes[idx // 2][idx % 2]
        df  = stations[stn][stations[stn]['month'].between(4, 10)].copy()
        early  = df[df['year'] <= mid_year]
        recent = df[df['year'] >= mid_year + 1]
        months = list(range(4, 11))
        mlabels = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct']

        def bp_kw(positions, color):
            return dict(positions=positions, widths=0.38, patch_artist=True,
                        medianprops=dict(color='white', lw=2),
                        whiskerprops=dict(color=color, lw=1.2),
                        capprops=dict(color=color),
                        boxprops=dict(facecolor=color, alpha=0.65),
                        flierprops=dict(marker='.', ms=2, alpha=0.25, color=color))

        ax.boxplot([early[early['month']  == m]['wtc'].dropna() for m in months],
                   **bp_kw([m - 0.22 for m in months], '#2980b9'))
        ax.boxplot([recent[recent['month'] == m]['wtc'].dropna() for m in months],
                   **bp_kw([m + 0.22 for m in months], '#c0392b'))

        ax.set_xticks(months)
        ax.set_xticklabels(mlabels)
        ax.set_title(STATION_LABELS[stn], fontsize=20, fontweight='bold',
                     color=C[stn], pad=6)
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.axvspan(6.6, 8.4, alpha=0.07, color='gold', zorder=0)
        ymax = ax.get_ylim()[1]
        for m in [7, 8]:
            e = early[early['month']  == m]['wtc'].dropna().values
            r = recent[recent['month'] == m]['wtc'].dropna().values
            if len(e) > 5 and len(r) > 5:
                _, wp = stats.mannwhitneyu(e, r, alternative='two-sided')
                ax.text(m, ymax * 0.98, f'{"★" if wp < 0.05 else ""}p={wp:.3f}',
                        ha='center', va='top', fontsize=16,
                        color='darkred' if wp < 0.05 else '#888', fontweight='bold')

    axes[1][0].set_xlabel('Month', fontsize=20)
    axes[1][1].set_xlabel('Month', fontsize=20)
    fig.legend(
        handles=[mpatches.Patch(facecolor='#2980b9', alpha=0.65, label=early_label),
                 mpatches.Patch(facecolor='#c0392b', alpha=0.65, label=recent_label),
                 Line2D([0], [0], color='white', lw=2, label='Median (white line in box)'),
                 Line2D([0], [0], color='gray', lw=1.2, label='Whiskers (1.5× IQR)'),
                 Line2D([0], [0], color='gray', marker='.', ms=4, lw=0, alpha=0.4, label='Outliers'),
                 mpatches.Patch(facecolor='gold', alpha=0.15, label='July–August Analysis Window'),
                 Line2D([0], [0], color='white', marker='', label='p = Mann-Whitney U test'),
                 Line2D([0], [0], color='white', marker='', label='★ = significant (p < 0.05)')],
        loc='lower center', ncol=4, fontsize=16,
        bbox_to_anchor=(0.5, -0.01), frameon=True, edgecolor='#cccccc'
    )
    plt.tight_layout(rect=[0, 0.04, 1, 0.97])
    plt.savefig(f'{OUTPUT_DIR}/Fig11_Period_Comparison_Boxplots.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("    ✓ Fig11_Period_Comparison_Boxplots.png")

    # ── Fig 12 — Monthly Projections: Projected Change (ΔTw) by Month ────────
    monthly_proj = results.get('monthly_proj', {})
    if monthly_proj:
        fig, ax = plt.subplots(figsize=(18, 9))

        jul_df = monthly_proj['July']
        aug_df = monthly_proj['August']

        # Exclude baseline row (delta = 0) for the delta chart
        jul_scen = jul_df[jul_df['Delta_Air_C'] > 0].reset_index(drop=True)
        aug_scen = aug_df[aug_df['Delta_Air_C'] > 0].reset_index(drop=True)

        scenario_labels = jul_scen['Scenario'].values
        n = len(scenario_labels)
        x = np.arange(n)
        bar_w = 0.35

        jul_dtw = jul_scen['Delta_Tw_C'].values
        aug_dtw = aug_scen['Delta_Tw_C'].values

        bars_jul = ax.bar(x - bar_w/2, jul_dtw, bar_w, label='July',
                          color='#e74c3c', alpha=0.85, edgecolor='white', linewidth=0.5)
        bars_aug = ax.bar(x + bar_w/2, aug_dtw, bar_w, label='August',
                          color='#2980b9', alpha=0.85, edgecolor='white', linewidth=0.5)

        # Add delta labels on each bar
        for bars in [bars_jul, bars_aug]:
            for bar in bars:
                h = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2, h + 0.01,
                        f'+{h:.2f}°C\n(+{h*9/5:.2f}°F)',
                        ha='center', va='bottom', fontsize=16, fontweight='bold')

        ax.set_xticks(x)
        ax.set_xticklabels(scenario_labels, fontsize=16)
        ax.set_xlabel('Climate Scenario', fontsize=20)
        ax.set_ylabel('Projected Change in Water Temperature (°C)', fontsize=20)
        ax.set_ylim(0, max(max(jul_dtw), max(aug_dtw)) * 1.45)

        # Add observed baselines as annotation
        jul_base = jul_df.loc[jul_df['Delta_Air_C'] == 0, 'Obs_Baseline_Tw_C'].values[0]
        aug_base = aug_df.loc[aug_df['Delta_Air_C'] == 0, 'Obs_Baseline_Tw_C'].values[0]
        ax.text(0.02, 0.97,
                f'Observed Baselines (1997–2025):\n'
                f'  July:    {jul_base:.2f}°C  ({jul_base*9/5+32:.1f}°F)\n'
                f'  August: {aug_base:.2f}°C  ({aug_base*9/5+32:.1f}°F)',
                transform=ax.transAxes, fontsize=16, va='top', family='monospace',
                bbox=dict(boxstyle='round,pad=0.4', facecolor='lightyellow', alpha=0.9))

        ax.set_title('Projected Change in July and August Water Temperature at CJD Forebay (CHJ)\n'
                     'Mohseni (1998) Delta Method + CMIP5 Climate Scenarios (RCP 4.5 / 8.5)',
                     fontsize=20, fontweight='bold')
        ax.legend(fontsize=20, loc='upper right')
        ax.grid(axis='y', alpha=0.3)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.04, 1, 0.97])
        plt.savefig(f'{OUTPUT_DIR}/Fig12_Monthly_Projections_Bar.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig12_Monthly_Projections_Bar.png")

        # ── Fig 13 — Mohseni sensitivity: delta arrows on curve ──────────────
        fig, ax = plt.subplots(figsize=(15, 10.5))
        Ta_r = np.linspace(15, 35, 500)
        ax.plot(Ta_r, mohseni(Ta_r, *popt_all), 'k-', lw=3, label='Mohseni Fitted Curve')
        ax.axhline(y=float(popt_all[1]), color='red', ls=':', alpha=0.6, lw=1,
                   label=f'Thermal Ceiling (α = {popt_all[1]:.1f}°C)')
        ax.axhline(y=float(popt_all[0]), color='blue', ls=':', alpha=0.6, lw=1,
                   label=f'Estimated Minimum (μ = {popt_all[0]:.1f}°C)')

        # Plot baseline points ON the Mohseni curve (model-predicted values)
        for month_name, mdf, marker_style in [('July', jul_df, 'o'), ('August', aug_df, 's')]:
            base_row = mdf[mdf['Delta_Air_C'] == 0].iloc[0]
            base_air = base_row['Baseline_Air_C']
            base_tw_mohseni = float(mohseni(base_air, *popt_all))

            # Plot baseline point on curve
            ax.scatter(base_air, base_tw_mohseni, s=150, marker=marker_style,
                      c='#2980b9', edgecolors='black', linewidth=1, zorder=10,
                      label=f'{month_name} Baseline Air ({base_air:.1f}°C)')

            # For the most extreme scenario, show the delta arrow on the curve
            extreme = mdf[mdf['Delta_Air_C'] > 0].iloc[-1]  # 2080s RCP 8.5
            ext_air = extreme['Proj_Air_C']
            ext_tw_mohseni = float(mohseni(ext_air, *popt_all))
            dtw = ext_tw_mohseni - base_tw_mohseni

            # Arrow from baseline to projected on the Mohseni curve
            ax.annotate('', xy=(ext_air, ext_tw_mohseni),
                       xytext=(base_air, base_tw_mohseni),
                       arrowprops=dict(arrowstyle='->', color='#e74c3c' if month_name == 'July' else '#2980b9',
                                      lw=2.5, connectionstyle='arc3,rad=0.1'))
            # Label the delta
            mid_air = (base_air + ext_air) / 2
            mid_tw = (base_tw_mohseni + ext_tw_mohseni) / 2
            offset_y = 0.6 if month_name == 'July' else -0.8
            ax.text(mid_air, mid_tw + offset_y,
                    f'{month_name}: ΔTw = +{dtw:.2f}°C\n(2080s RCP 8.5)',
                    fontsize=16, ha='center', fontweight='bold',
                    color='#e74c3c' if month_name == 'July' else '#2980b9',
                    bbox=dict(boxstyle='round,pad=0.2', facecolor='white', alpha=0.85))

            # Plot projected point on curve
            ax.scatter(ext_air, ext_tw_mohseni, s=150, marker=marker_style,
                      c='#8b0000', edgecolors='black', linewidth=1, zorder=10)

        ax.text(0.02, 0.98,
                'Arrows show Mohseni-derived ΔTw\n'
                'used in the delta method.\n'
                'Observed baselines differ from\n'
                'curve due to reservoir thermal lag.',
                transform=ax.transAxes, fontsize=16, va='top',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow', alpha=0.9))

        ax.set_xlabel('Air Temperature (°C)', fontsize=20)
        ax.set_ylabel('Mohseni-Predicted Water Temperature (°C)', fontsize=20)
        ax.set_title('Mohseni Model Sensitivity Used for Monthly Climate Projections\n'
                     'CHJ (CJD Forebay) — Delta Method (RCP 4.5 / 8.5)',
                     fontsize=20, fontweight='bold')
        ax.legend(loc='lower right', fontsize=16)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.04, 1, 0.97])
        plt.savefig(f'{OUTPUT_DIR}/Fig13_Monthly_Projection_Points.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig13_Monthly_Projection_Points.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 14 — Observed week-by-week min/max/mean, July through August
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 14 …")
    chj = stations['CHJ']
    week_stats_df = results.get('week_stats_df', None)

    if week_stats_df is not None and len(week_stats_df) > 0:
        wdf = week_stats_df.sort_values('week').reset_index(drop=True)
        n_weeks = len(wdf)
        x = np.arange(n_weeks)
        colors = ['#e74c3c' if m == 7 else '#2980b9' for m in wdf['primary_month']]

        fig, ax = plt.subplots(figsize=(21, 10.5))
        bars = ax.bar(x, wdf['mean_tw'].values, width=0.7, color=colors, alpha=0.85,
                      edgecolor='white', linewidth=0.5, zorder=3)
        err_lo = (wdf['mean_tw'] - wdf['min_tw']).values
        err_hi = (wdf['max_tw'] - wdf['mean_tw']).values
        ax.errorbar(x, wdf['mean_tw'].values, yerr=[err_lo, err_hi], fmt='none',
                    ecolor='#2c3e50', elinewidth=1.8, capsize=6, capthick=1.8, zorder=5)

        # Label min, max on each bar
        for i, (_, r) in enumerate(wdf.iterrows()):
            ax.text(x[i], r['max_tw'] + 0.08, f'{r["max_tw"]:.1f}',
                    ha='center', va='bottom', fontsize=16, fontweight='bold', color='#555')
            ax.text(x[i], r['min_tw'] - 0.08, f'{r["min_tw"]:.1f}',
                    ha='center', va='top', fontsize=16, fontweight='bold', color='#555')

        # Week labels with month context
        week_labels = []
        for _, r in wdf.iterrows():
            m_name = 'Jul' if r['primary_month'] == 7 else 'Aug'
            week_labels.append(f'Wk {int(r["week"])}\n({m_name})')
        ax.set_xticks(x)
        ax.set_xticklabels(week_labels, fontsize=16)
        ax.set_xlabel('Week', fontsize=20)
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.set_title('Observed Weekly Water Temperature at CHJ — July through August (1997–2025)',
                     fontsize=20, fontweight='bold')

        # Legend
        jul_patch = mpatches.Patch(color='#e74c3c', alpha=0.85, label='July weeks')
        aug_patch = mpatches.Patch(color='#2980b9', alpha=0.85, label='August weeks')
        err_line = Line2D([0], [0], color='#2c3e50', lw=1.8, marker='_', ms=10,
                          label='Min/Max range across years')
        ax.legend(handles=[jul_patch, aug_patch, err_line], loc='lower right',
                  fontsize=16, frameon=True, edgecolor='#cccccc')

        ax.grid(axis='y', alpha=0.3)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.04, 1, 1])
        plt.savefig(f'{OUTPUT_DIR}/Fig14_Observed_Weekly_MinMaxMean.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig14_Observed_Weekly_MinMaxMean.png")

    # ─────────────────────────────────────────────────────────────────────────
    # FIGURE 15 — Projected week-by-week under most extreme scenario (2080s RCP 8.5)
    # ─────────────────────────────────────────────────────────────────────────
    print("  Figure 15 …")
    monthly_proj_fig = results.get('monthly_proj', {})
    if week_stats_df is not None and len(week_stats_df) > 0 and monthly_proj_fig:
        wdf = week_stats_df.sort_values('week').reset_index(drop=True)
        n_weeks = len(wdf)
        x = np.arange(n_weeks)
        bar_w = 0.38

        # Get deltas for each month from monthly projections
        jul_proj_df = monthly_proj_fig['July']
        aug_proj_df = monthly_proj_fig['August']
        # Build a dict: scenario -> {7: delta, 8: delta}
        scenarios_to_plot = []
        jul_scen = jul_proj_df[jul_proj_df['Delta_Air_C'] > 0].reset_index(drop=True)
        aug_scen = aug_proj_df[aug_proj_df['Delta_Air_C'] > 0].reset_index(drop=True)
        for i in range(len(jul_scen)):
            scenarios_to_plot.append({
                'label': jul_scen.iloc[i]['Scenario'],
                7: jul_scen.iloc[i]['Delta_Tw_C'],
                8: aug_scen.iloc[i]['Delta_Tw_C'],
            })

        fig, ax = plt.subplots(figsize=(21, 10.5))

        # Observed baseline bars
        colors_obs = ['#e74c3c' if m == 7 else '#2980b9' for m in wdf['primary_month']]
        bars_obs = ax.bar(x - bar_w/2, wdf['mean_tw'].values, bar_w, color=colors_obs, alpha=0.5,
                          edgecolor='white', linewidth=0.5, zorder=3)
        err_lo_obs = (wdf['mean_tw'] - wdf['min_tw']).values
        err_hi_obs = (wdf['max_tw'] - wdf['mean_tw']).values
        ax.errorbar(x - bar_w/2, wdf['mean_tw'].values, yerr=[err_lo_obs, err_hi_obs], fmt='none',
                    ecolor='#888', elinewidth=1.2, capsize=4, capthick=1.2, zorder=5)

        # Use the most extreme scenario (last one = 2080s RCP 8.5) for the projected bars
        extreme = scenarios_to_plot[-1]
        proj_means = []
        proj_mins = []
        proj_maxs = []
        for _, r in wdf.iterrows():
            dtw = extreme[r['primary_month']]
            proj_means.append(r['mean_tw'] + dtw)
            proj_mins.append(r['min_tw'] + dtw)
            proj_maxs.append(r['max_tw'] + dtw)
        proj_means = np.array(proj_means)
        proj_mins = np.array(proj_mins)
        proj_maxs = np.array(proj_maxs)
        colors_proj = ['#c0392b' if m == 7 else '#1a5276' for m in wdf['primary_month']]

        bars_proj = ax.bar(x + bar_w/2, proj_means, bar_w, color=colors_proj, alpha=0.85,
                           edgecolor='white', linewidth=0.5, zorder=3)
        err_lo_proj = proj_means - proj_mins
        err_hi_proj = proj_maxs - proj_means
        ax.errorbar(x + bar_w/2, proj_means, yerr=[err_lo_proj, err_hi_proj], fmt='none',
                    ecolor='#2c3e50', elinewidth=1.8, capsize=4, capthick=1.8, zorder=5)

        # Min/max labels on projected bars
        for i in range(n_weeks):
            ax.text(x[i] + bar_w/2, proj_maxs[i] + 0.08, f'{proj_maxs[i]:.1f}',
                    ha='center', va='bottom', fontsize=16, fontweight='bold', color='#555')
            ax.text(x[i] + bar_w/2, proj_mins[i] - 0.08, f'{proj_mins[i]:.1f}',
                    ha='center', va='top', fontsize=16, fontweight='bold', color='#555')

        # Min/max labels on observed bars
        for i, (_, r) in enumerate(wdf.iterrows()):
            ax.text(x[i] - bar_w/2, r['max_tw'] + 0.08, f'{r["max_tw"]:.1f}',
                    ha='center', va='bottom', fontsize=16, fontweight='bold', color='#999')
            ax.text(x[i] - bar_w/2, r['min_tw'] - 0.08, f'{r["min_tw"]:.1f}',
                    ha='center', va='top', fontsize=16, fontweight='bold', color='#999')

        # Week labels
        week_labels = []
        for _, r in wdf.iterrows():
            m_name = 'Jul' if r['primary_month'] == 7 else 'Aug'
            week_labels.append(f'Wk {int(r["week"])}\n({m_name})')
        ax.set_xticks(x)
        ax.set_xticklabels(week_labels, fontsize=16)
        ax.set_xlabel('Week', fontsize=20)
        ax.set_ylabel('Water Temperature (°C)', fontsize=20)
        ax.set_title('Projected Weekly Water Temperature at CHJ — July through August\n'
                     'Observed vs. 2080s RCP 8.5',
                     fontsize=20, fontweight='bold')

        # Legend
        obs_jul = mpatches.Patch(color='#e74c3c', alpha=0.5, label='Observed July')
        obs_aug = mpatches.Patch(color='#2980b9', alpha=0.5, label='Observed August')
        proj_jul = mpatches.Patch(color='#c0392b', alpha=0.85, label=f'Projected July ({extreme["label"]})')
        proj_aug = mpatches.Patch(color='#1a5276', alpha=0.85, label=f'Projected August ({extreme["label"]})')
        err_line = Line2D([0], [0], color='#2c3e50', lw=1.8, marker='_', ms=10,
                          label='Min/Max range')
        ax.legend(handles=[obs_jul, obs_aug, proj_jul, proj_aug, err_line],
                  loc='lower right', fontsize=16, frameon=True, edgecolor='#cccccc')

        ax.grid(axis='y', alpha=0.3)
        add_footnote(fig)
        plt.tight_layout(rect=[0, 0.04, 1, 1])
        plt.savefig(f'{OUTPUT_DIR}/Fig15_Projected_Weekly_MinMaxMean.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("    ✓ Fig15_Projected_Weekly_MinMaxMean.png")

    print(f"\n  All figures saved to {OUTPUT_DIR}")


# ═════════════════════════════════════════════════════════════════════════════
# SECTION 6 — MAIN ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    print("=" * 70)
    print("CJD Tailrace Water Temperature — Full Analysis")
    print("=" * 70)

    # ── Load data ─────────────────────────────────────────────────────────────
    print("\n[1] Loading data …")
    stations = {s: load_station(s, STATION_FILES[s])
                for s in STATION_ORDER}
    air = load_air(AIR_TEMP_FILE)

    # ── Cap data at analysis end year (exclude any partial-year records) ───
    MAX_YEAR = 2025
    for s in stations:
        before = len(stations[s])
        stations[s] = stations[s][stations[s]['year'] <= MAX_YEAR].reset_index(drop=True)
        dropped = before - len(stations[s])
        if dropped > 0:
            print(f"    [{s}] Dropped {dropped:,} records from {MAX_YEAR+1}+")
    before_air = len(air)
    air = air[air['year'] <= MAX_YEAR].reset_index(drop=True)
    dropped_air = before_air - len(air)
    if dropped_air > 0:
        print(f"    [Air] Dropped {dropped_air:,} records from {MAX_YEAR+1}+")

    # ── Run analysis ──────────────────────────────────────────────────────────
    print("\n[2] Running statistical analysis …")
    results = run_analysis(stations, air)

    # ── Export Excel ──────────────────────────────────────────────────────────
    print("\n[3] Exporting results …")
    export_excel(results)

    # ── Generate figures ──────────────────────────────────────────────────────
    print("\n[4] Generating figures …")
    make_figures(results)

    print("\n" + "=" * 70)
    print("COMPLETE — all outputs saved to:", OUTPUT_DIR)
    print("=" * 70)
